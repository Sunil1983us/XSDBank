"""
PDF Table to Excel Extractor
Extracts all tables from a PDF and writes them to a formatted Excel workbook.
Each table gets its own sheet; a summary sheet lists all tables found.
"""

import os
import re
from pathlib import Path

import pdfplumber
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter


# ── colour palette ──────────────────────────────────────────────────────────
HEADER_FILL  = PatternFill("solid", start_color="1F3864")   # dark navy
HEADER_FONT  = Font(name="Arial", bold=True, color="FFFFFF", size=10)
ALT_FILL     = PatternFill("solid", start_color="EBF0FA")   # light blue
NORMAL_FONT  = Font(name="Arial", size=9)
TITLE_FONT   = Font(name="Arial", bold=True, size=11, color="1F3864")
SUMMARY_FILL = PatternFill("solid", start_color="D9E1F2")

THIN  = Side(style="thin",   color="BFC9E0")
THICK = Side(style="medium", color="1F3864")

def _cell_border(top=THIN, bottom=THIN, left=THIN, right=THIN):
    return Border(top=top, bottom=bottom, left=left, right=right)


def _clean(text):
    """Collapse whitespace; return empty string for None."""
    if text is None:
        return ""
    return re.sub(r"\s+", " ", str(text)).strip()


def _col_widths(ws):
    """Auto-fit column widths (capped at 60)."""
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                val = str(cell.value or "")
                max_len = max(max_len, min(len(val), 60))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = max(max_len + 2, 10)


def extract_tables_to_excel(pdf_path: str, output_path: str | None = None) -> dict:
    """
    Extract all tables from *pdf_path* into a formatted .xlsx file.

    Returns a result dict with:
        output_file  – path of the produced .xlsx
        total_tables – how many tables were extracted
        tables       – list of {page, table_index, rows, cols} per table
        warnings     – list of non-fatal messages
    """
    pdf_path = Path(pdf_path)
    if not pdf_path.exists():
        raise FileNotFoundError(f"PDF not found: {pdf_path}")

    if output_path is None:
        output_path = pdf_path.with_suffix(".xlsx")
    output_path = Path(output_path)

    wb = Workbook()
    summary_ws = wb.active
    summary_ws.title = "Summary"

    all_tables_meta = []
    warnings = []

    with pdfplumber.open(pdf_path) as pdf:
        total_pages = len(pdf.pages)
        sheet_idx = 1

        for page_num, page in enumerate(pdf.pages, start=1):
            try:
                tables = page.extract_tables()
            except Exception as e:
                warnings.append(f"Page {page_num}: extraction failed – {e}")
                continue

            for tbl_idx, raw_table in enumerate(tables, start=1):
                if not raw_table or len(raw_table) < 2:
                    continue

                # Clean every cell
                table = [[_clean(cell) for cell in row] for row in raw_table]

                sheet_name = f"P{page_num}_T{tbl_idx}"
                ws = wb.create_sheet(title=sheet_name)

                # ── page / table header ──────────────────────────────────
                ws.merge_cells(f"A1:{get_column_letter(len(table[0]))}1")
                title_cell = ws["A1"]
                title_cell.value = (
                    f"{pdf_path.name}  │  Page {page_num}  │  Table {tbl_idx}"
                )
                title_cell.font   = TITLE_FONT
                title_cell.fill   = PatternFill("solid", start_color="D9E1F2")
                title_cell.alignment = Alignment(horizontal="left",
                                                  vertical="center",
                                                  wrap_text=False)
                ws.row_dimensions[1].height = 18

                # ── column headers (first data row) ─────────────────────
                header_row = table[0]
                for col_idx, val in enumerate(header_row, start=1):
                    cell = ws.cell(row=2, column=col_idx, value=val)
                    cell.font      = HEADER_FONT
                    cell.fill      = HEADER_FILL
                    cell.alignment = Alignment(horizontal="center",
                                               vertical="center",
                                               wrap_text=True)
                    cell.border    = _cell_border(top=THICK, bottom=THICK)
                ws.row_dimensions[2].height = 22

                # ── data rows ────────────────────────────────────────────
                for row_i, row in enumerate(table[1:], start=3):
                    fill = ALT_FILL if row_i % 2 == 0 else None
                    for col_idx, val in enumerate(row, start=1):
                        cell = ws.cell(row=row_i, column=col_idx, value=val)
                        cell.font      = NORMAL_FONT
                        cell.alignment = Alignment(horizontal="left",
                                                   vertical="top",
                                                   wrap_text=True)
                        cell.border    = _cell_border()
                        if fill:
                            cell.fill = fill
                    ws.row_dimensions[row_i].height = 30

                # Freeze header rows
                ws.freeze_panes = "A3"
                _col_widths(ws)

                meta = {
                    "sheet":       sheet_name,
                    "page":        page_num,
                    "table_index": tbl_idx,
                    "rows":        len(table) - 1,   # excluding header
                    "cols":        len(header_row),
                }
                all_tables_meta.append(meta)
                sheet_idx += 1

    # ── Summary sheet ────────────────────────────────────────────────────────
    _build_summary(summary_ws, pdf_path.name, total_pages, all_tables_meta)

    wb.save(output_path)

    return {
        "output_file":  str(output_path),
        "total_tables": len(all_tables_meta),
        "total_pages":  total_pages,
        "tables":       all_tables_meta,
        "warnings":     warnings,
    }


def _build_summary(ws, pdf_name, total_pages, meta_list):
    ws["A1"] = "PDF Table Extraction – Summary"
    ws["A1"].font = Font(name="Arial", bold=True, size=14, color="1F3864")
    ws.merge_cells("A1:F1")
    ws.row_dimensions[1].height = 24

    ws["A2"] = f"Source: {pdf_name}   │   Total pages: {total_pages}   │   Tables found: {len(meta_list)}"
    ws["A2"].font = Font(name="Arial", italic=True, size=10, color="444444")
    ws.merge_cells("A2:F2")

    headers = ["Sheet", "Page", "Table #", "Data Rows", "Columns", "Go To"]
    for col_i, h in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=col_i, value=h)
        cell.font      = HEADER_FONT
        cell.fill      = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border    = _cell_border(top=THICK, bottom=THICK)
    ws.row_dimensions[4].height = 20

    for row_i, m in enumerate(meta_list, start=5):
        fill = ALT_FILL if row_i % 2 == 0 else None
        data = [m["sheet"], m["page"], m["table_index"], m["rows"], m["cols"],
                f"→ {m['sheet']}"]
        for col_i, val in enumerate(data, start=1):
            cell = ws.cell(row=row_i, column=col_i, value=val)
            cell.font      = Font(name="Arial", size=9)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border    = _cell_border()
            if fill:
                cell.fill = fill
        ws.row_dimensions[row_i].height = 16

    for col_letter, width in zip("ABCDEF", [18, 8, 10, 10, 10, 14]):
        ws.column_dimensions[col_letter].width = width
    ws.freeze_panes = "A5"


# ── CLI entry point ──────────────────────────────────────────────────────────
if __name__ == "__main__":
    import sys, json
    if len(sys.argv) < 2:
        print("Usage: python pdf_table_extractor.py <input.pdf> [output.xlsx]")
        sys.exit(1)

    out = sys.argv[2] if len(sys.argv) > 2 else None
    result = extract_tables_to_excel(sys.argv[1], out)
    print(json.dumps(result, indent=2))
