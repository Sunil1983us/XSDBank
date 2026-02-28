"""
xsd_ig_analyser.py  —  XSD vs IG Cross-Reference Analyser
Compares an ISO 20022 implementation-specific XSD (e.g. NPC pacs.008) against
an IG Extractor Excel output for the same message to surface:

  GAP CATEGORIES
  ──────────────
  ✅ ALIGNED       field present in both; status, multiplicity & rules agree
  🟡 STATUS DIFF   XSD says Mandatory but IG says Optional (or vice versa)
  🔵 RULES DIFF    Usage rule / format rule / length differs between sources
  🔴 EXCLUDED      field present in IG (ISO spec) but absent from NPC XSD (= Not Permitted by scheme)
  🟠 XSD ONLY      field defined in XSD but absent from IG (= sub-element or new type restriction)
  ⬜ MULT DIFF     multiplicity (minOccurs/maxOccurs) differs between XSD and IG

Output workbook sheets
──────────────────────
  Summary       — dashboard: counts by category + source metadata
  All Fields    — every field from both sources, side-by-side, colour-coded
  Gaps Only     — only rows where XSD and IG disagree (filtered view)
  Excluded      — fields in ISO IG but absent from NPC XSD (not permitted)
  XSD Only      — fields in XSD not in IG
  Mandatory     — all mandatory fields with XSD + IG data side by side

Usage
─────
    from xsd_ig_analyser import analyse
    result = analyse('pacs.008.xsd', 'NPC_IG.xlsx',
                     message_sheet='pacs_008_001_08',
                     scheme_label='NPC', version='2025 v1.1')
    # result keys: total, aligned, status_diff, rules_diff, excluded,
    #              xsd_only, mult_diff, output_path
"""

import os
import re
import xml.etree.ElementTree as ET
from collections import defaultdict

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl import load_workbook


# ── Constants ──────────────────────────────────────────────────────────────────

# XSD annotation source labels used in NPC/EPC schemas
_ANN_YELLOW      = 'Yellow Field'
_ANN_WHITE       = 'White Field'
_ANN_RED         = 'Red Field'
_ANN_USAGE_RULE  = 'Usage Rule'
_ANN_FORMAT_RULE = 'Format Rule'
_ANN_RULEBOOK    = 'Rulebook'

# Gap category labels and colours
_GAP = {
    'ALIGNED':      {'label': '✅ ALIGNED',         'bg': 'FFE2EFDA', 'fg': 'FF1A5632'},
    'STATUS_DIFF':  {'label': '🟡 STATUS DIFF',     'bg': 'FFFFF2CC', 'fg': 'FF7D4F00'},
    'RULES_DIFF':   {'label': '🔵 RULES DIFF',      'bg': 'FFD6E4F7', 'fg': 'FF1F3864'},
    'EXCLUDED':     {'label': '🔴 EXCLUDED',         'bg': 'FFFFCCCC', 'fg': 'FF8B0000'},
    'XSD_ONLY':     {'label': '🟠 XSD ONLY',        'bg': 'FFFCE4D6', 'fg': 'FF843C0C'},
    'MULT_DIFF':    {'label': '⬜ MULT DIFF',        'bg': 'FFF2F2F2', 'fg': 'FF595959'},
    'BOTH_DIFF':    {'label': '🟠 STATUS+RULES',     'bg': 'FFFCE8C3', 'fg': 'FF7F4B00'},
}

# XSD status display labels shown in the workbook cells
_XSD_STATUS_LABEL = {
    'Mandatory':     '🟡 Mandatory',
    'Conditional':   '🔶 Conditional',   # Yellow + minOccurs=0: must populate when condition is met
    'Optional':      '⬜ Optional',
    'Not Permitted': '🔴 Not Permitted',
}

_C = {
    'navy':       'FF1F3864',
    'blue':       'FF2E74B5',
    'lt_blue':    'FFD6E4F7',
    'white':      'FFFFFFFF',
    'lt_gray':    'FFF5F5F5',
    'dark_gray':  'FF595959',
    'green':      'FF375623',
}


# ── Style helpers ──────────────────────────────────────────────────────────────

def _fill(hex8): return PatternFill('solid', fgColor=hex8)
def _font(bold=False, color='FF000000', size=9):
    return Font(name='Arial', bold=bold, size=size, color=color)
def _border():
    s = Side(style='thin', color='FFD0D0D0')
    return Border(left=s, right=s, top=s, bottom=s)
def _align(h='left', v='top', wrap=True):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def _cell(ws, row, col, value, bg=None, fg='FF333333', bold=False,
          size=9, h='left', wrap=True):
    c = ws.cell(row=row, column=col, value=value)
    c.font      = _font(bold=bold, color=fg, size=size)
    c.alignment = _align(h=h, v='top', wrap=wrap)
    c.border    = _border()
    if bg:
        c.fill = _fill(bg)
    return c


# ── XSD parser ────────────────────────────────────────────────────────────────

_NS = {'xs': 'http://www.w3.org/2001/XMLSchema'}


def _detect_npc_suffix(root_elem):
    """
    Derive the scheme restriction suffix from Document root element type name.
    Document type = "Document_{SUFFIX}", suffix = everything after "Document".
    Falls back to frequency-based regex scan if Document element is absent.
    This avoids mistakenly capturing _2/_3 variant suffixes.
    """
    # Primary: derive from Document root element type attribute
    doc = root_elem.find('xs:element', _NS)
    if doc is not None:
        doc_type = doc.get('type', '')
        if doc_type.startswith('Document') and len(doc_type) > len('Document'):
            return doc_type[len('Document'):]
    # Fallback: regex scan — strip trailing _N variant, pick most-frequent base suffix
    suffix_counts: dict = {}
    for ct in root_elem.findall('.//xs:complexType', _NS):
        name = ct.get('name', '')
        m = re.search(r'(_(?:NPC|EPC|SEPA)[\w.-]+?)(_\d+)?$', name)
        if m and m.group(1):
            sfx = m.group(1)
            suffix_counts[sfx] = suffix_counts.get(sfx, 0) + 1
    if suffix_counts:
        return max(suffix_counts, key=suffix_counts.get)
    return ''


def _get_annotations(elem):
    """Return dict of {source: text|[texts]} from xs:annotation/xs:documentation."""
    docs = {}
    for doc in elem.findall('xs:annotation/xs:documentation', _NS):
        src = doc.get('source', '')
        txt = (doc.text or '').strip()
        if not src:
            continue
        if src in docs:
            prev = docs[src]
            docs[src] = (prev if isinstance(prev, list) else [prev]) + [txt]
        else:
            docs[src] = txt
    return docs


def _flatten(val):
    if isinstance(val, list):
        return '\n'.join(v for v in val if v)
    return val or ''


def _get_type_constraints(root_elem, type_name):
    """Extract length / pattern / enumeration constraints from a type."""
    info = {'length': '', 'pattern': '', 'enums': []}
    st = root_elem.find(f'.//xs:simpleType[@name="{type_name}"]', _NS)
    if st is None:
        return info
    for restr in st.findall('.//xs:restriction', _NS):
        mn = restr.find('xs:minLength', _NS)
        mx = restr.find('xs:maxLength', _NS)
        pt = restr.find('xs:pattern', _NS)
        if mn is not None and mx is not None:
            info['length'] = f"{mn.get('value','')} .. {mx.get('value','')}"
        elif mx is not None:
            info['length'] = f"1 .. {mx.get('value','')}"
        if pt is not None:
            info['pattern'] = pt.get('value', '')
        for en in restr.findall('xs:enumeration', _NS):
            info['enums'].append(en.get('value', ''))
    return info


def parse_xsd(xsd_path):
    """
    Parse an ISO 20022 implementation XSD and return a list of field dicts.

    Each dict contains:
      xml_tag, xpath, multiplicity, type_base, xsd_status,
      xsd_status_source, xsd_usage_rule, xsd_format_rule, xsd_rulebook,
      xsd_length, xsd_pattern, xsd_enums

    Status derivation rules (in priority order):
      1. Red Field annotation   → Not Permitted
      2. Yellow Field annotation → Mandatory
      3. White Field annotation  → Optional
      4. No annotation + minOccurs='1' → Mandatory (structurally required)
      5. No annotation + minOccurs='0' → Optional  (structurally optional)

    Visited key = (type_name, xpath) so the same type at different XPaths
    is walked independently — critical for types with multiple NPC variants
    (e.g. PostalAddress24_NPC vs PostalAddress24_NPC_2) which have different
    field-level colour annotations depending on context.
    """
    tree = ET.parse(xsd_path)
    root = tree.getroot()
    suffix = _detect_npc_suffix(root)

    def _resolve_child_type(type_name):
        """
        Given a type name from an element definition, return the best
        NPC-specific variant to walk into.
        If the type already has the NPC suffix, use it as-is.
        Otherwise try appending the suffix; if not found use the base type.
        """
        if not type_name:
            return type_name
        if suffix and suffix in type_name:
            return type_name          # already scheme-specific
        if suffix:
            npc = type_name + suffix
            if root.find(f'.//xs:complexType[@name="{npc}"]', _NS) is not None:
                return npc
        return type_name              # fall back to base ISO type

    def _base_type(type_name):
        """Strip NPC suffix (and trailing _2/_3 variant index) for display."""
        if not suffix:
            return type_name
        base = type_name.replace(suffix, '')
        # Remove trailing variant index like _2, _3 that follows the suffix
        base = re.sub(r'_\d+$', '', base)
        return base

    fields = []

    def walk(type_name, xpath_prefix, visited, depth=0):
        if depth > 16:
            return
        # Key on (type_name, xpath_prefix) so the same type at different
        # XPaths is walked independently — required for multi-variant types.
        key = (type_name, xpath_prefix)
        if key in visited:
            return
        visited = visited | {key}

        ct = root.find(f'.//xs:complexType[@name="{type_name}"]', _NS)
        if ct is None:
            return

        # Elements can live in restriction/sequence (NPC) or plain sequence (base ISO)
        elems = (
            ct.findall('xs:complexContent/xs:restriction/xs:sequence/xs:element', _NS)
            or ct.findall('xs:sequence/xs:element', _NS)
        )

        for elem in elems:
            name = elem.get('name', '')
            if not name:
                continue

            etype   = elem.get('type', '')
            min_occ = elem.get('minOccurs', '1')
            max_occ = elem.get('maxOccurs', '1')
            docs    = _get_annotations(elem)
            xpath   = f"{xpath_prefix}/{name}" if xpath_prefix else name
            tc      = _get_type_constraints(root, etype)

            # ── Status derivation ─────────────────────────────────────────
            # Priority: explicit annotation > structural (minOccurs)
            # Yellow + minOccurs=0 → "Conditional" — must populate when parent
            #   element is present / condition is met (e.g. cross-border, alias)
            # Yellow + minOccurs=1 → "Mandatory"   — unconditionally required
            # White  + any         → "Optional"
            # No annotation + minOccurs=1 → "Mandatory" (structural)
            # No annotation + minOccurs=0 → "Optional"  (structural)
            # Red  + any           → "Not Permitted"
            if _ANN_RED in docs:
                status     = 'Not Permitted'
                status_src = 'explicit (Red annotation)'
            elif _ANN_YELLOW in docs:
                if min_occ == '0':
                    status     = 'Conditional'
                    status_src = f'explicit (Yellow annotation, minOccurs={min_occ}..{max_occ})'
                else:
                    status     = 'Mandatory'
                    status_src = f'explicit (Yellow annotation, minOccurs={min_occ})'
            elif _ANN_WHITE in docs:
                status     = 'Optional'
                status_src = 'explicit (White annotation)'
            elif min_occ == '1':
                status     = 'Mandatory'
                status_src = 'derived (minOccurs=1, no colour annotation)'
            else:
                status     = 'Optional'
                status_src = f'derived (minOccurs={min_occ}, no colour annotation)'

            fields.append({
                'xml_tag':          name,
                'xpath':            xpath,
                'multiplicity':     f"{min_occ}..{max_occ}",
                'type_base':        _base_type(etype),
                'xsd_status':       status,
                'xsd_status_src':   status_src,
                'xsd_usage_rule':   _flatten(docs.get(_ANN_USAGE_RULE, '')),
                'xsd_format_rule':  _flatten(docs.get(_ANN_FORMAT_RULE, '')),
                'xsd_rulebook':     _flatten(docs.get(_ANN_RULEBOOK, '')),
                'xsd_length':       tc['length'],
                'xsd_pattern':      tc['pattern'],
                'xsd_enums':        ', '.join(tc['enums'][:15]),
            })

            # Recurse — use exact type reference from the element (may already
            # be a specific NPC variant like PostalAddress24_NPC_2).
            # Only auto-append suffix when the element type has no scheme suffix yet.
            child_type = _resolve_child_type(etype)
            if child_type:
                walk(child_type, xpath, visited, depth + 1)

    # Find document root element type
    doc_elem = root.find('xs:element', _NS)
    if doc_elem is not None:
        doc_type = doc_elem.get('type', '')
        if doc_type:
            walk(doc_type, '', set(), 0)

    return fields, suffix


# ── IG Excel reader ───────────────────────────────────────────────────────────

def _ig_status_from_fill(fill):
    """Return colour label from cell fill. 'None' means no colour coding present."""
    if fill is None or fill.fill_type != 'solid':
        return None
    rgb6 = (fill.fgColor.rgb or 'FFFFFFFF')[-6:].upper()
    return {'FFF2CC': 'Yellow', 'FFCCCC': 'Red'}.get(rgb6, None)


def _derive_ig_status(fill, mult_str, usage_rule):
    """
    Derive IG field status using three-priority cascade:

    Priority 1 — Cell fill colour (explicit, works for colour-coded EPC/NPC IGs):
        Yellow (#FFF2CC) → Mandatory
        Red    (#FFCCCC) → Not Permitted

    Priority 2 — Multiplicity column (structural, always reliable):
        min=1 (e.g. 1..1, 1..n) → Mandatory
        min=0 (e.g. 0..1, 0..n) → Optional / Conditional

    Priority 3 — Usage Rule text (conditional mandatory):
        Rule starts with "Mandatory" but min=0 → Conditional

    This handles IGs where colour coding was not captured from the PDF
    (all rows appear white) by falling back to the authoritative Multiplicity column.
    """
    # Priority 1: explicit cell colour
    colour = _ig_status_from_fill(fill)
    if colour == 'Yellow':
        return 'Mandatory'
    if colour == 'Red':
        return 'Not Permitted'

    # Priority 2: Multiplicity column
    mult = str(mult_str or '').strip()
    min_occ = mult.split('..')[0].strip() if '..' in mult else mult
    if min_occ == '1':
        return 'Mandatory'

    # Priority 3: Usage Rule text signals conditional mandatory
    rule_lower = str(usage_rule or '').strip().lower()
    if rule_lower.startswith('mandatory'):
        return 'Conditional'

    return 'Optional'


def parse_ig_sheet(ig_excel_path, sheet_name):
    """
    Read one sheet from an IG Extractor Excel and return list of field dicts.

    Status derivation (three-priority cascade — see _derive_ig_status):
      1. Cell fill colour: Yellow=Mandatory, Red=Not Permitted
      2. Multiplicity: min=1 → Mandatory, min=0 → Optional
      3. Usage Rule text starting with "Mandatory" → Conditional

    This correctly handles IGs where the PDF colour coding was not captured
    (all rows appear white), using the Multiplicity column as the authoritative
    structural source instead.

    Each dict contains:
      xml_tag, xpath, multiplicity, type,
      ig_status (Mandatory / Conditional / Optional / Not Permitted),
      ig_status_src (how status was derived),
      ig_length, ig_usage_rule, ig_rulebook, ig_format_rule,
      ig_code_restrictions, ig_iso_definition
    """
    wb = load_workbook(ig_excel_path, data_only=True)

    if sheet_name not in wb.sheetnames:
        for sn in wb.sheetnames:
            if sheet_name.lower() in sn.lower() or sn.lower() in sheet_name.lower():
                sheet_name = sn
                break

    if sheet_name not in wb.sheetnames:
        raise ValueError(
            f"Sheet '{sheet_name}' not found in {ig_excel_path}. "
            f"Available: {wb.sheetnames}"
        )

    ws = wb[sheet_name]
    headers = [ws.cell(row=3, column=c).value for c in range(1, ws.max_column + 1)]
    col_idx = {h: i + 1 for i, h in enumerate(headers) if h}

    def _v(row_num, col_name):
        ci = col_idx.get(col_name)
        return str(ws.cell(row=row_num, column=ci).value or '').strip() if ci else ''

    # Detect whether this IG Excel has any colour coding at all
    # Sample first 20 data rows — if all white, flag for audit log
    has_colour = False
    for r in range(4, min(ws.max_row + 1, 24)):
        fill = ws.cell(row=r, column=1).fill
        if _ig_status_from_fill(fill) is not None:
            has_colour = True
            break

    fields = []
    for r in range(4, ws.max_row + 1):
        xpath = _v(r, 'XPath')
        if not xpath:
            continue

        mult  = _v(r, 'Multiplicity')
        rule  = _v(r, 'SEPA/NPC Usage Rules')
        fill  = ws.cell(row=r, column=1).fill
        status = _derive_ig_status(fill, mult, rule)

        # Record how status was derived for transparency
        colour = _ig_status_from_fill(fill)
        if colour in ('Yellow', 'Red'):
            status_src = f'explicit (cell colour: {colour})'
        elif str(mult or '').strip().split('..')[0].strip() == '1':
            status_src = f'derived (Multiplicity={mult})'
        elif rule.strip().lower().startswith('mandatory'):
            status_src = 'derived (Usage Rule starts with "Mandatory")'
        else:
            status_src = f'derived (Multiplicity={mult}, no colour)'

        fields.append({
            'xml_tag':              _v(r, 'XML Tag'),
            'xpath':                xpath,
            'multiplicity':         mult,
            'type':                 _v(r, 'Type'),
            'ig_status':            status,
            'ig_status_src':        status_src,
            'ig_length':            _v(r, 'SEPA/NPC Length'),
            'ig_usage_rule':        rule,
            'ig_rulebook':          _v(r, 'SEPA/NPC Rulebook'),
            'ig_format_rule':       _v(r, 'SEPA/NPC Format Rules'),
            'ig_code_restrictions': _v(r, 'SEPA/NPC Code Restrictions'),
            'ig_iso_definition':    _v(r, 'ISO Definition'),
            'ig_index':             _v(r, 'Index'),
            'ig_element_name':      _v(r, 'Element Name'),
            'ig_has_colour':        has_colour,
        })

    return fields


# ── Cross-reference logic ─────────────────────────────────────────────────────

def _normalise_rule(text):
    """Normalise text for comparison: lowercase, collapse whitespace."""
    return re.sub(r'\s+', ' ', (text or '').lower().strip())


def _rules_differ(xsd_rule, ig_rule):
    """True if the two rule strings are meaningfully different."""
    xn = _normalise_rule(xsd_rule)
    ign = _normalise_rule(ig_rule)
    if not xn and not ign:
        return False
    if not xn or not ign:
        return True
    return xn != ign


def _mult_differs(xsd_mult, ig_mult):
    """True if multiplicities disagree (handles 'unbounded' vs 'n' etc.)."""
    def _norm(m):
        return m.replace('unbounded', 'n').replace(' ', '')
    return _norm(xsd_mult) != _norm(ig_mult) if ig_mult else False


def cross_reference(xsd_fields, ig_fields):
    """
    Cross-reference XSD fields against IG fields.
    Returns a list of comparison rows with 'gap_category' set.
    """
    # Build lookup by xml_tag
    xsd_by_tag = defaultdict(list)
    for f in xsd_fields:
        xsd_by_tag[f['xml_tag']].append(f)

    ig_by_tag = defaultdict(list)
    for f in ig_fields:
        ig_by_tag[f['xml_tag']].append(f)

    all_tags = sorted(set(list(xsd_by_tag.keys()) + list(ig_by_tag.keys())))
    rows = []

    for tag in all_tags:
        xsd_list = xsd_by_tag.get(tag, [])
        ig_list  = ig_by_tag.get(tag, [])

        if not xsd_list and ig_list:
            # In IG but NOT in NPC XSD → excluded by scheme
            for igf in ig_list:
                rows.append(_make_row(None, igf, 'EXCLUDED'))

        elif xsd_list and not ig_list:
            # In XSD but NOT in IG
            for xf in xsd_list:
                rows.append(_make_row(xf, None, 'XSD_ONLY'))

        else:
            # Present in both — compare
            # Match pairs by XPath similarity; if can't match, use first of each
            used_ig = set()
            for xf in xsd_list:
                # Find best IG match
                best_ig = None
                for i, igf in enumerate(ig_list):
                    if i in used_ig:
                        continue
                    # Exact tag match is good enough here
                    best_ig = igf
                    used_ig.add(i)
                    break

                if best_ig is None:
                    rows.append(_make_row(xf, None, 'XSD_ONLY'))
                    continue

                gap = _classify_gap(xf, best_ig)
                rows.append(_make_row(xf, best_ig, gap))

            # Any unmatched IG entries
            for i, igf in enumerate(ig_list):
                if i not in used_ig:
                    rows.append(_make_row(None, igf, 'EXCLUDED'))

    return rows


def _classify_gap(xf, igf):
    """Determine the gap category between an XSD field and an IG field."""
    xsd_st = xf['xsd_status']  # Mandatory / Conditional / Optional / Not Permitted
    ig_st  = igf['ig_status']  # Mandatory / Optional / Not Permitted / Conditional
    ig_src = igf.get('ig_status_src', '')  # how IG status was derived

    # Status difference rules:
    #   Mandatory    vs any-non-Mandatory  → diff
    #   Conditional  vs Optional           → NO diff (0..n Yellow = sometimes required, IG shows Optional)
    #   Conditional  vs Mandatory          → diff  (IG says always required, XSD says conditional)
    #   Optional     vs Mandatory          → diff
    #   Not Permitted vs non-Not Permitted → diff
    status_diff = False
    if xsd_st == 'Mandatory'     and ig_st != 'Mandatory':       status_diff = True
    if xsd_st == 'Not Permitted' and ig_st != 'Not Permitted':   status_diff = True
    if xsd_st == 'Optional'      and ig_st == 'Mandatory':       status_diff = True
    # XSD=Conditional + IG=Mandatory + IG source is explicit Yellow cell colour:
    # This is an EPC expression artefact — both XSD Yellow+0..n and IG Yellow mean
    # "required by scheme when in scope". Treat as ALIGNED (not a real diff).
    if xsd_st == 'Conditional' and ig_st == 'Mandatory':
        if ig_src == 'explicit (cell colour: Yellow)':
            pass  # aligned — same Yellow semantics expressed differently
        else:
            status_diff = True   # IG derived Mandatory but XSD says Conditional
    # Conditional vs Optional is NOT a diff — both mean "include when condition met"
    # Mandatory (XSD) vs Conditional (IG) IS a diff — IG underestimates the requirement
    if xsd_st == 'Mandatory'     and ig_st == 'Conditional':     status_diff = True

    # Multiplicity difference
    mult_diff = _mult_differs(xf['multiplicity'], igf['multiplicity'])

    # Rules difference: check usage rule and format rule
    rules_diff = (
        _rules_differ(xf['xsd_usage_rule'], igf['ig_usage_rule']) or
        _rules_differ(xf['xsd_format_rule'], igf['ig_format_rule']) or
        _rules_differ(xf['xsd_length'], igf['ig_length'])
    )

    if status_diff and rules_diff:
        return 'BOTH_DIFF'
    if status_diff:
        return 'STATUS_DIFF'
    if rules_diff:
        return 'RULES_DIFF'
    if mult_diff:
        return 'MULT_DIFF'
    return 'ALIGNED'


def _make_row(xf, igf, gap_cat):
    """Merge XSD and IG data into a single comparison row."""
    row = {
        'gap_category':   gap_cat,
        # XSD side
        'xml_tag':        (xf or igf or {}).get('xml_tag', ''),
        'xsd_xpath':      xf['xpath']           if xf else '',
        'xsd_mult':       xf['multiplicity']    if xf else '',
        'xsd_type':       xf['type_base']       if xf else '',
        'xsd_status':     _XSD_STATUS_LABEL.get(xf['xsd_status'], xf['xsd_status'])  if xf else '—',
        'xsd_status_src': xf['xsd_status_src']  if xf else '',
        'xsd_usage_rule': xf['xsd_usage_rule']  if xf else '',
        'xsd_format_rule':xf['xsd_format_rule'] if xf else '',
        'xsd_rulebook':   xf['xsd_rulebook']    if xf else '',
        'xsd_length':     xf['xsd_length']      if xf else '',
        'xsd_pattern':    xf['xsd_pattern']     if xf else '',
        'xsd_enums':      xf['xsd_enums']       if xf else '',
        # IG side
        'ig_xpath':       igf['xpath']              if igf else '',
        'ig_mult':        igf['multiplicity']       if igf else '',
        'ig_type':        igf['type']               if igf else '',
        'ig_status':      igf['ig_status']          if igf else '—',
        'ig_status_src':  igf.get('ig_status_src','')  if igf else '',
        'ig_length':      igf['ig_length']          if igf else '',
        'ig_usage_rule':  igf['ig_usage_rule']      if igf else '',
        'ig_format_rule': igf['ig_format_rule']     if igf else '',
        'ig_rulebook':    igf['ig_rulebook']        if igf else '',
        'ig_code_restr':  igf['ig_code_restrictions'] if igf else '',
        'ig_iso_def':     igf['ig_iso_definition']  if igf else '',
        'ig_index':       igf['ig_index']           if igf else '',
        'ig_element':     igf['ig_element_name']    if igf else '',
    }

    # Derive delta text
    row['delta'] = _delta_text(row, gap_cat)
    return row


def _delta_text(row, gap_cat):
    """Human-readable description of the difference found."""
    if gap_cat == 'ALIGNED':
        return ''
    if gap_cat == 'EXCLUDED':
        return (f"Field exists in ISO IG (index {row.get('ig_index','')}) "
                f"but is absent from NPC XSD — NPC scheme does not permit this field.")
    if gap_cat == 'XSD_ONLY':
        return 'Field defined in NPC XSD restriction but not found in IG Extractor output — may be a sub-element or new constraint not yet documented in IG.'

    parts = []
    xs, ig = row.get('xsd_status',''), row.get('ig_status','')
    xs_src = row.get('xsd_status_src','')
    if xs and xs != ig and xs != '—':
        ig_src = row.get('ig_status_src','')
        if xs == '🔶 Conditional':
            parts.append(f'XSD marks as Conditional (Yellow, {row.get("xsd_mult","")}) — mandatory when condition is met. IG shows "{ig}" [{ig_src}].')
        else:
            parts.append(f'Status: XSD says "{xs}" [{xs_src}]. IG says "{ig}" [{ig_src}].')

    xm, im = row.get('xsd_mult',''), row.get('ig_mult','')
    if xm and im and _mult_differs(xm, im):
        parts.append(f'Multiplicity: XSD={xm}, IG={im}.')

    xl, il = row.get('xsd_length',''), row.get('ig_length','')
    if xl and il and _normalise_rule(xl) != _normalise_rule(il):
        parts.append(f'Length: XSD="{xl}", IG="{il}".')
    elif xl and not il:
        parts.append(f'Length constraint in XSD ({xl}) not documented in IG.')
    elif il and not xl:
        parts.append(f'Length in IG ({il}) not reflected in XSD type.')

    xu, iu = row.get('xsd_usage_rule',''), row.get('ig_usage_rule','')
    if _rules_differ(xu, iu):
        if xu and not iu:
            parts.append('Usage rule in XSD annotation not present in IG.')
        elif iu and not xu:
            parts.append('Usage rule in IG not captured in XSD annotation.')
        else:
            parts.append('Usage rules differ between XSD and IG — review both.')

    xf_r, if_r = row.get('xsd_format_rule',''), row.get('ig_format_rule','')
    if _rules_differ(xf_r, if_r):
        parts.append('Format rules differ.')

    return ' '.join(parts) if parts else 'Difference detected — review both columns.'


# ── Excel writer ───────────────────────────────────────────────────────────────

# Column layout for the main analysis sheets
_COLS = [
    # Common
    ('Gap',            14, 'gap'),
    ('XML Tag',        16, 'common'),
    ('IG Index',        8, 'ig'),
    ('Element Name',   28, 'ig'),
    # XSD columns
    ('XSD XPath',      50, 'xsd'),
    ('XSD Mult',       10, 'xsd'),
    ('XSD Status',     18, 'xsd'),
    ('XSD Status Src', 36, 'xsd'),
    ('XSD Type',       22, 'xsd'),
    ('XSD Length',     14, 'xsd'),
    ('XSD Usage Rule', 42, 'xsd'),
    ('XSD Format Rule',28, 'xsd'),
    ('XSD Rulebook',   18, 'xsd'),
    ('XSD Pattern',    28, 'xsd'),
    ('XSD Enums',      30, 'xsd'),
    # IG columns
    ('IG XPath',       50, 'ig'),
    ('IG Mult',        10, 'ig'),
    ('IG Status',      18, 'ig'),
    ('IG Status Src',  36, 'ig'),
    ('IG Type',        22, 'ig'),
    ('IG Length',      14, 'ig'),
    ('IG Usage Rule',  42, 'ig'),
    ('IG Format Rule', 28, 'ig'),
    ('IG Rulebook',    18, 'ig'),
    ('IG Code Restr.', 30, 'ig'),
    ('IG ISO Def',     52, 'ig'),
    # Delta
    ('Delta / Finding', 60, 'delta'),
]

_XSD_HDR_BG  = 'FF1F3864'   # navy
_IG_HDR_BG   = 'FF375623'   # dark green
_GAP_HDR_BG  = 'FF7F4B00'   # dark orange
_CMN_HDR_BG  = 'FF595959'   # dark grey
_DELTA_HDR_BG= 'FF4B1C82'   # dark purple


def _section_bg(section):
    return {
        'xsd': _XSD_HDR_BG, 'ig': _IG_HDR_BG,
        'gap': _GAP_HDR_BG, 'delta': _DELTA_HDR_BG,
    }.get(section, _CMN_HDR_BG)


def _write_analysis_sheet(ws, rows, sheet_title, show_all=True):
    """Write one analysis sheet (All Fields or Gaps Only)."""
    total_cols = len(_COLS)

    # ── Title row ──────────────────────────────────────────────────────────────
    ws.cell(row=1, column=1, value=sheet_title)
    ws.merge_cells(f'A1:{get_column_letter(total_cols)}1')
    ws['A1'].font      = _font(bold=True, size=13, color='FFFFFFFF')
    ws['A1'].fill      = _fill(_C['navy'])
    ws['A1'].alignment = _align(h='left', v='center', wrap=False)
    ws.row_dimensions[1].height = 22

    # ── Legend row ─────────────────────────────────────────────────────────────
    legend_items = list(_GAP.items())
    span = max(1, total_cols // len(legend_items))
    col = 1
    for i, (key, cfg) in enumerate(legend_items):
        end = col + span - 1 if i < len(legend_items) - 1 else total_cols
        ws.cell(row=2, column=col, value=cfg['label'])
        if end > col:
            ws.merge_cells(f'{get_column_letter(col)}2:{get_column_letter(end)}2')
        c = ws[f'{get_column_letter(col)}2']
        c.font      = _font(bold=True, size=8, color=cfg['fg'])
        c.fill      = _fill(cfg['bg'])
        c.alignment = _align(h='center', v='center', wrap=False)
        c.border    = _border()
        col = end + 1
    ws.row_dimensions[2].height = 13

    # ── Section header row 3 ──────────────────────────────────────────────────
    # Build CONTIGUOUS section runs (IG columns appear in two separate blocks)
    sec_runs = []
    cur_sec, cur_start, cur_end = None, None, None
    for ci, (_, _, section) in enumerate(_COLS, 1):
        if section != cur_sec:
            if cur_sec is not None:
                sec_runs.append((cur_sec, cur_start, cur_end))
            cur_sec, cur_start, cur_end = section, ci, ci
        else:
            cur_end = ci
    if cur_sec is not None:
        sec_runs.append((cur_sec, cur_start, cur_end))

    sec_labels = {
        'xsd':    '◀  XSD  (Schema Ground Truth)',
        'ig':     'IG  (Rules Ground Truth)  ▶',
        'gap':    'Gap',
        'common': 'Field',
        'delta':  'Delta / Finding',
    }
    for section, start, end in sec_runs:
        bg  = _section_bg(section)
        lbl = sec_labels.get(section, section.upper())
        ws.cell(row=3, column=start, value=lbl)
        if end > start:
            ws.merge_cells(f'{get_column_letter(start)}3:{get_column_letter(end)}3')
        c = ws[f'{get_column_letter(start)}3']
        c.font      = _font(bold=True, size=9, color='FFFFFFFF')
        c.fill      = _fill(bg)
        c.alignment = _align(h='center', v='center', wrap=False)
        c.border    = _border()
    ws.row_dimensions[3].height = 16

    # ── Column headers row 4 ──────────────────────────────────────────────────
    for ci, (col_name, _, section) in enumerate(_COLS, 1):
        bg = _section_bg(section)
        c  = ws.cell(row=4, column=ci, value=col_name)
        c.font      = _font(bold=True, size=8, color='FFFFFFFF')
        c.fill      = _fill(bg)
        c.alignment = _align(h='center', v='center', wrap=True)
        c.border    = _border()
    ws.row_dimensions[4].height = 30
    ws.freeze_panes = 'A5'

    # ── Data rows ──────────────────────────────────────────────────────────────
    data_row = 5
    for row in rows:
        gap  = row.get('gap_category', 'ALIGNED')
        if not show_all and gap == 'ALIGNED':
            continue
        cfg  = _GAP.get(gap, _GAP['ALIGNED'])
        bg   = cfg['bg']
        fg   = cfg['fg']

        values = [
            cfg['label'],
            row.get('xml_tag', ''),
            row.get('ig_index', ''),
            row.get('ig_element', ''),
            row.get('xsd_xpath', ''),
            row.get('xsd_mult', ''),
            row.get('xsd_status', ''),
            row.get('xsd_status_src', ''),
            row.get('xsd_type', ''),
            row.get('xsd_length', ''),
            row.get('xsd_usage_rule', ''),
            row.get('xsd_format_rule', ''),
            row.get('xsd_rulebook', ''),
            row.get('xsd_pattern', ''),
            row.get('xsd_enums', ''),
            row.get('ig_xpath', ''),
            row.get('ig_mult', ''),
            row.get('ig_status', ''),
            row.get('ig_status_src', ''),
            row.get('ig_type', ''),
            row.get('ig_length', ''),
            row.get('ig_usage_rule', ''),
            row.get('ig_format_rule', ''),
            row.get('ig_rulebook', ''),
            row.get('ig_code_restr', ''),
            row.get('ig_iso_def', ''),
            row.get('delta', ''),
        ]

        for ci, val in enumerate(values, 1):
            section = _COLS[ci - 1][2]
            if section == 'gap':
                cell_bg, cell_fg, cell_bold = bg, fg, True
            elif section == 'delta':
                cell_bg = 'FFFFFACD' if gap not in ('ALIGNED',) else _C['lt_gray']
                cell_fg = 'FF4B1C82' if gap not in ('ALIGNED',) else _C['dark_gray']
                cell_bold = False
            elif section == 'xsd':
                cell_bg = bg if gap not in ('EXCLUDED',) else _C['lt_gray']
                cell_fg = fg if gap not in ('EXCLUDED',) else _C['dark_gray']
                cell_bold = False
            elif section == 'ig':
                cell_bg = _C['lt_gray'] if gap == 'XSD_ONLY' else bg
                cell_fg = _C['dark_gray'] if gap == 'XSD_ONLY' else fg
                cell_bold = False
            else:
                cell_bg = bg; cell_fg = fg; cell_bold = False

            _cell(ws, data_row, ci, val, bg=cell_bg, fg=cell_fg,
                  bold=cell_bold, size=8,
                  h='center' if ci <= 2 else 'left')

        ws.row_dimensions[data_row].height = 55
        data_row += 1

    # ── Column widths ──────────────────────────────────────────────────────────
    for ci, (_, width, _) in enumerate(_COLS, 1):
        ws.column_dimensions[get_column_letter(ci)].width = width

    return data_row - 5


def _write_summary_sheet(ws, stats, xsd_meta, ig_meta, scheme, version):
    """Write the dashboard Summary sheet."""
    total_cols = 10

    ws.cell(row=1, column=1,
            value=f'XSD vs IG Analysis  —  {scheme}  {version}')
    ws.merge_cells(f'A1:{get_column_letter(total_cols)}1')
    ws['A1'].font      = _font(bold=True, size=14, color='FFFFFFFF')
    ws['A1'].fill      = _fill(_C['navy'])
    ws['A1'].alignment = _align(h='left', v='center', wrap=False)
    ws.row_dimensions[1].height = 28

    # Source files
    row = 2
    for label, value in [
        ('XSD File',        xsd_meta.get('filename','')),
        ('XSD Scheme ID',   xsd_meta.get('scheme_id','')),
        ('IG Excel File',   ig_meta.get('filename','')),
        ('IG Sheet',        ig_meta.get('sheet','')),
        ('Message',         ig_meta.get('message','')),
        ('Scheme',          scheme),
        ('Version',         version),
    ]:
        _cell(ws, row, 1, label, bg=_C['lt_gray'], fg=_C['dark_gray'], bold=True, h='right')
        _cell(ws, row, 2, value, bg='FFFFFFFF', fg='FF222222')
        ws.merge_cells(f'B{row}:{get_column_letter(total_cols)}{row}')
        ws.row_dimensions[row].height = 15
        row += 1

    row += 1

    # ── Gap counts dashboard ──────────────────────────────────────────────────
    ws.cell(row=row, column=1, value='Analysis Results')
    ws.merge_cells(f'A{row}:{get_column_letter(total_cols)}{row}')
    ws[f'A{row}'].font      = _font(bold=True, size=11, color='FFFFFFFF')
    ws[f'A{row}'].fill      = _fill(_C['blue'])
    ws[f'A{row}'].alignment = _align(h='center', v='center')
    ws.row_dimensions[row].height = 18
    row += 1

    gap_order = ['ALIGNED', 'EXCLUDED', 'STATUS_DIFF', 'RULES_DIFF',
                 'MULT_DIFF', 'BOTH_DIFF', 'XSD_ONLY']

    # Header
    for ci, lbl in enumerate(['Category', 'Count', '% of Total', 'Meaning', 'Action'], 1):
        _cell(ws, row, ci, lbl, bg=_C['blue'], fg='FFFFFFFF', bold=True, h='center')
    ws.merge_cells(f'E{row}:{get_column_letter(total_cols)}{row}')
    ws.row_dimensions[row].height = 22
    row += 1

    total = sum(stats.values())
    meanings = {
        'ALIGNED':     ('XSD and IG agree on status, rules, and multiplicity',
                        'No action needed'),
        'EXCLUDED':    ('Field exists in ISO IG but NPC XSD excludes it — Not Permitted',
                        '⚠️ Do NOT send — ensure source systems do not populate'),
        'STATUS_DIFF': ('XSD and IG disagree on field status (Mandatory vs Optional)',
                        '⚠️ Investigate — XSD is structural ground truth; IG may need update'),
        'RULES_DIFF':  ('Usage rule / format rule / length differs between XSD and IG',
                        '🔍 Review both sources and align implementation with the stricter rule'),
        'MULT_DIFF':   ('Multiplicity (min/max occurrences) differs',
                        '🔍 Check occurrence handling in your implementation'),
        'BOTH_DIFF':   ('Both status and rules differ',
                        '⚠️ High priority — both structural and textual differences'),
        'XSD_ONLY':    ('In XSD restriction but not in IG output — may be sub-element',
                        '🔍 Verify whether this element needs to be implemented'),
    }

    for gap_key in gap_order:
        cfg     = _GAP.get(gap_key, _GAP['ALIGNED'])
        count   = stats.get(gap_key, 0)
        pct     = f'{count/total*100:.1f}%' if total else '0%'
        meaning, action = meanings.get(gap_key, ('', ''))
        bg = cfg['bg']; fg = cfg['fg']

        _cell(ws, row, 1, cfg['label'],  bg=bg, fg=fg, bold=True)
        _cell(ws, row, 2, count,         bg=bg, fg=fg, bold=True, h='center')
        _cell(ws, row, 3, pct,           bg=bg, fg=fg, h='center')
        _cell(ws, row, 4, meaning,       bg=bg, fg=fg, size=8)
        ws.merge_cells(f'D{row}:G{row}')
        _cell(ws, row, 8, action,        bg=bg, fg=fg, size=8)
        ws.merge_cells(f'H{row}:{get_column_letter(total_cols)}{row}')
        ws.row_dimensions[row].height = 22
        row += 1

    # Totals
    _cell(ws, row, 1, 'TOTAL', bg=_C['lt_blue'], fg=_C['navy'], bold=True)
    _cell(ws, row, 2, total,   bg=_C['lt_blue'], fg=_C['navy'], bold=True, h='center')
    ws.row_dimensions[row].height = 20
    row += 2

    # ── How to use ────────────────────────────────────────────────────────────
    ws.cell(row=row, column=1, value='📖 How to Use This Analysis')
    ws.merge_cells(f'A{row}:{get_column_letter(total_cols)}{row}')
    ws[f'A{row}'].font = _font(bold=True, size=10, color='FFFFFFFF')
    ws[f'A{row}'].fill = _fill(_C['green'])
    ws[f'A{row}'].alignment = _align(h='left', v='center')
    ws.row_dimensions[row].height = 16
    row += 1

    guide = [
        ('All Fields sheet',   'Every field from both XSD and IG side by side. Sort by Gap column to prioritise.'),
        ('Gaps Only sheet',     'Filtered view — only rows where XSD and IG disagree. Start here for action items.'),
        ('Excluded sheet',      'Fields that exist in the ISO standard but are NOT permitted by the NPC scheme.'),
        ('Mandatory sheet',     'All fields the XSD marks as Mandatory — your minimum implementation checklist.'),
        ('XSD is authoritative','For structural questions (what\'s mandatory, what\'s excluded) trust the XSD.'),
        ('IG is authoritative', 'For business rules, AT-references, and textual usage rules trust the IG text.'),
        ('STATUS DIFF rows',    'These are highest priority: XSD and IG disagree on whether a field is required.'),
        ('Conditional Fields',   '🔶 Yellow + minOccurs=0 in XSD — populate when the condition in the usage rule is met (e.g. cross-border, alias). Check the XSD Status Src column for details.'),
        ('RULES DIFF rows',     'Review length constraints — XSD type restriction may be stricter than IG text.'),
    ]
    for heading, text in guide:
        _cell(ws, row, 1, heading, bg='FFE8F5E9', fg=_C['green'], bold=True, h='left', size=9)
        ws.merge_cells(f'B{row}:{get_column_letter(total_cols)}{row}')
        _cell(ws, row, 2, text, bg='FFFAFAFA', fg='FF333333', size=8)
        ws.row_dimensions[row].height = 20
        row += 1

    # Column widths
    for ci, w in enumerate([28, 30, 12, 45, 45, 12, 12, 45, 12, 12], 1):
        ws.column_dimensions[get_column_letter(ci)].width = w


def _write_simple_sheet(ws, rows, title, gap_filter):
    """Write a simple filtered sheet (Excluded, XSD Only, Mandatory)."""
    filtered = [r for r in rows if r.get('gap_category') == gap_filter] if gap_filter else rows
    if not filtered:
        ws.cell(row=1, column=1, value=f'{title} — No entries')
        return 0
    return _write_analysis_sheet(ws, filtered, title, show_all=True)


def _write_mandatory_sheet(ws, rows, title='Mandatory Fields Checklist'):
    """Write all fields that XSD or IG marks as Mandatory."""
    mandatory = [r for r in rows
                 if r.get('xsd_status') in ('🟡 Mandatory', 'Mandatory',
                                             '🔶 Conditional', 'Conditional')
                 or r.get('ig_status') in ('Mandatory', 'Conditional')]
    return _write_analysis_sheet(ws, mandatory, title, show_all=True)


# ── Public API ────────────────────────────────────────────────────────────────

def analyse(
    xsd_path: str,
    ig_excel_path: str,
    output_path: str,
    message_sheet: str = '',
    scheme_label: str = '',
    version: str = '',
) -> dict:
    """
    Cross-reference an XSD against an IG Excel and write an analysis workbook.

    Parameters
    ----------
    xsd_path       : Path to the ISO 20022 XSD (implementation-specific)
    ig_excel_path  : Path to IG Extractor Excel output
    output_path    : Where to save the analysis workbook
    message_sheet  : Sheet name in the IG Excel (e.g. 'pacs_008_001_08')
                     If empty, the first non-Summary sheet is used
    scheme_label   : e.g. 'NPC', 'EPC'
    version        : e.g. '2025 v1.1'

    Returns
    -------
    dict with keys: total, aligned, excluded, status_diff, rules_diff,
                    mult_diff, both_diff, xsd_only, output_path, sheet_used
    """
    # ── Parse inputs ──────────────────────────────────────────────────────────
    xsd_fields, suffix = parse_xsd(xsd_path)

    # Auto-detect sheet
    ig_wb = load_workbook(ig_excel_path, data_only=True, read_only=True)
    if not message_sheet:
        for sn in ig_wb.sheetnames:
            if sn != 'Summary':
                message_sheet = sn
                break
    ig_wb.close()

    ig_fields = parse_ig_sheet(ig_excel_path, message_sheet)

    # ── Auto-detect metadata ──────────────────────────────────────────────────
    xsd_fname = os.path.basename(xsd_path)
    ig_fname  = os.path.basename(ig_excel_path)
    message   = message_sheet.replace('_', '.')

    if not scheme_label:
        fname_up = xsd_fname.upper()
        scheme_label = ('NPC' if 'NPC' in fname_up else
                        'EPC' if 'EPC' in fname_up else 'SEPA')
    if not version:
        m = re.search(r'(\d{4}[_-]?V[\d.]+)', xsd_fname, re.I)
        if m:
            version = m.group(1).replace('_', ' ')

    # ── Cross-reference ───────────────────────────────────────────────────────
    rows = cross_reference(xsd_fields, ig_fields)

    # Count categories
    from collections import Counter
    cat_counts = Counter(r['gap_category'] for r in rows)
    stats = {k: cat_counts.get(k, 0) for k in _GAP}

    # ── Build workbook ────────────────────────────────────────────────────────
    wb = Workbook()
    wb.remove(wb.active)

    xsd_meta = {'filename': xsd_fname, 'scheme_id': suffix}
    ig_meta  = {'filename': ig_fname,  'sheet': message_sheet, 'message': message}

    # Summary
    sum_ws = wb.create_sheet('Summary')
    _write_summary_sheet(sum_ws, stats, xsd_meta, ig_meta, scheme_label, version)

    # All Fields
    all_ws = wb.create_sheet('All Fields')
    _write_analysis_sheet(all_ws, rows, f'All Fields — {message} XSD vs IG', show_all=True)

    # Gaps Only (exclude ALIGNED)
    gaps_ws = wb.create_sheet('Gaps Only')
    _write_analysis_sheet(gaps_ws, rows,
                          f'Gaps Only — {message} XSD vs IG (ALIGNED rows hidden)',
                          show_all=False)

    # Excluded (in IG, not in XSD)
    excl_ws = wb.create_sheet('Excluded by Scheme')
    _write_simple_sheet(excl_ws, rows, f'Excluded by {scheme_label} — Not Permitted', 'EXCLUDED')

    # Mandatory
    mand_ws = wb.create_sheet('Mandatory Checklist')
    _write_mandatory_sheet(mand_ws, rows, f'Mandatory Fields — {scheme_label} {message}')

    # Conditional Fields (Yellow + 0..n: required when condition met)
    cond_ws = wb.create_sheet('Conditional Fields')
    cond_rows = [r for r in rows if '🔶 Conditional' in str(r.get('xsd_status',''))]
    _write_analysis_sheet(cond_ws, cond_rows,
                          f'Conditional Fields — {scheme_label} {message}  (Yellow + 0..n)',
                          show_all=True)

    # XSD Only
    xsd_only_ws = wb.create_sheet('XSD Only')
    _write_simple_sheet(xsd_only_ws, rows, 'In XSD, Not in IG', 'XSD_ONLY')

    os.makedirs(os.path.dirname(os.path.abspath(output_path)), exist_ok=True)
    wb.save(output_path)

    return {
        'total':       len(rows),
        'aligned':     stats['ALIGNED'],
        'excluded':    stats['EXCLUDED'],
        'status_diff': stats['STATUS_DIFF'],
        'rules_diff':  stats['RULES_DIFF'],
        'mult_diff':   stats['MULT_DIFF'],
        'both_diff':   stats['BOTH_DIFF'],
        'xsd_only':    stats['XSD_ONLY'],
        'sheet_used':  message_sheet,
        'output_path': output_path,
    }
