#!/usr/bin/env python3
"""
ISO 20022 Payment Message - Comprehensive Schema Analyzer
Extracts ALL metadata including Yellow/White fields from XSD annotations
"""

import xml.etree.ElementTree as ET
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
import argparse
from pathlib import Path

class ISO20022ComprehensiveAnalyzer:
    """Extract complete metadata from ISO 20022 payment message schemas"""
    
    def __init__(self, xsd_file):
        self.xsd_file = xsd_file
        self.tree = ET.parse(xsd_file)
        self.root = self.tree.getroot()
        self.ns = {'xs': 'http://www.w3.org/2001/XMLSchema'}
        self.elements = []
        self.type_definitions = {}
        self.sequence_counter = 0
        
        self._build_type_map()
    
    def _build_type_map(self):
        """Build map of all type definitions"""
        for complex_type in self.root.findall('.//xs:complexType', self.ns):
            name = complex_type.get('name')
            if name:
                self.type_definitions[name] = complex_type
        
        for simple_type in self.root.findall('.//xs:simpleType', self.ns):
            name = simple_type.get('name')
            if name:
                self.type_definitions[name] = simple_type
    
    def extract_all_metadata(self):
        """Extract comprehensive metadata in XSD sequence order"""
        print("\n⏳ Extracting comprehensive metadata from ISO 20022 schema...")
        
        root_elements = self.root.findall('./xs:element', self.ns)
        
        for root_elem in root_elements:
            elem_name = root_elem.get('name', '')
            if elem_name:
                self._process_element(root_elem, '', 0, False)
        
        print(f"   ✅ Extracted {len(self.elements)} elements")
        print(f"   ✅ XSD sequence order maintained")
    
    def _process_element(self, element, parent_path, level, in_choice):
        """Process element and recurse into its type"""
        elem_name = element.get('name', element.get('ref', ''))
        if not elem_name:
            return
        
        if ':' in elem_name:
            elem_name = elem_name.split(':')[-1]
        
        current_path = f"{parent_path}/{elem_name}" if parent_path else elem_name
        
        elem_type = element.get('type', '')
        if ':' in elem_type:
            elem_type = elem_type.split(':')[-1]
        
        min_occurs = element.get('minOccurs', '1')
        max_occurs = element.get('maxOccurs', '1')
        
        # Extract annotation (including Yellow/White classification)
        annotation = self._extract_annotation(element)
        
        # Classify field from XSD annotation
        field_class = self._classify_field_from_xsd(element, elem_name, min_occurs, annotation)
        
        # Extract restrictions
        restriction_info = self._extract_restriction_from_element(element)
        
        self.sequence_counter += 1
        element_info = {
            'sequence': self.sequence_counter,
            'level': level,
            'path': current_path,
            'element': elem_name,
            'type': elem_type,
            'min_occurs': min_occurs,
            'max_occurs': max_occurs,
            'required': min_occurs != '0',
            'in_choice': in_choice,
            'annotation': annotation.get('documentation', ''),
            'usage_rules': annotation.get('usage_rules', ''),
            'business_rules': annotation.get('business_rules', ''),
            'field_classification': field_class,
            'pattern': restriction_info.get('pattern', ''),
            'min_length': restriction_info.get('min_length', ''),
            'max_length': restriction_info.get('max_length', ''),
            'min_value': restriction_info.get('min_value', ''),
            'max_value': restriction_info.get('max_value', ''),
            'enumeration': restriction_info.get('enumeration', ''),
            'fraction_digits': restriction_info.get('fraction_digits', ''),
            'total_digits': restriction_info.get('total_digits', ''),
            'whitespace': restriction_info.get('whitespace', ''),
        }
        
        self.elements.append(element_info)
        
        # Recurse into type definition
        inline_complex = element.find('xs:complexType', self.ns)
        if inline_complex is not None:
            self._process_complex_type(inline_complex, current_path, level + 1, in_choice)
        elif elem_type and elem_type in self.type_definitions:
            type_def = self.type_definitions[elem_type]
            if type_def.tag.endswith('complexType'):
                self._process_complex_type(type_def, current_path, level + 1, in_choice)
    
    def _process_complex_type(self, complex_type, parent_path, level, in_choice):
        """Process complex type maintaining sequence order"""
        for child in complex_type:
            tag = child.tag.replace('{http://www.w3.org/2001/XMLSchema}', '')
            
            if tag == 'sequence':
                self._process_sequence(child, parent_path, level, in_choice)
            elif tag == 'choice':
                self._process_choice(child, parent_path, level)
            elif tag == 'complexContent':
                self._process_complex_content(child, parent_path, level, in_choice)
    
    def _process_sequence(self, sequence, parent_path, level, in_choice):
        """Process sequence maintaining order"""
        for child in sequence:
            tag = child.tag.replace('{http://www.w3.org/2001/XMLSchema}', '')
            if tag == 'element':
                self._process_element(child, parent_path, level, in_choice)
    
    def _process_choice(self, choice, parent_path, level):
        """Process choice maintaining order"""
        for child in choice:
            tag = child.tag.replace('{http://www.w3.org/2001/XMLSchema}', '')
            if tag == 'element':
                self._process_element(child, parent_path, level, in_choice=True)
    
    def _process_complex_content(self, complex_content, parent_path, level, in_choice):
        """Process complexContent"""
        for child in complex_content:
            tag = child.tag.replace('{http://www.w3.org/2001/XMLSchema}', '')
            if tag in ['extension', 'restriction']:
                for sub_child in child:
                    sub_tag = sub_child.tag.replace('{http://www.w3.org/2001/XMLSchema}', '')
                    if sub_tag == 'sequence':
                        self._process_sequence(sub_child, parent_path, level, in_choice)
                    elif sub_tag == 'choice':
                        self._process_choice(sub_child, parent_path, level)
    
    def _extract_annotation(self, element):
        """Extract all annotation information"""
        result = {
            'documentation': '',
            'usage_rules': '',
            'business_rules': '',
        }
        
        annotation = element.find('xs:annotation', self.ns)
        if annotation is not None:
            docs = annotation.findall('xs:documentation', self.ns)
            doc_texts = []
            
            for doc in docs:
                source = doc.get('source', '').strip()
                text = doc.text.strip() if doc.text else ''
                
                # Skip Yellow/White field markers in documentation text
                if source not in ['Yellow Field', 'White Field']:
                    if text:
                        doc_texts.append(text)
                
                # Capture business/usage rules
                if 'rulebook' in source.lower() or 'rule' in source.lower():
                    if text:
                        result['business_rules'] = text
                elif 'usage' in source.lower():
                    if text:
                        result['usage_rules'] = text
            
            result['documentation'] = ' | '.join(doc_texts) if doc_texts else ''
        
        return result
    
    def _classify_field_from_xsd(self, element, elem_name, min_occurs, annotation):
        """
        Classify field from XSD annotation (PRIORITY #1)
        Falls back to heuristics only if not in XSD
        
        ISSUE 2 FIX: Read from <xs:documentation source="Yellow Field"/> or "White Field"
        """
        # PRIORITY 1: Check XSD annotation source attribute
        ann_elem = element.find('xs:annotation', self.ns)
        if ann_elem is not None:
            docs = ann_elem.findall('xs:documentation', self.ns)
            for doc in docs:
                source = doc.get('source', '').strip()
                
                # Exact match
                if source == 'Yellow Field':
                    return '🟡 Yellow (ISO Specified)'
                elif source == 'White Field':
                    return '⚪ White (ISO Specified)'
                
                # Case-insensitive match
                source_lower = source.lower()
                if 'yellow' in source_lower and 'field' in source_lower:
                    return '🟡 Yellow (ISO Specified)'
                elif 'white' in source_lower and 'field' in source_lower:
                    return '⚪ White (ISO Specified)'
        
        # PRIORITY 2: Heuristic fallback for non-annotated fields
        element_lower = elem_name.lower()
        core_fields = [
            'id', 'identification', 'iban', 'bic', 'account',
            'amt', 'amount', 'currency', 'ccy',
            'debtor', 'creditor', 'party', 'agent', 'name', 'nm',
            'date', 'dt', 'datetime', 'crdt',
            'ref', 'reference', 'msgid', 'messageid', 'endtoendid', 'txid',
            'code', 'cd', 'status', 'sts'
        ]
        
        if any(core in element_lower for core in core_fields):
            return '🟡 Yellow (Inferred)'
        
        # PRIORITY 3: Not specified in XSD
        return '⚪ NA (Not Specified)'
    
    def _extract_restriction_from_element(self, element):
        """Extract restriction information"""
        result = {}
        
        simple_type = element.find('xs:simpleType', self.ns)
        if simple_type is not None:
            restriction = simple_type.find('xs:restriction', self.ns)
            if restriction is not None:
                result = self._extract_restriction_details(restriction)
        
        return result
    
    def _extract_restriction_details(self, restriction):
        """Extract all restriction details"""
        result = {}
        
        pattern = restriction.find('xs:pattern', self.ns)
        if pattern is not None:
            result['pattern'] = pattern.get('value', '')
        
        min_length = restriction.find('xs:minLength', self.ns)
        if min_length is not None:
            result['min_length'] = min_length.get('value', '')
        
        max_length = restriction.find('xs:maxLength', self.ns)
        if max_length is not None:
            result['max_length'] = max_length.get('value', '')
        
        length = restriction.find('xs:length', self.ns)
        if length is not None:
            result['min_length'] = length.get('value', '')
            result['max_length'] = length.get('value', '')
        
        min_inclusive = restriction.find('xs:minInclusive', self.ns)
        if min_inclusive is not None:
            result['min_value'] = min_inclusive.get('value', '')
        
        max_inclusive = restriction.find('xs:maxInclusive', self.ns)
        if max_inclusive is not None:
            result['max_value'] = max_inclusive.get('value', '')
        
        enums = restriction.findall('xs:enumeration', self.ns)
        if enums:
            enum_values = [e.get('value', '') for e in enums]
            result['enumeration'] = ', '.join(enum_values[:10])
            if len(enum_values) > 10:
                result['enumeration'] += f' ... ({len(enum_values)} total)'
        
        fraction_digits = restriction.find('xs:fractionDigits', self.ns)
        if fraction_digits is not None:
            result['fraction_digits'] = fraction_digits.get('value', '')
        
        total_digits = restriction.find('xs:totalDigits', self.ns)
        if total_digits is not None:
            result['total_digits'] = total_digits.get('value', '')
        
        whitespace = restriction.find('xs:whiteSpace', self.ns)
        if whitespace is not None:
            result['whitespace'] = whitespace.get('value', '')
        
        return result
    
    def generate_comprehensive_excel(self, output_file):
        """Generate comprehensive Excel report"""
        print("\n⏳ Generating comprehensive Excel report...")
        
        wb = Workbook()
        
        # Sheet 1: Complete Structure
        ws1 = wb.active
        ws1.title = "Complete Structure"
        
        headers = [
            'Seq', 'Level', 'Full Path', 'Element', 'Type',
            'Required', 'Min', 'Max', 'In Choice',
            'Field Class', 'Documentation', 'Usage Rules', 'Business Rules',
            'Pattern', 'Min Length', 'Max Length', 'Enumeration',
            'Min Value', 'Max Value', 'Fraction Digits', 'Total Digits'
        ]
        
        ws1.append(headers)
        
        # Style header
        for cell in ws1[1]:
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
            cell.alignment = Alignment(horizontal='center', wrap_text=True)
        
        # Add data
        for elem in self.elements:
            row = [
                elem['sequence'],
                elem['level'],
                elem['path'],
                elem['element'],
                elem['type'],
                'Yes' if elem['required'] else 'No',
                elem['min_occurs'],
                elem['max_occurs'],
                'Yes' if elem['in_choice'] else 'No',
                elem['field_classification'],
                elem['annotation'],
                elem['usage_rules'],
                elem['business_rules'],
                elem['pattern'],
                elem['min_length'],
                elem['max_length'],
                elem['enumeration'],
                elem['min_value'],
                elem['max_value'],
                elem['fraction_digits'],
                elem['total_digits']
            ]
            ws1.append(row)
            
            # Highlight Yellow fields
            row_num = ws1.max_row
            if '🟡' in elem['field_classification']:
                for col_letter in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T', 'U']:
                    ws1[f'{col_letter}{row_num}'].fill = PatternFill(start_color='FFF8DC', end_color='FFF8DC', fill_type='solid')
        
        # Column widths
        column_widths = [6, 6, 60, 30, 35, 10, 6, 6, 10, 28, 50, 40, 40, 35, 10, 10, 40, 10, 10, 10, 10]
        for i, width in enumerate(column_widths, 1):
            ws1.column_dimensions[get_column_letter(i)].width = width
        
        ws1.freeze_panes = 'D2'
        
        # Sheet 2: Field Classification
        ws2 = wb.create_sheet("Field Classification")
        ws2.append(['Classification', 'Element', 'Full Path', 'Required', 'Documentation'])
        
        for cell in ws2[1]:
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        
        for elem in self.elements:
            ws2.append([
                elem['field_classification'],
                elem['element'],
                elem['path'],
                'Yes' if elem['required'] else 'No',
                elem['annotation']
            ])
            
            row_num = ws2.max_row
            if '🟡' in elem['field_classification']:
                for col in ['A', 'B', 'C', 'D', 'E']:
                    ws2[f'{col}{row_num}'].fill = PatternFill(start_color='FFF8DC', end_color='FFF8DC', fill_type='solid')
        
        ws2.column_dimensions['A'].width = 28
        ws2.column_dimensions['B'].width = 35
        ws2.column_dimensions['C'].width = 60
        ws2.column_dimensions['D'].width = 10
        ws2.column_dimensions['E'].width = 50
        
        # Save
        wb.save(output_file)
        
        yellow_count = sum(1 for e in self.elements if '🟡' in e['field_classification'])
        white_count = sum(1 for e in self.elements if '⚪ White (ISO' in e['field_classification'])
        na_count = sum(1 for e in self.elements if 'NA' in e['field_classification'])
        
        print(f"\n✅ Comprehensive analysis saved: {output_file}")
        print(f"   📊 Total elements: {len(self.elements)}")
        print(f"   🟡 Yellow fields: {yellow_count} (ISO/Inferred)")
        print(f"   ⚪ White fields: {white_count} (ISO Specified)")
        print(f"   ⚪ NA fields: {na_count} (Not Specified)")
        print(f"   ✅ Sequence order: XSD maintained")
        print(f"   ✅ Full paths: Complete")


def main():
    parser = argparse.ArgumentParser(description='ISO 20022 Payment Message - Comprehensive Analysis')
    parser.add_argument('xsd_file', help='ISO 20022 XSD schema file')
    parser.add_argument('-o', '--output', help='Output Excel file', default='comprehensive_analysis.xlsx')
    
    args = parser.parse_args()
    
    if not Path(args.xsd_file).exists():
        print(f"❌ Error: File '{args.xsd_file}' not found")
        return
    
    print(f"\n{'='*70}")
    print("ISO 20022 PAYMENT MESSAGE - COMPREHENSIVE ANALYSIS")
    print("Extracts ALL metadata with Yellow/White from XSD annotations")
    print(f"{'='*70}")
    print(f"\n📂 Schema: {args.xsd_file}")
    
    analyzer = ISO20022ComprehensiveAnalyzer(args.xsd_file)
    analyzer.extract_all_metadata()
    analyzer.generate_comprehensive_excel(args.output)
    
    print(f"\n{'='*70}")
    print("✅ ANALYSIS COMPLETE!")
    print(f"{'='*70}\n")


if __name__ == '__main__':
    main()
