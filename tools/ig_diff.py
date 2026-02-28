"""
ig_diff.py  —  ISO 20022 IG Diff Tool
Compares two IG Extractor Excel outputs (e.g. EPC vs NPC) and produces a
colour-coded diff workbook: one sheet per shared message type, plus a Summary.

Change categories
-----------------
  🟢  ADDED       – field exists in File B but not File A
  🔴  REMOVED     – field exists in File A but not File B
  🟡  STATUS      – mandatory/optional/not-permitted status changed (row colour)
  🔵  RULES       – usage rules, rulebook, format rules, length changed
  🟠  BOTH        – both status AND rules changed
  ⬜  UNCHANGED   – identical in both files (hidden by default via row grouping)

Usage
-----
    from ig_diff import diff_ig
    result = diff_ig('EPC_IG.xlsx', 'NPC_IG.xlsx', 'diff_output.xlsx',
                     label_a='EPC', label_b='NPC')
    # result = {'total_changes': N, 'sheets': [...], 'messages_compared': [...]}
"""

import re
import os
from pathlib import Path
from typing import Optional

from openpyxl import Workbook, load_workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting.rule import Rule

# ── Colour palette ────────────────────────────────────────────────────────────
_C = {
    'navy':       'FF1F3864',
    'blue':       'FF2E74B5',
    'white':      'FFFFFFFF',
    'lt_gray':    'FFF2F2F2',
    'gray':       'FFD9D9D9',
    'dark_gray':  'FF595959',

    # diff status backgrounds
    'added':      'FFE2EFDA',   # soft green
    'added_txt':  'FF375623',
    'removed':    'FFFCE4D6',   # soft red/salmon
    'removed_txt':'FF843C0C',
    'status':     'FFFFF2CC',   # yellow
    'status_txt': 'FF7D6608',
    'rules':      'FFD6E4F7',   # light blue
    'rules_txt':  'FF1F4E79',
    'both':       'FFFCE8C3',   # orange-ish
    'both_txt':   'FF7F4B00',
    'unchanged':  'FFFFFFFF',
    'unchanged_txt': 'FF999999',

    # IG row colours (matching ig_extractor)
    'ig_yellow':  'FFFFF2CC',
    'ig_red':     'FFFFCCCC',
    'ig_white':   'FFFFFFFF',
}

# ── Diff category constants ───────────────────────────────────────────────────
ADDED     = 'ADDED'
REMOVED   = 'REMOVED'
STATUS    = 'STATUS'
RULES     = 'RULES'
BOTH      = 'BOTH'
UNCHANGED = 'UNCHANGED'

_DIFF_LABELS = {
    ADDED:     '🟢 ADDED',
    REMOVED:   '🔴 REMOVED',
    STATUS:    '🟡 STATUS',
    RULES:     '🔵 RULES',
    BOTH:      '🟠 BOTH',
    UNCHANGED: '⬜ SAME',
}

_DIFF_BG = {
    ADDED:     _C['added'],
    REMOVED:   _C['removed'],
    STATUS:    _C['status'],
    RULES:     _C['rules'],
    BOTH:      _C['both'],
    UNCHANGED: _C['unchanged'],
}

_DIFF_FG = {
    ADDED:     _C['added_txt'],
    REMOVED:   _C['removed_txt'],
    STATUS:    _C['status_txt'],
    RULES:     _C['rules_txt'],
    BOTH:      _C['both_txt'],
    UNCHANGED: _C['unchanged_txt'],
}

# Fields compared for RULES changes
_RULE_FIELDS = [
    'Multiplicity',
    'SEPA/NPC Length',
    'SEPA/NPC Usage Rules',
    'SEPA/NPC Rulebook',
    'SEPA/NPC Format Rules',
    'SEPA/NPC Code Restrictions',
]

# All IG data columns in display order
_IG_COLS = [
    'Index', 'Multiplicity', 'XPath', 'Element Name',
    'ISO Name', 'ISO Definition', 'XML Tag', 'Type',
    'ISO Length', 'SEPA/NPC Length',
    'SEPA/NPC Usage Rules', 'SEPA/NPC Rulebook',
    'SEPA/NPC Format Rules', 'SEPA/NPC FractDigits',
    'SEPA/NPC Inclusive', 'SEPA/NPC Code Restrictions',
]


# ── Helpers ───────────────────────────────────────────────────────────────────

def _fill(hex8: str) -> PatternFill:
    return PatternFill('solid', fgColor=hex8)

def _font(bold=False, color='FF000000', size=9, name='Arial') -> Font:
    return Font(name=name, bold=bold, size=size, color=color)

def _border_thin() -> Border:
    s = Side(style='thin', color='FFD0D0D0')
    return Border(left=s, right=s, top=s, bottom=s)

def _border_thick() -> Border:
    s = Side(style='medium', color='FF9E9E9E')
    return Border(left=s, right=s, top=s, bottom=s)

def _align(h='left', v='top', wrap=True) -> Alignment:
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def _cell(ws, row, col, value='', bold=False, bg=None, fg='FF000000',
          size=9, h='left', v='top', wrap=True, border=True):
    c = ws.cell(row=row, column=col, value=value)
    c.font      = _font(bold=bold, color=fg, size=size)
    c.alignment = _align(h=h, v=v, wrap=wrap)
    if bg:
        c.fill = _fill(bg)
    if border:
        c.border = _border_thin()
    return c


# ── Load IG workbook into structured dict ─────────────────────────────────────

def _ig_color_name(rgb6: str) -> str:
    """Convert 6-char hex to status label."""
    m = {'FFF2CC': 'Yellow/Core', 'FFCCCC': 'Red/NotPermitted'}
    return m.get(rgb6.upper(), 'White/Optional')

def _read_ig_workbook(path: str) -> dict:
    """
    Returns:
        {
          'sheets': {
              'pacs_008_001_08': {
                  'meta': {'message': ..., 'label': ...},
                  'rows':  { xpath: {col: value, ..., '_color': 'FFF2CC', '_color_name': ...} }
              }, ...
          },
          'summary': [{'sheet':..,'message':..,'label':..,'total':..,'yellow':..,'white':..,'red':..}]
        }
    """
    wb = load_workbook(path, data_only=True)
    result = {'sheets': {}, 'summary': []}

    # Read summary sheet
    if 'Summary' in wb.sheetnames:
        ws = wb['Summary']
        for r in range(4, ws.max_row + 1):
            row = [ws.cell(row=r, column=c).value for c in range(1, 9)]
            if row[0]:
                result['summary'].append({
                    'sheet':   str(row[0] or ''),
                    'message': str(row[1] or ''),
                    'label':   str(row[2] or ''),
                    'total':   row[3] or 0,
                    'yellow':  row[4] or 0,
                    'white':   row[5] or 0,
                    'red':     row[6] or 0,
                    'pages':   str(row[7] or ''),
                })

    # Read each data sheet
    for sheet_name in wb.sheetnames:
        if sheet_name == 'Summary':
            continue
        ws = wb[sheet_name]
        # Headers are on row 3
        headers = [ws.cell(row=3, column=c).value for c in range(1, ws.max_column + 1)]
        if not any(h for h in headers if h in _IG_COLS):
            continue  # not an IG data sheet

        rows = {}
        for r in range(4, ws.max_row + 1):
            row = {}
            for c, h in enumerate(headers, 1):
                if h:
                    row[h] = ws.cell(row=r, column=c).value or ''
            # Extract fill colour
            fill = ws.cell(row=r, column=1).fill
            if fill and fill.fill_type == 'solid' and fill.fgColor:
                rgb = fill.fgColor.rgb or 'FFFFFFFF'
                rgb6 = rgb[-6:].upper()
            else:
                rgb6 = 'FFFFFF'
            row['_color']      = rgb6
            row['_color_name'] = _ig_color_name(rgb6)

            xpath = str(row.get('XPath', '') or '').strip()
            if xpath:
                rows[xpath] = row

        result['sheets'][sheet_name] = {'rows': rows}

    return result


# ── Diff engine ───────────────────────────────────────────────────────────────

def _diff_sheets(rows_a: dict, rows_b: dict) -> list:
    """
    Returns list of diff records:
    {
        'xpath', 'change', 'changed_fields': {field: (a_val, b_val)},
        'row_a': dict|None, 'row_b': dict|None
    }
    """
    all_xpaths = list(dict.fromkeys(list(rows_a.keys()) + list(rows_b.keys())))
    records = []

    for xpath in all_xpaths:
        a = rows_a.get(xpath)
        b = rows_b.get(xpath)

        if a is None:
            records.append({'xpath': xpath, 'change': ADDED,
                             'changed_fields': {}, 'row_a': None, 'row_b': b})
            continue
        if b is None:
            records.append({'xpath': xpath, 'change': REMOVED,
                             'changed_fields': {}, 'row_a': a, 'row_b': None})
            continue

        # Compare rule fields
        changed_fields = {}
        for field in _RULE_FIELDS:
            va = str(a.get(field, '') or '').strip()
            vb = str(b.get(field, '') or '').strip()
            if va != vb:
                changed_fields[field] = (va, vb)

        # Compare status (row colour)
        status_changed = a['_color'] != b['_color']
        if status_changed:
            changed_fields['Status'] = (a['_color_name'], b['_color_name'])

        if not changed_fields:
            change = UNCHANGED
        elif status_changed and len(changed_fields) > 1:
            change = BOTH
        elif status_changed:
            change = STATUS
        else:
            change = RULES

        records.append({'xpath': xpath, 'change': change,
                         'changed_fields': changed_fields,
                         'row_a': a, 'row_b': b})

    return records


# ── Excel writer ──────────────────────────────────────────────────────────────

# ── Business impact descriptions ────────────────────────────────────────────

def _business_impact(change: str, changed_fields: dict,
                     row_a: dict, row_b: dict) -> str:
    """Return a plain-English business impact sentence for this diff row."""
    if change == ADDED:
        elem = (row_b or {}).get('Element Name', '') or (row_b or {}).get('XML Tag', '')
        mult = (row_b or {}).get('Multiplicity', '')
        color_b = (row_b or {}).get('_color_name', '')
        status_str = ''
        if 'Core' in color_b:
            status_str = ' It is Core Mandatory — must be included in every NPC message.'
        elif 'NotPermitted' in color_b:
            status_str = ' It is blocked in NPC — do not include this field.'
        else:
            status_str = ' It is optional in NPC.'
        mult_str = f' (multiplicity {mult})' if mult else ''
        return f'Field "{elem}"{mult_str} is new in NPC and does not exist in EPC.{status_str}'

    if change == REMOVED:
        elem = (row_a or {}).get('Element Name', '') or (row_a or {}).get('XML Tag', '')
        color_a = (row_a or {}).get('_color_name', '')
        was_str = ' It was Core Mandatory in EPC.' if 'Core' in color_a else ''
        return f'Field "{elem}" exists in EPC but is not present in NPC.{was_str} NPC processing can omit this field.'

    status_change = changed_fields.get('Status')
    rule_changes  = {k: v for k, v in changed_fields.items() if k != 'Status'}

    lines = []

    if status_change:
        a_status, b_status = status_change
        if 'Core' in a_status and 'Optional' in b_status:
            lines.append('Status RELAXED: was Core Mandatory in EPC, is Optional in NPC. NPC implementations are not required to populate this field.')
        elif 'Optional' in a_status and 'Core' in b_status:
            lines.append('Status ELEVATED: was Optional in EPC, is Core Mandatory in NPC. This field must be present in every NPC message.')
        elif 'NotPermitted' in b_status:
            lines.append('Field BLOCKED in NPC: must NOT be included in NPC messages. EPC implementations that populate this field need a separate NPC code path.')
        elif 'NotPermitted' in a_status and 'Core' in b_status:
            lines.append('Field UNBLOCKED in NPC: was Not Permitted in EPC, is now Core Mandatory. NPC must include this field.')
        elif 'NotPermitted' in a_status:
            lines.append('Field UNBLOCKED in NPC: was Not Permitted in EPC, is now Optional in NPC.')
        else:
            lines.append(f'Status changed from {a_status} (EPC) to {b_status} (NPC).')

    if 'SEPA/NPC Length' in rule_changes:
        old_len, new_len = rule_changes['SEPA/NPC Length']
        if old_len and new_len:
            lines.append(f'Length constraint changed: EPC allows {old_len}, NPC allows {new_len}. Validate field values against the applicable limit per scheme.')
        elif new_len:
            lines.append(f'NPC adds a length constraint: {new_len}. No constraint existed in EPC.')
        elif old_len:
            lines.append(f'EPC had a length constraint ({old_len}) that is not present in NPC.')

    if 'Multiplicity' in rule_changes:
        old_m, new_m = rule_changes['Multiplicity']
        lines.append(f'Multiplicity changed: EPC {old_m} → NPC {new_m}. Check cardinality handling in your implementation.')

    if 'SEPA/NPC Usage Rules' in rule_changes:
        old_r, new_r = rule_changes['SEPA/NPC Usage Rules']
        if old_r and new_r:
            lines.append('Usage rules differ between EPC and NPC. Review the rules in both columns and ensure your implementation satisfies the applicable scheme.')
        elif new_r:
            lines.append('NPC adds a usage rule not present in EPC. Review the NPC constraint.')
        elif old_r:
            lines.append('EPC had a usage rule that is not present in NPC — the constraint is relaxed.')

    if 'SEPA/NPC Rulebook' in rule_changes:
        old_rb, new_rb = rule_changes['SEPA/NPC Rulebook']
        lines.append(f'Rulebook reference differs: EPC references "{old_rb}", NPC references "{new_rb}". Ensure the correct AT-number is used per scheme.')

    if 'SEPA/NPC Format Rules' in rule_changes:
        lines.append('Format rules differ. Review format constraints per scheme before implementing.')

    if 'SEPA/NPC Code Restrictions' in rule_changes:
        lines.append('Code set restrictions differ between EPC and NPC. Validate allowed code values per scheme.')

    if not lines:
        return 'No business-relevant differences.'

    return ' '.join(lines)


# Diff sheet column layout - note: ISO Definition and XML Tag are added as
# shared contextual columns (not split A/B) since they are typically identical.
_SHOW_COLS = ['Index', 'XPath', 'Element Name', 'Multiplicity',
              'SEPA/NPC Length', 'SEPA/NPC Usage Rules',
              'SEPA/NPC Rulebook', 'SEPA/NPC Format Rules',
              'SEPA/NPC Code Restrictions']

def _write_diff_sheet(ws, diff_records: list, label_a: str, label_b: str,
                      message: str, section_label: str):
    """Write one diff sheet."""

    n_show = len(_SHOW_COLS)
    # Column layout:
    # 1: Change marker
    # 2..1+n_show: File A columns
    # 2+n_show..1+2*n_show: File B columns
    # 2+2*n_show: What Changed
    # 3+2*n_show: ISO Definition (shared context)
    # 4+2*n_show: XML Tag (shared context)
    # 5+2*n_show: Business Impact
    COL_CHANGE  = 1
    COL_A_START = 2
    COL_B_START = 2 + n_show
    COL_WHAT    = 2 + 2 * n_show
    COL_ISO_DEF = 3 + 2 * n_show
    COL_XML_TAG = 4 + 2 * n_show
    COL_IMPACT  = 5 + 2 * n_show

    total_cols = COL_IMPACT

    # ── Title ──────────────────────────────────────────────────────────────────
    t = ws['A1']
    t.value     = f'{message.upper()}  —  {section_label}  │  {label_a} vs {label_b}'
    ws.merge_cells(f'A1:{get_column_letter(total_cols)}1')
    t.font      = _font(bold=True, size=13, color=_C['white'])
    t.fill      = _fill(_C['navy'])
    t.alignment = _align(h='left', v='center')
    ws.row_dimensions[1].height = 22

    # ── Legend ──────────────────────────────────────────────────────────────────
    legend = [
        (ADDED,     f'🟢 ADDED — in {label_b} only'),
        (REMOVED,   f'🔴 REMOVED — in {label_a} only'),
        (STATUS,    '🟡 STATUS — mandatory/optional changed'),
        (RULES,     '🔵 RULES — length/usage/rulebook changed'),
        (BOTH,      '🟠 BOTH — status + rules changed'),
        (UNCHANGED, '⬜ SAME — identical in both'),
    ]
    leg_col_w = max(1, total_cols // len(legend))
    for i, (cat, text) in enumerate(legend):
        col_start = i * leg_col_w + 1
        col_end   = min((i + 1) * leg_col_w, total_cols)
        if i == len(legend) - 1:
            col_end = total_cols
        start_letter = get_column_letter(col_start)
        end_letter   = get_column_letter(col_end)
        c = ws.cell(row=2, column=col_start, value=text)
        if start_letter != end_letter:
            ws.merge_cells(f'{start_letter}2:{end_letter}2')
        c.font      = _font(bold=True, size=8, color=_DIFF_FG[cat])
        c.fill      = _fill(_DIFF_BG[cat])
        c.alignment = _align(h='center', v='center', wrap=False)
        c.border    = _border_thin()
    ws.row_dimensions[2].height = 14

    # ── Section headers (A-side / B-side) ──────────────────────────────────────
    c_chg = ws.cell(row=3, column=COL_CHANGE, value='Change')
    c_chg.font      = _font(bold=True, size=9, color=_C['white'])
    c_chg.fill      = _fill(_C['dark_gray'])
    c_chg.alignment = _align(h='center', v='center', wrap=False)

    hdr_a = ws.cell(row=3, column=COL_A_START, value=f'◀  {label_a.upper()}')
    hdr_a.font      = _font(bold=True, size=10, color=_C['white'])
    hdr_a.fill      = _fill(_C['blue'])
    hdr_a.alignment = _align(h='center', v='center', wrap=False)
    hdr_a.border    = _border_thin()
    ws.merge_cells(f'{get_column_letter(COL_A_START)}3:{get_column_letter(COL_B_START-1)}3')

    hdr_b = ws.cell(row=3, column=COL_B_START, value=f'{label_b.upper()}  ▶')
    hdr_b.font      = _font(bold=True, size=10, color=_C['white'])
    hdr_b.fill      = _fill('FF2E8B57')
    hdr_b.alignment = _align(h='center', v='center', wrap=False)
    hdr_b.border    = _border_thin()
    ws.merge_cells(f'{get_column_letter(COL_B_START)}3:{get_column_letter(COL_WHAT-1)}3')

    c_what = ws.cell(row=3, column=COL_WHAT, value='What Changed')
    c_what.font      = _font(bold=True, size=9, color=_C['white'])
    c_what.fill      = _fill(_C['dark_gray'])
    c_what.alignment = _align(h='center', v='center', wrap=False)
    c_what.border    = _border_thin()
    ws.row_dimensions[3].height = 18

    # ── Column headers row 4 ───────────────────────────────────────────────────
    for i, col_name in enumerate(_SHOW_COLS):
        # A-side header
        ca = ws.cell(row=4, column=COL_A_START + i, value=col_name)
        ca.font      = _font(bold=True, size=8, color=_C['white'])
        ca.fill      = _fill(_C['blue'])
        ca.alignment = _align(h='center', v='center', wrap=True)
        ca.border    = _border_thin()
        # B-side header
        cb = ws.cell(row=4, column=COL_B_START + i, value=col_name)
        cb.font      = _font(bold=True, size=8, color=_C['white'])
        cb.fill      = _fill('FF2E8B57')
        cb.alignment = _align(h='center', v='center', wrap=True)
        cb.border    = _border_thin()

    ws.cell(row=4, column=COL_CHANGE).font      = _font(bold=True, size=8, color=_C['white'])
    ws.cell(row=4, column=COL_CHANGE).fill      = _fill(_C['dark_gray'])
    ws.cell(row=4, column=COL_CHANGE).alignment = _align(h='center', wrap=False)
    ws.cell(row=4, column=COL_CHANGE).border    = _border_thin()
    ws.cell(row=4, column=COL_WHAT).font        = _font(bold=True, size=8, color=_C['white'])
    ws.cell(row=4, column=COL_WHAT).fill        = _fill(_C['dark_gray'])
    ws.cell(row=4, column=COL_WHAT).alignment   = _align(h='center', wrap=False)
    ws.cell(row=4, column=COL_WHAT).border      = _border_thin()

    # Shared context columns header
    _green_hdr = 'FF1A5C38'
    for col, label in [(COL_ISO_DEF, 'ISO Definition'), (COL_XML_TAG, 'XML Tag')]:
        c = ws.cell(row=3, column=col, value='Field Info')
        c.font = _font(bold=True, size=10, color=_C['white'])
        c.fill = _fill('FF4A4A4A')
        c.alignment = _align(h='center', v='center', wrap=False)
        c.border = _border_thin()
        ch = ws.cell(row=4, column=col, value=label)
        ch.font = _font(bold=True, size=8, color=_C['white'])
        ch.fill = _fill('FF4A4A4A')
        ch.alignment = _align(h='center', wrap=False)
        ch.border = _border_thin()

    # Business Impact column header
    bi3 = ws.cell(row=3, column=COL_IMPACT, value='Business Impact')
    bi3.font = _font(bold=True, size=10, color=_C['white'])
    bi3.fill = _fill('FF1A5C38')
    bi3.alignment = _align(h='center', v='center', wrap=False)
    bi3.border = _border_thin()
    bi4 = ws.cell(row=4, column=COL_IMPACT, value='Business Impact & Implementation Note')
    bi4.font = _font(bold=True, size=8, color=_C['white'])
    bi4.fill = _fill('FF1A5C38')
    bi4.alignment = _align(h='center', wrap=True)
    bi4.border = _border_thin()

    ws.row_dimensions[4].height = 30
    ws.freeze_panes = 'A5'

    # ── Data rows ──────────────────────────────────────────────────────────────
    data_row = 5
    unchanged_rows = []  # for row grouping

    for rec in diff_records:
        change   = rec['change']
        row_a    = rec['row_a'] or {}
        row_b    = rec['row_b'] or {}
        chg_flds = rec['changed_fields']

        bg   = _DIFF_BG[change]
        fg   = _DIFF_FG[change]
        lbl  = _DIFF_LABELS[change]

        # Change marker cell
        cm = ws.cell(row=data_row, column=COL_CHANGE, value=lbl)
        cm.font      = _font(bold=True, size=8, color=fg)
        cm.fill      = _fill(bg)
        cm.alignment = _align(h='center', v='center', wrap=False)
        cm.border    = _border_thin()

        # A-side values
        for i, col_name in enumerate(_SHOW_COLS):
            val  = str(row_a.get(col_name, '') or '')
            col  = COL_A_START + i
            # Highlight cell if this field changed
            if col_name in chg_flds or (col_name == 'Index' and change in (STATUS, BOTH)):
                cell_bg = _DIFF_BG.get(change, bg)
            else:
                # Use IG original colour for unchanged fields in A
                ig_color = row_a.get('_color', 'FFFFFF')
                ig_map   = {'FFF2CC': _C['ig_yellow'], 'FFCCCC': _C['ig_red']}
                cell_bg  = ig_map.get(ig_color, _C['white']) if change == UNCHANGED else bg
            c = ws.cell(row=data_row, column=col, value=val)
            c.font      = _font(size=8, color='FF404040')
            c.fill      = _fill(cell_bg)
            c.alignment = _align(wrap=True)
            c.border    = _border_thin()

        # B-side values — highlight changed fields in stronger colour
        for i, col_name in enumerate(_SHOW_COLS):
            val = str(row_b.get(col_name, '') or '')
            col = COL_B_START + i
            if col_name in chg_flds and change not in (ADDED, REMOVED):
                # Show the change prominently
                cell_bg = bg
                cell_fg = fg
                bold    = True
            else:
                ig_color = row_b.get('_color', 'FFFFFF')
                ig_map   = {'FFF2CC': _C['ig_yellow'], 'FFCCCC': _C['ig_red']}
                cell_bg  = ig_map.get(ig_color, _C['white']) if change == UNCHANGED else bg
                cell_fg  = '404040'
                bold     = False
            c = ws.cell(row=data_row, column=col, value=val)
            c.font      = _font(bold=bold, size=8, color=f'FF{cell_fg.lstrip("F")}' if not cell_fg.startswith('FF') else cell_fg)
            c.fill      = _fill(cell_bg)
            c.alignment = _align(wrap=True)
            c.border    = _border_thin()

        # What Changed summary
        if chg_flds:
            lines = []
            for field, (va, vb) in chg_flds.items():
                if va and vb:
                    lines.append(f'{field}:\n  A: {va[:60]}\n  B: {vb[:60]}')
                elif vb:
                    lines.append(f'{field}: (new) {vb[:60]}')
                else:
                    lines.append(f'{field}: (removed)')
            summary_text = '\n'.join(lines)
        else:
            summary_text = '—'

        wc = ws.cell(row=data_row, column=COL_WHAT, value=summary_text)
        wc.font      = _font(size=7, color=fg)
        wc.fill      = _fill(bg)
        wc.alignment = _align(wrap=True)
        wc.border    = _border_thin()

        # ISO Definition — prefer B (NPC) if available, else A (EPC)
        iso_def = str((row_b.get('ISO Definition') or row_a.get('ISO Definition') or '')).strip()
        xml_tag = str((row_b.get('XML Tag') or row_a.get('XML Tag') or '')).strip()
        # If they differ between A and B, show both
        iso_def_a = str(row_a.get('ISO Definition') or '').strip()
        iso_def_b = str(row_b.get('ISO Definition') or '').strip()
        if iso_def_a and iso_def_b and iso_def_a != iso_def_b:
            iso_def = 'A: ' + iso_def_a + '\nB: ' + iso_def_b
        xml_tag_a = str(row_a.get('XML Tag') or '').strip()
        xml_tag_b = str(row_b.get('XML Tag') or '').strip()
        if xml_tag_a and xml_tag_b and xml_tag_a != xml_tag_b:
            xml_tag = f'A: {xml_tag_a} / B: {xml_tag_b}'

        ci = ws.cell(row=data_row, column=COL_ISO_DEF, value=iso_def)
        ci.font      = _font(size=8, color='FF333333')
        ci.fill      = _fill('FFF5F5F5' if change == UNCHANGED else 'FFFAFAFA')
        ci.alignment = _align(wrap=True)
        ci.border    = _border_thin()

        ct = ws.cell(row=data_row, column=COL_XML_TAG, value=xml_tag)
        ct.font      = _font(bold=True, size=8, color='FF1F3864')
        ct.fill      = _fill('FFF5F5F5' if change == UNCHANGED else 'FFFAFAFA')
        ct.alignment = _align(h='center', wrap=False)
        ct.border    = _border_thin()

        # Business Impact
        impact_text = _business_impact(change, chg_flds, row_a, row_b) if change != UNCHANGED else ''
        impact_fg = 'FF1A5C38' if change != UNCHANGED else _C['unchanged_txt']
        impact_bg = 'FFEBF5ED' if change not in (UNCHANGED,) else _C['unchanged']
        if change == REMOVED:
            impact_bg = 'FFFFF0ED'
            impact_fg = _C['removed_txt']
        elif change == ADDED:
            impact_bg = 'FFF0FFF4'

        cb = ws.cell(row=data_row, column=COL_IMPACT, value=impact_text)
        cb.font      = _font(size=8, color=impact_fg)
        cb.fill      = _fill(impact_bg)
        cb.alignment = _align(wrap=True)
        cb.border    = _border_thin()

        ws.row_dimensions[data_row].height = 60
        if change == UNCHANGED:
            unchanged_rows.append(data_row)

        data_row += 1

    # Group unchanged rows so they can be collapsed
    if unchanged_rows:
        # Find contiguous blocks and group them
        groups = []
        start = unchanged_rows[0]
        prev  = unchanged_rows[0]
        for r in unchanged_rows[1:]:
            if r == prev + 1:
                prev = r
            else:
                groups.append((start, prev))
                start = prev = r
        groups.append((start, prev))
        for g_start, g_end in groups:
            ws.row_dimensions.group(g_start, g_end, outline_level=1, hidden=True)

    # ── Column widths ──────────────────────────────────────────────────────────
    widths = {
        'Index': 7, 'XPath': 40, 'Element Name': 22,
        'Multiplicity': 10, 'SEPA/NPC Length': 12,
        'SEPA/NPC Usage Rules': 35, 'SEPA/NPC Rulebook': 28,
        'SEPA/NPC Format Rules': 25, 'SEPA/NPC Code Restrictions': 22,
    }
    ws.column_dimensions[get_column_letter(COL_CHANGE)].width = 14
    for i, col_name in enumerate(_SHOW_COLS):
        w = widths.get(col_name, 20)
        ws.column_dimensions[get_column_letter(COL_A_START + i)].width = w
        ws.column_dimensions[get_column_letter(COL_B_START + i)].width = w
    ws.column_dimensions[get_column_letter(COL_WHAT)].width = 42
    ws.column_dimensions[get_column_letter(COL_ISO_DEF)].width = 55
    ws.column_dimensions[get_column_letter(COL_XML_TAG)].width = 14
    ws.column_dimensions[get_column_letter(COL_IMPACT)].width = 65

    return data_row - 5  # total rows written


def _write_summary_sheet(ws, sheet_stats: list, label_a: str, label_b: str):
    """Write the Summary sheet."""

    total_cols = 10

    # Title
    t = ws['A1']
    t.value     = f'IG Diff Summary  —  {label_a}  vs  {label_b}'
    ws.merge_cells(f'A1:{get_column_letter(total_cols)}1')
    t.font      = _font(bold=True, size=14, color=_C['white'])
    t.fill      = _fill(_C['navy'])
    t.alignment = _align(h='left', v='center')
    ws.row_dimensions[1].height = 26

    # Stats boxes row 2
    totals = {'added': 0, 'removed': 0, 'status': 0, 'rules': 0,
              'both': 0, 'unchanged': 0, 'sheets': len(sheet_stats)}
    for s in sheet_stats:
        for k in ('added', 'removed', 'status', 'rules', 'both', 'unchanged'):
            totals[k] += s.get(k, 0)

    boxes = [
        (ADDED,     'Added',     totals['added']),
        (REMOVED,   'Removed',   totals['removed']),
        (STATUS,    'Status Δ',  totals['status']),
        (RULES,     'Rules Δ',   totals['rules']),
        (BOTH,      'Both Δ',    totals['both']),
        (UNCHANGED, 'Same',      totals['unchanged']),
    ]
    box_w = total_cols // len(boxes)
    for i, (cat, lbl, cnt) in enumerate(boxes):
        col = i * box_w + 1
        end = min((i + 1) * box_w, total_cols)
        if i == len(boxes) - 1:
            end = total_cols
        c = ws.cell(row=2, column=col, value=f'{cnt:,}\n{lbl}')
        if col != end:
            ws.merge_cells(f'{get_column_letter(col)}2:{get_column_letter(end)}2')
        c.font      = _font(bold=True, size=11, color=_DIFF_FG[cat])
        c.fill      = _fill(_DIFF_BG[cat])
        c.alignment = _align(h='center', v='center', wrap=True)
        c.border    = _border_thick()
    ws.row_dimensions[2].height = 36

    # Column headers row 3
    headers = ['Message Sheet', 'Message ID', 'Section Label',
               '🟢 Added', '🔴 Removed', '🟡 Status Δ',
               '🔵 Rules Δ', '🟠 Both Δ', '⬜ Same', '∑ Changes']
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=3, column=ci, value=h)
        c.font      = _font(bold=True, size=9, color=_C['white'])
        c.fill      = _fill(_C['blue'])
        c.alignment = _align(h='center', v='center', wrap=True)
        c.border    = _border_thin()
    ws.row_dimensions[3].height = 30
    ws.freeze_panes = 'A4'

    # Data rows
    for ri, s in enumerate(sheet_stats, 4):
        total_chg = s['added'] + s['removed'] + s['status'] + s['rules'] + s['both']
        row_vals = [
            s['sheet'], s['message'], s['label'],
            s['added'], s['removed'], s['status'],
            s['rules'], s['both'], s['unchanged'], total_chg,
        ]
        bg = _C['lt_gray'] if ri % 2 == 0 else _C['white']
        for ci, val in enumerate(row_vals, 1):
            c = ws.cell(row=ri, column=ci, value=val)
            c.font      = _font(size=9)
            c.alignment = _align(h='center' if ci > 3 else 'left')
            c.border    = _border_thin()
            # Colour the change-count cells
            cat_map = {4: ADDED, 5: REMOVED, 6: STATUS, 7: RULES, 8: BOTH, 9: UNCHANGED}
            if ci in cat_map and isinstance(val, int) and val > 0:
                c.fill = _fill(_DIFF_BG[cat_map[ci]])
                c.font = _font(bold=True, size=9, color=_DIFF_FG[cat_map[ci]])
            elif ci == 10 and isinstance(val, int) and val > 0:
                c.fill = _fill(_C['lt_gray'])
                c.font = _font(bold=True, size=9)
            else:
                c.fill = _fill(bg)
        ws.row_dimensions[ri].height = 18

    # Column widths
    col_widths = [28, 20, 42, 10, 10, 10, 10, 10, 10, 12]
    for ci, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w


# ── Public API ────────────────────────────────────────────────────────────────

def diff_ig(
    path_a: str,
    path_b: str,
    output_path: str,
    label_a: str = 'File A',
    label_b: str = 'File B',
    include_unchanged: bool = True,
) -> dict:
    """
    Compare two IG Extractor Excel files and write a diff workbook.

    Parameters
    ----------
    path_a          : Path to first IG Excel (e.g. EPC)
    path_b          : Path to second IG Excel (e.g. NPC)
    output_path     : Where to save the diff workbook
    label_a/label_b : Display labels (shown in headers)
    include_unchanged: Whether to include unchanged rows (grouped/hidden)

    Returns
    -------
    dict with keys: total_changes, sheets, messages_compared
    """
    ig_a = _read_ig_workbook(path_a)
    ig_b = _read_ig_workbook(path_b)

    def _msg_key(sheet_name: str) -> str:
        """
        Normalise sheet name to a matching key.
        Strip duplicate-instance suffixes (_2, _3, _4) but keep version numbers.
        e.g. pacs_008_001_08   → pacs_008_001_08
             pacs_028_001_03_2 → pacs_028_001_03
        """
        # Only strip a trailing _N if N is a single digit (duplicate instance marker)
        base = re.sub(r'_([2-9])$', '', sheet_name)
        return base.lower()

    # Build mapping: key → list of sheet names (may have duplicates like _2, _3)
    def _sheets_map(ig_data):
        mapping = {}
        for k in ig_data['sheets']:
            key = _msg_key(k)
            if key not in mapping:
                mapping[key] = k   # use first occurrence as canonical
        return mapping

    sheets_a = _sheets_map(ig_a)
    sheets_b = _sheets_map(ig_b)

    all_keys = sorted(set(sheets_a) | set(sheets_b))

    wb_out = Workbook()
    wb_out.remove(wb_out.active)  # remove default sheet

    summary_ws = wb_out.create_sheet('Summary', 0)
    sheet_stats = []
    total_changes = 0
    messages_compared = []

    for key in all_keys:
        name_a = sheets_a.get(key)
        name_b = sheets_b.get(key)

        rows_a = ig_a['sheets'][name_a]['rows'] if name_a else {}
        rows_b = ig_b['sheets'][name_b]['rows'] if name_b else {}

        # Message metadata from summary
        def _meta(ig_data, sheet_name):
            for s in ig_data.get('summary', []):
                if s['sheet'] == sheet_name:
                    return s['message'], s['label']
            if sheet_name:
                return sheet_name.replace('_', '.'), sheet_name
            return key.replace('_', '.'), key

        msg_a, lbl_a = _meta(ig_a, name_a)
        msg_b, lbl_b = _meta(ig_b, name_b)
        message = msg_a or msg_b
        label   = lbl_a or lbl_b

        diff_records = _diff_sheets(rows_a, rows_b)
        if not include_unchanged:
            diff_records = [r for r in diff_records if r['change'] != UNCHANGED]

        # Sheet name: use key, max 31 chars (Excel limit)
        ws_name = key[:31]
        # Avoid duplicate sheet names
        existing = [s.title for s in wb_out.worksheets]
        base = ws_name
        counter = 2
        while ws_name in existing:
            ws_name = f'{base[:28]}_{counter}'
            counter += 1

        ws = wb_out.create_sheet(ws_name)
        _write_diff_sheet(ws, diff_records, label_a, label_b, message, label)

        # Tally stats
        stats = {cat: sum(1 for r in diff_records if r['change'] == cat)
                 for cat in (ADDED, REMOVED, STATUS, RULES, BOTH, UNCHANGED)}
        sheet_total = stats[ADDED] + stats[REMOVED] + stats[STATUS] + stats[RULES] + stats[BOTH]
        total_changes += sheet_total

        sheet_stats.append({
            'sheet':     ws_name,
            'message':   message,
            'label':     label,
            'added':     stats[ADDED],
            'removed':   stats[REMOVED],
            'status':    stats[STATUS],
            'rules':     stats[RULES],
            'both':      stats[BOTH],
            'unchanged': stats[UNCHANGED],
        })
        messages_compared.append(message)

    _write_summary_sheet(summary_ws, sheet_stats, label_a, label_b)

    os.makedirs(os.path.dirname(os.path.abspath(output_path)), exist_ok=True)
    wb_out.save(output_path)

    return {
        'total_changes':    total_changes,
        'sheets':           sheet_stats,
        'messages_compared': list(dict.fromkeys(messages_compared)),
    }
