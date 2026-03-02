#!/usr/bin/env python3
"""
YAML / JSON Explorer – Interactive Structure Analyser
══════════════════════════════════════════════════════
Features:
  • Interactive collapsible tree with type colour-coding
  • JSONPath per node (click to copy)
  • Right-hand detail panel – value, type, path, child count
  • Statistics panel – type distribution heatmap, largest arrays, top keys
  • Search + filter by type toolbar
  • Excel flat-structure report
"""

import sys, os, json
from pathlib import Path

try:
    import yaml as _yaml
    HAS_YAML = True
except ImportError:
    HAS_YAML = False

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ── Type metadata ─────────────────────────────────────────────────────────────

TYPE_COLOR = {
    'object':  '#4f8ef7',
    'array':   '#34d997',
    'string':  '#fbbf24',
    'integer': '#fb923c',
    'number':  '#fb923c',
    'boolean': '#a78bfa',
    'null':    '#f87171',
    'unknown': '#6b7280',
}

TYPE_ICON = {
    'object':  '{}',
    'array':   '[]',
    'string':  '"…"',
    'integer': '123',
    'number':  '1.0',
    'boolean': 'T/F',
    'null':    '∅',
    'unknown': '?',
}


# ── Parser ────────────────────────────────────────────────────────────────────

_counter = [0]


def _nid():
    _counter[0] += 1
    return f"n{_counter[0]}"


def _infer_type(v):
    if v is None:           return 'null'
    if isinstance(v, bool): return 'boolean'
    if isinstance(v, int):  return 'integer'
    if isinstance(v, float):return 'number'
    if isinstance(v, str):  return 'string'
    if isinstance(v, list): return 'array'
    if isinstance(v, dict): return 'object'
    return 'unknown'


def _trunc(v, n=100):
    s = str(v)
    return s[:n] + '…' if len(s) > n else s


def _build(key, value, path, depth):
    t    = _infer_type(value)
    node = {
        'id':       _nid(),
        'key':      str(key),
        'type':     t,
        'path':     path,
        'depth':    depth,
        'children': [],
        'value':    '',
        'extra':    '',
    }
    if t == 'object':
        node['extra'] = f'{len(value)} keys'
        for k, v in value.items():
            cp = f'{path}.{k}' if path else str(k)
            node['children'].append(_build(k, v, cp, depth + 1))
    elif t == 'array':
        node['extra'] = f'{len(value)} items'
        for i, v in enumerate(value):
            node['children'].append(_build(f'[{i}]', v, f'{path}[{i}]', depth + 1))
    else:
        node['value'] = _trunc(value)
    return node


def parse_file(file_path):
    """Parse a YAML or JSON file. Returns (roots, stats, raw_text)."""
    _counter[0] = 0
    p        = Path(file_path)
    raw_text = p.read_text(encoding='utf-8', errors='replace')
    ext      = p.suffix.lower()

    if ext in ('.yaml', '.yml'):
        if not HAS_YAML:
            raise ImportError('PyYAML not installed. Run: pip install pyyaml')
        data = _yaml.safe_load(raw_text)
    else:
        data = json.loads(raw_text)

    if isinstance(data, dict):
        roots = [_build(k, v, k, 0) for k, v in data.items()]
    elif isinstance(data, list):
        roots = [_build(f'[{i}]', v, f'[{i}]', 0) for i, v in enumerate(data)]
    else:
        roots = [_build('root', data, 'root', 0)]

    def flat(nodes):
        for n in nodes:
            yield n
            yield from flat(n['children'])

    all_nodes = list(flat(roots))

    type_cnts = {}
    for n in all_nodes:
        type_cnts[n['type']] = type_cnts.get(n['type'], 0) + 1

    arr_sizes = sorted(
        [{'key': n['key'], 'path': n['path'],
          'size': int(n['extra'].split()[0])}
         for n in all_nodes if n['type'] == 'array'],
        key=lambda x: -x['size']
    )[:20]

    key_freq = {}
    for n in all_nodes:
        k = n['key']
        if not k.startswith('['):
            key_freq[k] = key_freq.get(k, 0) + 1
    top_keys = sorted(key_freq.items(), key=lambda x: -x[1])[:20]

    stats = {
        'total_nodes':   len(all_nodes),
        'total_leaves':  sum(1 for n in all_nodes if not n['children']),
        'total_objects': type_cnts.get('object', 0),
        'total_arrays':  type_cnts.get('array', 0),
        'max_depth':     max((n['depth'] for n in all_nodes), default=0),
        'type_counts':   type_cnts,
        'array_sizes':   arr_sizes,
        'top_keys':      top_keys,
        'file_type':     'YAML' if ext in ('.yaml', '.yml') else 'JSON',
    }
    return roots, stats, raw_text


# ── HTML tree renderer ────────────────────────────────────────────────────────

def _esc(s):
    return str(s).replace('&', '&amp;').replace('<', '&lt;').replace('>', '&gt;').replace('"', '&quot;').replace("'", '&#39;')


def render_node(node, depth=0):
    nid      = node['id']
    key      = node['key']
    ntype    = node['type']
    kids     = node['children']
    has_kids = bool(kids)
    value    = node['value']
    extra    = node['extra']
    path     = node['path']

    col  = TYPE_COLOR.get(ntype, '#6b7280')
    icon = TYPE_ICON.get(ntype, '?')

    data_json = json.dumps({
        'id': nid, 'key': key, 'type': ntype,
        'path': path, 'depth': depth,
        'value': value, 'extra': extra,
        'child_count': len(kids),
    }, ensure_ascii=False).replace("'", '&#39;')

    toggle  = '<span class="tog">▶</span>' if has_kids else '<span class="tog leaf">·</span>'
    tbadge  = f'<span class="tbadge" style="color:{col};border-color:{col}44">{icon} {ntype}</span>'

    val_part = ''
    if value and not has_kids:
        ve = _esc(value)
        val_part = f'<span class="nval" title="{ve}">{ve}</span>'
    elif extra:
        val_part = f'<span class="nextra">{extra}</span>'

    click = 'onclick="toggleNode(this)"' if has_kids else 'onclick="selectNode(this)"'

    h = (f'<li class="node {"has-ch" if has_kids else "leaf"}" id="node-{nid}" data-depth="{depth}" data-type="{ntype}">'
         f'<div class="nrow" data-nid="{nid}" data-json=\'{data_json}\' {click}>'
         f'{toggle} '
         f'<span class="nkey">{_esc(key)}</span>'
         f'{tbadge}'
         f'{val_part}'
         f'</div>')
    if has_kids:
        h += '<ul class="kids" style="display:none">'
        for ch in kids:
            h += render_node(ch, depth + 1)
        h += '</ul>'
    h += '</li>'
    return h


# ── HTML generator ────────────────────────────────────────────────────────────

def generate_html(roots, stats, raw_text, file_name, out_path):
    tree_html = ''.join(render_node(n) for n in roots)
    stats_j   = json.dumps(stats, ensure_ascii=False)
    file_type = stats['file_type']

    HTML = f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<title>{file_type} Explorer – {_esc(file_name)}</title>
<style>
@import url('https://fonts.googleapis.com/css2?family=JetBrains+Mono:wght@400;600&family=Sora:wght@300;400;600;700&display=swap');

:root {{
  --bg0:#0a0e1a; --bg1:#0f1526; --bg2:#151d35; --bg3:#1c2645;
  --border:#243060; --border2:#2e3d7a;
  --txt:#c8d4f0; --txt2:#8a9bc8; --txt3:#4a5a8a;
  --blue:#4f8ef7; --cyan:#38c4e8; --green:#34d997;
  --purple:#a78bfa; --orange:#fb923c; --yellow:#fbbf24; --red:#f87171;
  --obj:#4f8ef7; --arr:#34d997; --str:#fbbf24; --num:#fb923c; --bool:#a78bfa; --null:#f87171;
}}
*{{box-sizing:border-box;margin:0;padding:0}}
html,body{{height:100%;overflow:hidden;font-family:'Sora',sans-serif;background:var(--bg0);color:var(--txt);font-size:13px}}

/* ── Layout ── */
#app{{display:flex;flex-direction:column;height:100vh}}
#header{{
  background:linear-gradient(135deg,var(--bg2),var(--bg1));
  border-bottom:1px solid var(--border);
  padding:10px 20px;display:flex;align-items:center;gap:16px;flex-shrink:0
}}
#header h1{{font-size:1.1rem;font-weight:700;color:var(--cyan);letter-spacing:.03em}}
#header .sub{{color:var(--txt2);font-size:.8rem;font-family:'JetBrains Mono',monospace}}
.hbadge{{background:var(--bg3);border:1px solid var(--border);border-radius:6px;padding:3px 10px;font-size:.75rem;color:var(--txt2)}}
.hbadge b{{color:var(--cyan)}}

/* ── Tab bar ── */
#tabs{{display:flex;background:var(--bg1);border-bottom:1px solid var(--border);flex-shrink:0;gap:2px;padding:0 16px}}
.tab{{padding:8px 18px;cursor:pointer;font-size:.8rem;font-weight:600;color:var(--txt3);border-bottom:2px solid transparent;transition:all .15s;letter-spacing:.04em;text-transform:uppercase}}
.tab:hover{{color:var(--txt)}}
.tab.active{{color:var(--cyan);border-bottom-color:var(--cyan)}}

/* ── Toolbar ── */
#toolbar{{background:var(--bg2);border-bottom:1px solid var(--border);padding:8px 16px;display:flex;gap:8px;align-items:center;flex-shrink:0;flex-wrap:wrap}}
.tbtn{{background:var(--bg3);color:var(--txt);border:1px solid var(--border);padding:5px 12px;border-radius:6px;cursor:pointer;font-size:.76rem;font-family:'Sora',sans-serif;transition:all .15s}}
.tbtn:hover{{border-color:var(--cyan);color:var(--cyan);background:rgba(56,196,232,.08)}}
.tbtn.active-filter{{border-color:var(--cyan);color:var(--cyan);background:rgba(56,196,232,.12)}}
#search{{background:var(--bg0);color:var(--txt);border:1px solid var(--border);padding:5px 12px;border-radius:6px;font-size:.76rem;width:240px;outline:none;font-family:'Sora',sans-serif}}
#search:focus{{border-color:var(--cyan)}}
#searchCnt{{color:var(--txt3);font-size:.75rem}}
.sep{{width:1px;height:20px;background:var(--border);margin:0 4px}}

/* ── Views ── */
#views{{flex:1;overflow:hidden;display:flex}}

/* Tree + Detail split */
#tree-view{{display:flex;width:100%;height:100%}}
#tree-panel{{flex:1;overflow:auto;padding:12px 6px 12px 14px;min-width:0}}
#detail-panel{{width:360px;border-left:1px solid var(--border);display:flex;flex-direction:column;background:var(--bg1);flex-shrink:0}}
#detail-tabs{{display:flex;border-bottom:1px solid var(--border)}}
.dtab{{padding:7px 14px;cursor:pointer;font-size:.74rem;font-weight:600;color:var(--txt3);border-bottom:2px solid transparent;text-transform:uppercase;letter-spacing:.04em;transition:all .15s}}
.dtab:hover{{color:var(--txt)}}
.dtab.active{{color:var(--cyan);border-bottom-color:var(--cyan)}}
#detail-content{{flex:1;overflow:auto;padding:14px}}

/* ── Tree nodes ── */
ul{{list-style:none;padding-left:18px}}
ul.root-ul{{padding-left:0}}
.node{{margin:1px 0}}
.nrow{{display:flex;align-items:center;gap:5px;padding:3px 8px;border-radius:5px;cursor:pointer;flex-wrap:nowrap;transition:background .1s;position:relative;min-height:26px}}
.nrow:hover{{background:var(--bg2)}}
.nrow.selected{{background:rgba(56,196,232,.1);border-left:2px solid var(--cyan);padding-left:6px}}
.nrow.hl{{background:rgba(251,191,36,.08);border-left:2px solid var(--yellow)}}
.tog{{width:12px;flex-shrink:0;color:var(--txt3);font-size:.7rem;user-select:none}}
.tog.leaf{{color:var(--bg3)}}
.nkey{{color:#79c0ff;font-weight:600;font-size:.83rem;font-family:'JetBrains Mono',monospace;flex-shrink:0}}
.tbadge{{font-size:.67rem;padding:1px 7px;border-radius:10px;font-weight:700;border:1px solid;flex-shrink:0;letter-spacing:.03em}}
.nval{{font-size:.75rem;color:var(--txt2);font-family:'JetBrains Mono',monospace;overflow:hidden;text-overflow:ellipsis;white-space:nowrap;max-width:320px;flex-shrink:1}}
.nextra{{font-size:.72rem;color:var(--txt3);font-family:'JetBrains Mono',monospace;flex-shrink:0}}
.kids{{border-left:1px dashed var(--border2);margin-left:8px}}

/* ── Detail panel ── */
.dp-section{{margin-bottom:16px}}
.dp-label{{font-size:.67rem;text-transform:uppercase;letter-spacing:.08em;color:var(--txt3);margin-bottom:6px;font-weight:700}}
.dp-key{{font-size:1.05rem;font-weight:700;color:var(--cyan);font-family:'JetBrains Mono',monospace;margin-bottom:4px}}
.dp-type-chip{{display:inline-block;padding:2px 10px;border-radius:10px;font-size:.72rem;font-weight:700;margin-bottom:10px;border:1px solid}}
.dp-path{{font-family:'JetBrains Mono',monospace;font-size:.72rem;color:var(--cyan);background:var(--bg0);padding:6px 10px;border-radius:5px;border:1px solid var(--border);word-break:break-all;cursor:pointer;position:relative}}
.dp-path:hover::after{{content:'📋 copied!';position:absolute;right:8px;top:6px;font-size:.65rem;color:var(--green)}}
.dp-value{{font-family:'JetBrains Mono',monospace;font-size:.8rem;color:var(--txt);background:var(--bg0);padding:8px 10px;border-radius:5px;border:1px solid var(--border);word-break:break-all;white-space:pre-wrap;max-height:200px;overflow:auto}}
.prop-grid{{display:grid;grid-template-columns:1fr 1fr;gap:6px}}
.prop-card{{background:var(--bg0);border:1px solid var(--border);border-radius:6px;padding:7px 10px}}
.pk{{font-size:.65rem;color:var(--txt3);text-transform:uppercase;letter-spacing:.06em}}
.pv{{font-size:.8rem;color:var(--txt);font-family:'JetBrains Mono',monospace}}

/* ── Stats panel ── */
#stats-view{{width:100%;height:100%;overflow:auto;padding:20px;display:none}}
.stat-grid{{display:grid;grid-template-columns:repeat(auto-fill,minmax(150px,1fr));gap:12px;margin-bottom:24px}}
.stat-card{{background:var(--bg2);border:1px solid var(--border);border-radius:10px;padding:16px;text-align:center}}
.stat-big{{font-size:2rem;font-weight:700;color:var(--cyan)}}
.stat-lbl{{font-size:.75rem;color:var(--txt2);margin-top:4px;text-transform:uppercase;letter-spacing:.06em}}
.heatmap-row{{display:flex;align-items:center;gap:10px;margin-bottom:6px;font-size:.8rem}}
.hm-bar{{height:14px;border-radius:4px;min-width:4px;transition:width .3s}}
.hm-name{{font-family:'JetBrains Mono',monospace;color:var(--cyan);width:220px;overflow:hidden;text-overflow:ellipsis;white-space:nowrap;flex-shrink:0}}
.hm-count{{color:var(--txt3);width:50px;text-align:right;flex-shrink:0}}
.section-title{{font-size:.9rem;font-weight:700;color:var(--txt);margin-bottom:12px;letter-spacing:.04em;border-bottom:1px solid var(--border);padding-bottom:8px}}

/* ── Scrollbar ── */
::-webkit-scrollbar{{width:7px;height:7px}}
::-webkit-scrollbar-track{{background:var(--bg0)}}
::-webkit-scrollbar-thumb{{background:var(--border2);border-radius:4px}}
</style>
</head>
<body>
<div id="app">

<!-- Header -->
<div id="header">
  <div>
    <h1>◈ {file_type} Explorer</h1>
    <div class="sub">{_esc(file_name)}</div>
  </div>
  <span class="hbadge">Nodes: <b id="hdr-nodes">–</b></span>
  <span class="hbadge">Objects: <b id="hdr-obj">–</b></span>
  <span class="hbadge">Arrays: <b id="hdr-arr">–</b></span>
  <span class="hbadge">Max depth: <b id="hdr-depth">–</b></span>
</div>

<!-- Tab bar -->
<div id="tabs">
  <div class="tab active" onclick="switchTab('tree')">🌲 Tree</div>
  <div class="tab" onclick="switchTab('stats')">📊 Statistics</div>
</div>

<!-- Toolbar -->
<div id="toolbar">
  <button class="tbtn" onclick="expandAll()">⊞ Expand All</button>
  <button class="tbtn" onclick="collapseAll()">⊟ Collapse All</button>
  <button class="tbtn" onclick="expandLevel(1)">L1</button>
  <button class="tbtn" onclick="expandLevel(2)">L2</button>
  <button class="tbtn" onclick="expandLevel(3)">L3</button>
  <div class="sep"></div>
  <button class="tbtn" id="flt-all"     onclick="filterType('all',this)">All</button>
  <button class="tbtn" id="flt-object"  onclick="filterType('object',this)">Objects</button>
  <button class="tbtn" id="flt-array"   onclick="filterType('array',this)">Arrays</button>
  <button class="tbtn" id="flt-string"  onclick="filterType('string',this)">Strings</button>
  <button class="tbtn" id="flt-number"  onclick="filterType('number',this)">Numbers</button>
  <button class="tbtn" id="flt-boolean" onclick="filterType('boolean',this)">Booleans</button>
  <button class="tbtn" id="flt-null"    onclick="filterType('null',this)">Nulls</button>
  <div class="sep"></div>
  <input type="text" id="search" placeholder="🔍  Search key / value / path…" oninput="doSearch(this.value)">
  <span id="searchCnt"></span>
</div>

<!-- Views -->
<div id="views">

  <!-- TREE VIEW -->
  <div id="tree-view">
    <div id="tree-panel">
      <ul class="root-ul">{tree_html}</ul>
    </div>
    <!-- Detail right panel -->
    <div id="detail-panel">
      <div id="detail-tabs">
        <div class="dtab active" onclick="switchDTab('info')">Info</div>
        <div class="dtab" onclick="switchDTab('value')">Value</div>
        <div class="dtab" onclick="switchDTab('path')">Path</div>
      </div>
      <div id="detail-content">
        <div id="dtab-info"><p style="color:var(--txt3);font-size:.82rem">Click any node to inspect it.</p></div>
        <div id="dtab-value" style="display:none"><p style="color:var(--txt3);font-size:.82rem">Select a leaf node to see its value.</p></div>
        <div id="dtab-path"  style="display:none"><p style="color:var(--txt3);font-size:.82rem">Select a node to see its JSONPath.</p></div>
      </div>
    </div>
  </div>

  <!-- STATS VIEW -->
  <div id="stats-view">
    <div class="stat-grid" id="stat-cards"></div>
    <div class="section-title">Type Distribution</div>
    <div id="type-dist" style="margin-bottom:24px"></div>
    <div class="section-title">Largest Arrays (Top 20)</div>
    <div id="arr-sizes" style="margin-bottom:24px"></div>
    <div class="section-title">Most Frequent Keys (Top 20)</div>
    <div id="top-keys"></div>
  </div>

</div><!-- /views -->
</div><!-- /app -->

<script>
// ── Data ──────────────────────────────────────────────────────────────────────
const STATS = {stats_j};
const TYPE_COLOR = {json.dumps(TYPE_COLOR)};

// ── Boot ──────────────────────────────────────────────────────────────────────
document.getElementById('hdr-nodes').textContent = STATS.total_nodes;
document.getElementById('hdr-obj').textContent   = STATS.total_objects;
document.getElementById('hdr-arr').textContent   = STATS.total_arrays;
document.getElementById('hdr-depth').textContent = STATS.max_depth;

// ── Tab switching ─────────────────────────────────────────────────────────────
function switchTab(name) {{
  document.querySelectorAll('.tab').forEach((t,i) => {{
    t.classList.toggle('active', ['tree','stats'][i] === name);
  }});
  document.getElementById('toolbar').style.display   = name === 'tree'  ? 'flex'  : 'none';
  document.getElementById('tree-view').style.display  = name === 'tree'  ? 'flex'  : 'none';
  document.getElementById('stats-view').style.display = name === 'stats' ? 'block' : 'none';
  if (name === 'stats') initStats();
}}

// ── Detail sub-tab ────────────────────────────────────────────────────────────
function switchDTab(name) {{
  document.querySelectorAll('.dtab').forEach((t,i) => {{
    t.classList.toggle('active', ['info','value','path'][i] === name);
  }});
  ['info','value','path'].forEach(n => {{
    document.getElementById('dtab-' + n).style.display = n === name ? 'block' : 'none';
  }});
}}

// ── Tree toggle / select ──────────────────────────────────────────────────────
function toggleNode(row) {{
  const li  = row.parentElement;
  const ul  = li.querySelector(':scope > ul');
  const tog = row.querySelector('.tog');
  if (!ul) return;
  const open = ul.style.display !== 'none';
  ul.style.display = open ? 'none' : 'block';
  tog.textContent  = open ? '▶' : '▼';
  selectNode(row);
}}

function selectNode(row) {{
  document.querySelectorAll('.nrow.selected').forEach(r => r.classList.remove('selected'));
  row.classList.add('selected');
  try {{
    const nd = JSON.parse(row.dataset.json);
    renderDetail(nd);
  }} catch(e) {{}}
}}

// ── Detail panel ─────────────────────────────────────────────────────────────
function renderDetail(nd) {{
  const col = TYPE_COLOR[nd.type] || '#6b7280';

  // Info tab
  let info = `
    <div class="dp-section">
      <div class="dp-key">${{escH(nd.key)}}</div>
      <span class="dp-type-chip" style="background:${{col}}22;color:${{col}};border-color:${{col}}44">${{nd.type}}</span>
    </div>
    <div class="dp-section">
      <div class="dp-label">Properties</div>
      <div class="prop-grid">
        <div class="prop-card"><div class="pk">depth</div><div class="pv">${{nd.depth}}</div></div>
        <div class="prop-card"><div class="pk">type</div><div class="pv">${{nd.type}}</div></div>`;
  if (nd.child_count > 0)
    info += `<div class="prop-card"><div class="pk">children</div><div class="pv">${{nd.child_count}}</div></div>`;
  if (nd.extra)
    info += `<div class="prop-card"><div class="pk">size</div><div class="pv">${{nd.extra}}</div></div>`;
  info += `</div></div>`;

  if (nd.value !== '') {{
    info += `<div class="dp-section">
      <div class="dp-label">Value</div>
      <div class="dp-value">${{escH(nd.value)}}</div>
    </div>`;
  }}

  document.getElementById('dtab-info').innerHTML = info;

  // Value tab
  const vhtml = nd.value !== ''
    ? `<div class="dp-section"><div class="dp-label">Value</div><div class="dp-value">${{escH(nd.value)}}</div></div>`
    : nd.child_count > 0
      ? `<p style="color:var(--txt3);font-size:.82rem">This node has ${{nd.child_count}} child(ren). Expand it in the tree to see them.</p>`
      : `<p style="color:var(--txt3);font-size:.82rem">No value.</p>`;
  document.getElementById('dtab-value').innerHTML = vhtml;

  // Path tab
  const phtml = `
    <div class="dp-section">
      <div class="dp-label">JSONPath</div>
      <div class="dp-path" onclick="copyPath(this,'${{nd.path.replace(/'/g,"&apos;")}}')" title="Click to copy">
        ${{escH(nd.path)}}
      </div>
    </div>
    <div class="dp-section" style="margin-top:10px">
      <div class="dp-label">Depth</div>
      <div style="font-size:.8rem;color:var(--txt);font-family:'JetBrains Mono',monospace">${{nd.depth}}</div>
    </div>`;
  document.getElementById('dtab-path').innerHTML = phtml;
}}

function copyPath(el, path) {{
  navigator.clipboard.writeText(path).catch(() => {{}});
  el.style.background = 'rgba(52,217,151,.1)';
  setTimeout(() => el.style.background = '', 1200);
}}

function escH(s) {{
  return String(s).replace(/&/g,'&amp;').replace(/</g,'&lt;').replace(/>/g,'&gt;');
}}

// ── Tree controls ─────────────────────────────────────────────────────────────
function expandAll() {{
  document.querySelectorAll('.kids').forEach(u => u.style.display = 'block');
  document.querySelectorAll('.tog:not(.leaf)').forEach(t => t.textContent = '▼');
}}
function collapseAll() {{
  document.querySelectorAll('.kids').forEach(u => u.style.display = 'none');
  document.querySelectorAll('.tog:not(.leaf)').forEach(t => t.textContent = '▶');
}}
function expandLevel(max) {{
  collapseAll();
  document.querySelectorAll('.node[data-depth]').forEach(li => {{
    const d = parseInt(li.dataset.depth);
    if (d < max) {{
      const ul = li.querySelector(':scope > ul');
      const t  = li.querySelector('.nrow .tog');
      if (ul) {{ ul.style.display = 'block'; if (t && !t.classList.contains('leaf')) t.textContent = '▼'; }}
    }}
  }});
}}

// ── Filter by type ────────────────────────────────────────────────────────────
function filterType(type, btn) {{
  // Update button active state
  document.querySelectorAll('#toolbar .tbtn[id^="flt-"]').forEach(b => b.classList.remove('active-filter'));
  btn.classList.add('active-filter');

  document.querySelectorAll('.node').forEach(li => {{
    if (type === 'all') {{ li.style.display = ''; return; }}
    // integer and number share the 'number' filter
    const t = li.dataset.type || '';
    const match = (type === 'number') ? (t === 'number' || t === 'integer') : t === type;
    li.style.display = match ? '' : 'none';
  }});
  if (type !== 'all') _revealMatchedAncestors();
}}

// Ensure ancestor li.node and ul.kids are visible for matched nodes
function _revealMatchedAncestors() {{
  document.querySelectorAll('.node').forEach(li => {{
    if (li.style.display === 'none') return;
    let el = li.parentElement;
    while (el) {{
      if (el.tagName === 'UL' && el.classList.contains('kids')) {{
        el.style.display = 'block';
        const tog = el.previousElementSibling?.querySelector('.tog');
        if (tog && !tog.classList.contains('leaf')) tog.textContent = '▼';
      }}
      if (el.tagName === 'LI' && el.classList.contains('node')) {{
        el.style.display = '';
      }}
      el = el.parentElement;
    }}
  }});
}}

// ── Search ────────────────────────────────────────────────────────────────────
function doSearch(q) {{
  document.querySelectorAll('.nrow').forEach(r => r.classList.remove('hl'));
  if (!q) {{ document.getElementById('searchCnt').textContent = ''; return; }}
  const lq = q.toLowerCase();
  let n = 0;
  document.querySelectorAll('.nrow').forEach(row => {{
    if (row.textContent.toLowerCase().includes(lq)) {{
      row.classList.add('hl');
      // Reveal ancestors
      let el = row.parentElement;
      while (el) {{
        if (el.classList && el.classList.contains('kids')) {{
          el.style.display = 'block';
          const t = el.previousElementSibling?.querySelector('.tog');
          if (t && !t.classList.contains('leaf')) t.textContent = '▼';
        }}
        el = el.parentElement;
      }}
      n++;
    }}
  }});
  document.getElementById('searchCnt').textContent = `${{n}} match${{n !== 1 ? 'es' : ''}}`;
}}

// ── Statistics ────────────────────────────────────────────────────────────────
let _statsInited = false;
function initStats() {{
  if (_statsInited) return;
  _statsInited = true;
  const s = STATS;

  // Summary cards
  const cardData = [
    ['Total Nodes',   s.total_nodes,   'var(--cyan)'],
    ['Leaf Values',   s.total_leaves,  'var(--green)'],
    ['Objects',       s.total_objects, 'var(--blue)'],
    ['Arrays',        s.total_arrays,  'var(--green)'],
    ['Max Depth',     s.max_depth,     'var(--orange)'],
  ];
  const cg = document.getElementById('stat-cards');
  cardData.forEach(([lbl,val,col]) => {{
    cg.innerHTML += `<div class="stat-card">
      <div class="stat-big" style="color:${{col}}">${{val}}</div>
      <div class="stat-lbl">${{lbl}}</div></div>`;
  }});

  // Type distribution
  const tc = s.type_counts || {{}};
  const total = s.total_nodes || 1;
  const td = document.getElementById('type-dist');
  const COLORS = {json.dumps(TYPE_COLOR)};
  Object.entries(tc).sort((a,b) => b[1]-a[1]).forEach(([t,cnt]) => {{
    const pct = Math.round(cnt / total * 300);
    const col = COLORS[t] || '#6b7280';
    td.innerHTML += `<div class="heatmap-row">
      <div class="hm-name" title="${{t}}">${{t}}</div>
      <div class="hm-bar" style="width:${{pct}}px;background:${{col}}"></div>
      <div class="hm-count">${{cnt}} (${{Math.round(cnt/total*100)}}%)</div>
    </div>`;
  }});

  // Largest arrays
  const as = document.getElementById('arr-sizes');
  if (!s.array_sizes || !s.array_sizes.length) {{
    as.innerHTML = '<p style="color:var(--txt3);font-size:.8rem">No arrays found.</p>';
  }} else {{
    const maxSz = s.array_sizes[0].size || 1;
    s.array_sizes.forEach(a => {{
      const pct = Math.round(a.size / maxSz * 300);
      as.innerHTML += `<div class="heatmap-row">
        <div class="hm-name" title="${{a.path}}">${{a.key}} <span style="color:var(--txt3);font-size:.7rem">${{a.path}}</span></div>
        <div class="hm-bar" style="width:${{pct}}px;background:var(--green)"></div>
        <div class="hm-count">${{a.size}}</div>
      </div>`;
    }});
  }}

  // Top keys
  const tk = document.getElementById('top-keys');
  if (!s.top_keys || !s.top_keys.length) {{
    tk.innerHTML = '<p style="color:var(--txt3);font-size:.8rem">No named keys found.</p>';
  }} else {{
    const maxK = s.top_keys[0][1] || 1;
    s.top_keys.forEach(([k,cnt]) => {{
      const pct = Math.round(cnt / maxK * 300);
      tk.innerHTML += `<div class="heatmap-row">
        <div class="hm-name" title="${{k}}">${{k}}</div>
        <div class="hm-bar" style="width:${{pct}}px;background:var(--blue)"></div>
        <div class="hm-count">${{cnt}}</div>
      </div>`;
    }});
  }}
}}
</script>
</body>
</html>"""

    with open(out_path, 'w', encoding='utf-8') as f:
        f.write(HTML)
    print(f"✅ HTML written: {out_path}")


# ── Excel generator ───────────────────────────────────────────────────────────

def _flatten_nodes(nodes, rows=None):
    if rows is None:
        rows = []
    for n in nodes:
        rows.append({
            'path':        n['path'],
            'key':         n['key'],
            'type':        n['type'],
            'depth':       n['depth'],
            'value':       n['value'],
            'child_count': len(n['children']),
            'extra':       n['extra'],
        })
        _flatten_nodes(n['children'], rows)
    return rows


def generate_excel(roots, file_name, out_path):
    rows = _flatten_nodes(roots)
    wb   = Workbook()
    ws   = wb.active
    ws.title = 'Structure'

    thin   = Side(style='thin', color='D0D7DE')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    cols   = ['JSONPath', 'Key', 'Type', 'Depth', 'Value / Size']
    hfont  = Font(name='Segoe UI', bold=True, color='FFFFFF', size=10)
    hfill  = PatternFill('solid', fgColor='1B2A5E')
    halign = Alignment(horizontal='center', vertical='center', wrap_text=True)

    ws.append(cols)
    for c in range(1, len(cols) + 1):
        cell = ws.cell(1, c)
        cell.font = hfont; cell.fill = hfill
        cell.alignment = halign; cell.border = border
    ws.row_dimensions[1].height = 26
    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = f"A1:{get_column_letter(len(cols))}1"

    col_widths = {'JSONPath': 55, 'Key': 28, 'Type': 12, 'Depth': 7, 'Value / Size': 50}
    for c, k in enumerate(cols, 1):
        ws.column_dimensions[get_column_letter(c)].width = col_widths.get(k, 15)

    TYPE_FILL = {
        'object':  'EEF3FF', 'array': 'EDFFF7',
        'string':  'FFFBEB', 'integer': 'FFF4ED', 'number': 'FFF4ED',
        'boolean': 'F5F0FF', 'null':  'FFF0F0',
    }
    path_font  = Font(name='Consolas', size=9, color='1F5C8B')
    value_font = Font(name='Consolas', size=9)
    key_font   = Font(name='Consolas', size=9, bold=True)

    for ri, row in enumerate(rows, 2):
        depth     = row['depth']
        ntype     = row['type']
        val_disp  = row['value'] if row['value'] else row['extra']
        vals      = [row['path'], ('  ' * depth) + row['key'], ntype, depth, val_disp]
        ws.append(vals)
        fill_hex  = TYPE_FILL.get(ntype, 'FFFFFF')
        rfill     = PatternFill('solid', fgColor=fill_hex)
        for ci, _ in enumerate(vals, 1):
            cell = ws.cell(ri, ci)
            cell.border    = border
            cell.alignment = Alignment(vertical='center', wrap_text=(ci == 5))
            cell.fill      = rfill
            if ci == 1: cell.font = path_font
            elif ci == 2: cell.font = key_font
            else: cell.font = value_font
        ws.row_dimensions[ri].height = 15

    # Summary sheet
    ws2 = wb.create_sheet('Summary')
    ws2.column_dimensions['A'].width = 28
    ws2.column_dimensions['B'].width = 20
    ws2.append([f'YAML / JSON Explorer – {file_name}', ''])
    ws2['A1'].font = Font(name='Segoe UI', bold=True, size=12, color='1F3864')
    ws2.merge_cells('A1:B1')
    type_cnts = {}
    for r in rows:
        type_cnts[r['type']] = type_cnts.get(r['type'], 0) + 1
    summary_rows = [
        ('File', file_name),
        ('Total Nodes', len(rows)),
        ('Max Depth', max((r['depth'] for r in rows), default=0)),
    ] + [(f'  {t}', cnt) for t, cnt in sorted(type_cnts.items())]
    for i, (k, v) in enumerate(summary_rows, 3):
        ws2.cell(i, 1, k).font = Font(name='Segoe UI', bold=True, size=10)
        ws2.cell(i, 2, v).font = Font(name='Segoe UI', size=10)

    wb.save(out_path)
    print(f"✅ Excel written: {out_path}")


# ── CLI entry point ───────────────────────────────────────────────────────────

def main():
    if len(sys.argv) < 2:
        print("Usage: python yaml_json_explorer.py <file.yaml|file.json> [output_dir]")
        sys.exit(1)
    fp      = sys.argv[1]
    out_dir = sys.argv[2] if len(sys.argv) > 2 else os.path.dirname(fp) or '.'
    os.makedirs(out_dir, exist_ok=True)
    fname   = os.path.basename(fp)
    base    = os.path.splitext(fname)[0]

    print(f"📂 Parsing: {fp}")
    roots, stats, raw = parse_file(fp)
    print(f"   Nodes: {stats['total_nodes']}  |  Max depth: {stats['max_depth']}")

    generate_html(roots, stats, raw, fname,
                  os.path.join(out_dir, f"{base}_explorer.html"))
    generate_excel(roots, fname,
                   os.path.join(out_dir, f"{base}_structure.xlsx"))
    print(f"\n🎉 Done!")


if __name__ == '__main__':
    main()