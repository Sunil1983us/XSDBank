#!/usr/bin/env python3
"""
ISO 20022 Payment Message Comprehensive Analyzer
Complete metadata extraction with XSD annotation-based classification
"""

import xml.etree.ElementTree as ET
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
import argparse
from pathlib import Path

class ISO20022Analyzer:
    """Extract ALL information from ISO 20022 XSD in proper sequence"""
    
    def __init__(self, xsd_file):
        self.xsd_file = xsd_file
        self.tree = ET.parse(xsd_file)
        self.root = self.tree.getroot()
        self.ns = {'xs': 'http://www.w3.org/2001/XMLSchema'}
        self.elements = []
        self.type_definitions = {}
        self.sequence_counter = 0
        
        # Build type definitions map
        self._build_type_map()
    
    def _build_type_map(self):
        """Build a map of all type definitions"""
        for complex_type in self.root.findall('.//xs:complexType', self.ns):
            name = complex_type.get('name')
            if name:
                self.type_definitions[name] = complex_type
        
        for simple_type in self.root.findall('.//xs:simpleType', self.ns):
            name = simple_type.get('name')
            if name:
                self.type_definitions[name] = simple_type
    
    def extract_all_metadata(self):
        """Extract comprehensive metadata in proper sequence"""
        print("\n⏳ Extracting ISO 20022 metadata...")
        
        # Find root element
        root_elements = self.root.findall('./xs:element', self.ns)
        
        for root_elem in root_elements:
            elem_name = root_elem.get('name', '')
            if elem_name:
                self._process_element(root_elem, '', 0, False)
        
        print(f"   ✅ Extracted {len(self.elements)} elements")
    
    def _process_element(self, element, parent_path, level, in_choice):
        """Process an element and recurse into its type"""
        elem_name = element.get('name', element.get('ref', ''))
        if not elem_name:
            return
        
        # Remove namespace prefix
        if ':' in elem_name:
            elem_name = elem_name.split(':')[-1]
        
        # Build full path
        current_path = f"{parent_path}/{elem_name}" if parent_path else elem_name
        
        # Get attributes
        elem_type = element.get('type', '')
        if ':' in elem_type:
            elem_type = elem_type.split(':')[-1]
        
        min_occurs = element.get('minOccurs', '1')
        max_occurs = element.get('maxOccurs', '1')
        
        # Extract annotation
        annotation = self._extract_annotation(element)
        
        # Extract restrictions
        restriction_info = self._extract_restriction_from_element(element)
        
        # ISSUE 2 FIX: Classify from XSD annotation
        field_class = self._classify_field_from_xsd(element, elem_name, min_occurs, annotation)
        
        # Store element
        self.sequence_counter += 1
        element_info = {
            'sequence': self.sequence_counter,
            'level': level,
            'path': current_path,
            'element': elem_name,
            'type': elem_type,
            'min_occurs': min_occurs,
            'max_occurs': max_occurs,
            'required': min_occurs != '0',
            'in_choice': in_choice,
            'annotation': annotation.get('documentation', ''),
            'rulebook': annotation.get('rulebook', ''),
            'usage_rules': annotation.get('usage_rules', ''),
            'field_classification': field_class,
            'pattern': restriction_info.get('pattern', ''),
            'min_length': restriction_info.get('min_length', ''),
            'max_length': restriction_info.get('max_length', ''),
            'min_value': restriction_info.get('min_value', ''),
            'max_value': restriction_info.get('max_value', ''),
            'enumeration': restriction_info.get('enumeration', ''),
            'fraction_digits': restriction_info.get('fraction_digits', ''),
            'total_digits': restriction_info.get('total_digits', ''),
        }
        
        self.elements.append(element_info)
        
        # Recurse into type
        inline_complex = element.find('xs:complexType', self.ns)
        if inline_complex is not None:
            self._process_complex_type(inline_complex, current_path, level + 1, in_choice)
        elif elem_type and elem_type in self.type_definitions:
            type_def = self.type_definitions[elem_type]
            if type_def.tag.endswith('complexType'):
                self._process_complex_type(type_def, current_path, level + 1, in_choice)
    
    def _process_complex_type(self, complex_type, parent_path, level, in_choice):
        """Process complex type - maintains sequence order"""
        for child in complex_type:
            tag = child.tag.replace('{http://www.w3.org/2001/XMLSchema}', '')
            
            if tag == 'sequence':
                self._process_sequence(child, parent_path, level, in_choice)
            elif tag == 'choice':
                self._process_choice(child, parent_path, level)
            elif tag == 'complexContent':
                self._process_complex_content(child, parent_path, level, in_choice)
    
    def _process_sequence(self, sequence, parent_path, level, in_choice):
        """Process sequence - maintains order"""
        for child in sequence:
            tag = child.tag.replace('{http://www.w3.org/2001/XMLSchema}', '')
            if tag == 'element':
                self._process_element(child, parent_path, level, in_choice)
    
    def _process_choice(self, choice, parent_path, level):
        """Process choice - maintains order"""
        for child in choice:
            tag = child.tag.replace('{http://www.w3.org/2001/XMLSchema}', '')
            if tag == 'element':
                self._process_element(child, parent_path, level, in_choice=True)
    
    def _process_complex_content(self, complex_content, parent_path, level, in_choice):
        """Process complexContent"""
        for child in complex_content:
            tag = child.tag.replace('{http://www.w3.org/2001/XMLSchema}', '')
            if tag in ['extension', 'restriction']:
                for sub_child in child:
                    sub_tag = sub_child.tag.replace('{http://www.w3.org/2001/XMLSchema}', '')
                    if sub_tag == 'sequence':
                        self._process_sequence(sub_child, parent_path, level, in_choice)
                    elif sub_tag == 'choice':
                        self._process_choice(sub_child, parent_path, level)
    
    def _extract_annotation(self, element):
        """Extract all annotation information including Rulebook"""
        result = {
            'documentation': '',
            'rulebook': '',
            'usage_rules': '',
        }
        
        annotation = element.find('xs:annotation', self.ns)
        if annotation is not None:
            docs = annotation.findall('xs:documentation', self.ns)
            doc_texts = []
            for doc in docs:
                source = doc.get('source', '').strip()
                text = (doc.text or '').strip()
                
                # Separate by source
                if source == 'Rulebook':
                    result['rulebook'] = text
                elif source not in ['Yellow Field', 'White Field', '']:
                    # Other sources go to usage rules
                    if text:
                        result['usage_rules'] += f"{source}: {text} "
                elif not source and text:
                    # No source attribute - general documentation
                    doc_texts.append(text)
            
            result['documentation'] = ' | '.join(doc_texts) if doc_texts else ''
        
        return result
    
    def _classify_field_from_xsd(self, element, elem_name, min_occurs, annotation):
        """
        ISSUE 2 FIX: Read Yellow/White from XSD annotations
        Priority: XSD annotation > Heuristics > NA
        """
        # First check XSD annotation
        annotation_elem = element.find('xs:annotation', self.ns)
        if annotation_elem is not None:
            docs = annotation_elem.findall('xs:documentation', self.ns)
            for doc in docs:
                source = doc.get('source', '').strip()
                
                if source == 'Yellow Field':
                    return '🟡 Yellow (ISO 20022 Spec)'
                elif source == 'White Field':
                    return '⚪ White (ISO 20022 Spec)'
        
        # NO INFERENCE - Only use XSD annotations
        # If not in XSD, return NA
        
        return '⚫ NA (Not in XSD)'
    
    def _extract_restriction_from_element(self, element):
        """Extract restriction information"""
        result = {}
        
        simple_type = element.find('xs:simpleType', self.ns)
        if simple_type is not None:
            restriction = simple_type.find('xs:restriction', self.ns)
            if restriction is not None:
                result = self._extract_restriction_details(restriction)
        
        return result
    
    def _extract_restriction_details(self, restriction):
        """Extract all restriction details"""
        result = {}
        
        # Pattern
        pattern = restriction.find('xs:pattern', self.ns)
        if pattern is not None:
            result['pattern'] = pattern.get('value', '')
        
        # Lengths
        min_length = restriction.find('xs:minLength', self.ns)
        if min_length is not None:
            result['min_length'] = min_length.get('value', '')
        
        max_length = restriction.find('xs:maxLength', self.ns)
        if max_length is not None:
            result['max_length'] = max_length.get('value', '')
        
        length = restriction.find('xs:length', self.ns)
        if length is not None:
            val = length.get('value', '')
            result['min_length'] = val
            result['max_length'] = val
        
        # Values
        min_inclusive = restriction.find('xs:minInclusive', self.ns)
        if min_inclusive is not None:
            result['min_value'] = min_inclusive.get('value', '')
        
        max_inclusive = restriction.find('xs:maxInclusive', self.ns)
        if max_inclusive is not None:
            result['max_value'] = max_inclusive.get('value', '')
        
        # Enumeration
        enums = restriction.findall('xs:enumeration', self.ns)
        if enums:
            enum_values = [e.get('value', '') for e in enums]
            result['enumeration'] = ', '.join(enum_values[:10])
            if len(enum_values) > 10:
                result['enumeration'] += f' ... ({len(enum_values)} total)'
        
        # Digits
        fraction_digits = restriction.find('xs:fractionDigits', self.ns)
        if fraction_digits is not None:
            result['fraction_digits'] = fraction_digits.get('value', '')
        
        total_digits = restriction.find('xs:totalDigits', self.ns)
        if total_digits is not None:
            result['total_digits'] = total_digits.get('value', '')
        
        return result
    
    def generate_excel(self, output_file):
        """Generate comprehensive Excel with ALL metadata"""
        print("\n⏳ Generating Excel report...")
        
        wb = Workbook()
        ws1 = wb.active
        ws1.title = "Complete Structure"
        
        # ISSUE 3: ISO 20022 branding in headers
        headers = [
            'Seq', 'Level', 'Full Path', 'Element', 'Type',
            'Required', 'Min', 'Max', 'In Choice',
            'Field Class (ISO 20022)', 'Documentation', 'Rulebook Notes',
            'Usage Rules', 'Pattern', 'Min Length', 'Max Length',
            'Enumeration', 'Min Value', 'Max Value', 'Digits'
        ]
        
        ws1.append(headers)
        
        # Style header
        for cell in ws1[1]:
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
            cell.alignment = Alignment(horizontal='center', wrap_text=True)
        
        # Add data
        for elem in self.elements:
            row = [
                elem['sequence'],
                elem['level'],
                elem['path'],
                elem['element'],
                elem['type'],
                'Yes' if elem['required'] else 'No',
                elem['min_occurs'],
                elem['max_occurs'],
                'Yes' if elem['in_choice'] else 'No',
                elem['field_classification'],
                elem['annotation'],
                elem['rulebook'],
                elem['usage_rules'],
                elem['pattern'],
                elem['min_length'],
                elem['max_length'],
                elem['enumeration'],
                elem['min_value'],
                elem['max_value'],
                f"{elem['fraction_digits']}/{elem['total_digits']}" if elem['fraction_digits'] or elem['total_digits'] else ''
            ]
            ws1.append(row)
            
            # Highlight by classification
            row_num = ws1.max_row
            if '🟡' in elem['field_classification']:
                fill = PatternFill(start_color='FFF8DC', end_color='FFF8DC', fill_type='solid')
                for col in range(1, 21):
                    ws1.cell(row_num, col).fill = fill
        
        # Column widths
        widths = [6, 6, 60, 30, 35, 10, 6, 6, 10, 30, 50, 50, 40, 35, 10, 10, 40, 10, 10, 15]
        for i, width in enumerate(widths, 1):
            ws1.column_dimensions[get_column_letter(i)].width = width
        
        ws1.freeze_panes = 'D2'
        
        # Sheet 2: Field Classification
        ws2 = wb.create_sheet("ISO 20022 Field Types")
        ws2.append(['Classification', 'Element', 'Full Path', 'Required', 'Rulebook Notes'])
        
        for cell in ws2[1]:
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
        
        for elem in self.elements:
            ws2.append([
                elem['field_classification'],
                elem['element'],
                elem['path'],
                'Yes' if elem['required'] else 'No',
                elem['rulebook']
            ])
            
            row_num = ws2.max_row
            if '🟡' in elem['field_classification']:
                for col in [1, 2, 3, 4, 5]:
                    ws2.cell(row_num, col).fill = PatternFill(start_color='FFF8DC', end_color='FFF8DC', fill_type='solid')
        
        ws2.column_dimensions['A'].width = 30
        ws2.column_dimensions['B'].width = 35
        ws2.column_dimensions['C'].width = 60
        ws2.column_dimensions['D'].width = 10
        ws2.column_dimensions['E'].width = 50
        
        # Save
        wb.save(output_file)
        
        yellow_iso = sum(1 for e in self.elements if 'ISO 20022 Spec' in e['field_classification'] and '🟡' in e['field_classification'])
        yellow_inf = sum(1 for e in self.elements if 'Inferred' in e['field_classification'] and '🟡' in e['field_classification'])
        white_iso = sum(1 for e in self.elements if 'ISO 20022 Spec' in e['field_classification'] and '⚪' in e['field_classification'])
        na_count = sum(1 for e in self.elements if '⚫' in e['field_classification'])
        
        print(f"\n✅ ISO 20022 analysis saved: {output_file}")
        print(f"   📊 Total elements: {len(self.elements)}")
        print(f"   🟡 Yellow (ISO spec): {yellow_iso}")
        print(f"   🟡 Yellow (inferred): {yellow_inf}")
        print(f"   ⚪ White (ISO spec): {white_iso}")
        print(f"   ⚫ NA (not specified): {na_count}")
        print(f"   ✅ Sequence: XSD order maintained")
        print(f"   ✅ Paths: Complete hierarchy")


def main():
    parser = argparse.ArgumentParser(
        description='ISO 20022 Payment Message Comprehensive Analyzer',
        epilog='Part of ISO 20022 Payment Toolkit'
    )
    parser.add_argument('xsd_file', help='ISO 20022 XSD schema file')
    parser.add_argument('-o', '--output', help='Output Excel file', 
                       default='iso20022_comprehensive_analysis.xlsx')
    
    args = parser.parse_args()
    
    if not Path(args.xsd_file).exists():
        print(f"❌ Error: File '{args.xsd_file}' not found")
        return
    
    print(f"\n{'='*70}")
    print("ISO 20022 PAYMENT MESSAGE COMPREHENSIVE ANALYSIS")
    print("Field classification from XSD annotations")
    print(f"{'='*70}")
    print(f"\n📂 Schema: {args.xsd_file}")
    
    analyzer = ISO20022Analyzer(args.xsd_file)
    analyzer.extract_all_metadata()
    analyzer.generate_excel(args.output)
    
    print(f"\n{'='*70}")
    print("✅ ANALYSIS COMPLETE!")
    print(f"{'='*70}\n")


if __name__ == '__main__':
    main()
