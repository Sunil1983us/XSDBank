"""
rulebook_change_tracker.py  -  ISO 20022 Rulebook Change Tracker
Extracts the 'List of Changes' section from one or two IG PDFs
and produces a structured, colour-coded Excel workbook.

Change Type Coding:
  CHAN  Orange  - Content / Rulebook change
  CLAR  Blue    - Clarification only
  TYPO  Yellow  - Typo / layout fix

Usage:
    from rulebook_change_tracker import track_changes
    result = track_changes('EPC_v1.pdf', '/out/changes.xlsx', pdf_b='EPC_v2.pdf')
"""

import re, os
from typing import Optional
import pdfplumber
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Colours
_C = {
    'navy': 'FF1F3864', 'blue': 'FF2E74B5', 'white': 'FFFFFFFF',
    'lt_gray': 'FFF2F2F2', 'dark_gray': 'FF595959',
    'chan_bg': 'FFFCE4D6', 'chan_fg': 'FF843C0C',
    'clar_bg': 'FFD6E4F7', 'clar_fg': 'FF1F4E79',
    'typo_bg': 'FFFFF2CC', 'typo_fg': 'FF7D6608',
    'ds_bg': 'FFE2EFDA',  'ds_fg': 'FF375623',
}

_TYPE_BG    = {'CHAN': _C['chan_bg'], 'CLAR': _C['clar_bg'], 'TYPO': _C['typo_bg']}
_TYPE_FG    = {'CHAN': _C['chan_fg'], 'CLAR': _C['clar_fg'], 'TYPO': _C['typo_fg']}
_TYPE_LABEL = {
    'CHAN': '🟠 CHAN - Content Change',
    'CLAR': '🔵 CLAR - Clarification',
    'TYPO': '🟡 TYPO - Typo / Layout',
}


def _fill(h): return PatternFill('solid', fgColor=h)
def _font(bold=False, color='FF000000', size=9): return Font(name='Arial', bold=bold, size=size, color=color)
def _border():
    s = Side(style='thin', color='FFD0D0D0')
    return Border(left=s, right=s, top=s, bottom=s)
def _border_med():
    s = Side(style='medium', color='FF9E9E9E')
    return Border(left=s, right=s, top=s, bottom=s)
def _align(h='left', v='top', wrap=True): return Alignment(horizontal=h, vertical=v, wrap_text=wrap)


def _extract_metadata(pdf_path):
    with pdfplumber.open(pdf_path) as pdf:
        text = pdf.pages[0].extract_text() or ''
    meta = {'doc_number': '', 'version': '', 'date_issued': '',
            'date_effective': '', 'scheme': '', 'filename': os.path.basename(pdf_path)}
    m = re.search(r'((?:EPC|NPC)\d+-\d+)\s*/\s*(\d{4}\s+Version\s+[\d.]+)', text)
    if m:
        meta['doc_number'] = m.group(1)
        meta['version']    = m.group(2).strip()
    m = re.search(r'Date issued[:\s]+([\d]+\s+\w+\s+\d{4}|\d{4}-\d{2}-\d{2})', text, re.IGNORECASE)
    if m: meta['date_issued'] = m.group(1).strip()
    m = re.search(r'Date effective[:\s]+([\d]+\s+\w+\s+\d{4}|\d{4}-\d{2}-\d{2})', text, re.IGNORECASE)
    if m: meta['date_effective'] = m.group(1).strip()
    if re.search(r'Nordic|NPC', text): meta['scheme'] = 'NCT Inst (NPC)'
    elif re.search(r'SEPA|SCT\s*Inst', text): meta['scheme'] = 'SCT Inst (EPC)'
    return meta


def _find_changes_start(pdf):
    for i, page in enumerate(pdf.pages):
        text = page.extract_text() or ''
        if re.search(r'list of changes', text, re.IGNORECASE) and i > 5:
            return i
    return -1


def _infer_dataset(text):
    patterns = [
        (r'DS-02|Inter-PSP Payment|pacs\.008|Credit Transfer V08',  'DS-02  pacs.008  (Credit Transfer)'),
        (r'Negative Confirmation|DS-03.*Neg|pacs\.002',             'DS-03  pacs.002  (Negative Confirmation)'),
        (r'Positive Confirmation|DS-03.*Pos',                        'DS-03  pacs.002  (Positive Confirmation)'),
        (r'DS-05|Recall.*Dataset|camt\.056|Payment Cancellation',   'DS-05  camt.056  (Recall Request)'),
        (r'DS-06|Response.*Recall.*Neg|camt\.029|Resolution.*Inv',  'DS-06  camt.029  (Negative Recall Response)'),
        (r'DS-06|Response.*Recall.*Pos|pacs\.004|Payment Return',   'DS-06  pacs.004  (Positive Recall Response)'),
        (r'DS-07|Status Investigation|pacs\.028',                   'DS-07  pacs.028  (Status Investigation)'),
        (r'DS-08|Request for Recall.*Originator',                   'DS-08  camt.056  (Recall by Originator)'),
        (r'DS-09|Negative Response.*Originator',                    'DS-09  camt.029  (Negative Response - Originator Recall)'),
        (r'DS-10|Positive Response.*Originator',                    'DS-10  pacs.004  (Positive Response - Originator Recall)'),
        (r'DS-11|Status Update.*Recall',                            'DS-11  pacs.028  (Status Update - Originator Recall)'),
    ]
    for pattern, label in patterns:
        if re.search(pattern, text, re.IGNORECASE):
            return label
    return ''


def _business_note(chg_type, element_ref, description):
    d = description.lower()
    prefix = {
        'CHAN': 'Content change — may require implementation update.',
        'CLAR': 'Clarification only — no functional change expected; review for alignment.',
        'TYPO': 'Typo/layout fix — no functional impact expected.',
    }.get(chg_type, '')
    notes = []
    if re.search(r'mandatory|1\.\.1|made yellow',     d): notes.append('Field is now MANDATORY — must always be populated.')
    if re.search(r'\bmade white\b|now optional',       d): notes.append('Field is now OPTIONAL — not required.')
    if re.search(r'not permitted|forbidden|removed from ig|made red', d): notes.append('Field BLOCKED — do not include.')
    if re.search(r'new (section|attribute|field|element|sub)', d): notes.append('New element added — implement handling.')
    if re.search(r'remov(ed?|al)\b',                  d): notes.append('Element removed — stop sending for this scheme.')
    if re.search(r'length|max \d+|\d+\s*\.\.\s*\d+',  d): notes.append('Length constraint changed — revalidate field values.')
    if re.search(r'usage rule',                        d): notes.append('Usage rule updated — re-read constraint.')
    if re.search(r'format rule',                       d): notes.append('Format rule changed — check regex/pattern validation.')
    if re.search(r'amount|currency',                   d): notes.append('Monetary field impacted — review amount/currency logic.')
    if re.search(r'timestamp|datetime',                d): notes.append('Timestamp rule changed — check date/time formatting.')
    if re.search(r'address|postal',                    d): notes.append('Address handling impacted — review structured vs unstructured logic.')
    if re.search(r'reason code|r-message|rjct|rtn',   d): notes.append('Reason code/R-message affected — update code validation lists.')
    if re.search(r'\buetr\b',                          d): notes.append('UETR handling changed — check end-to-end transaction ID logic.')
    if re.search(r'iban|account.*id|bic',              d): notes.append('Account/BIC handling affected — verify validation logic.')
    if re.search(r'charge|fee',                        d): notes.append('Charges/fees impacted — review charge bearer logic.')
    if re.search(r'purpose',                           d): notes.append('Purpose code affected — check allowed purpose values.')
    if re.search(r'remittance',                        d): notes.append('Remittance information changed — update remittance handling.')
    if notes:
        return prefix + '  ' + '  '.join(notes)
    return prefix


def extract_changes(pdf_path):
    """Returns (metadata_dict, list_of_entry_dicts)."""
    meta    = _extract_metadata(pdf_path)
    entries = []

    with pdfplumber.open(pdf_path) as pdf:
        start = _find_changes_start(pdf)
        if start < 0:
            return meta, entries

        current    = None
        current_ds = 'General / Cross-Message'
        npc_fmt    = None

        for pg_idx in range(start, len(pdf.pages)):
            tables = pdf.pages[pg_idx].extract_tables()
            for table in tables:
                if not table: continue
                for row in table:
                    if not row: continue
                    cells = [str(c or '').replace('\n', ' ').strip() for c in row]
                    if not any(cells): continue

                    c0 = cells[0].lower().rstrip('.')

                    # Skip pure header rows
                    if c0 in ('n', 'no', '#', 'section /', 'message', 'element', ''):
                        if len(cells) < 4 or not cells[3].strip() or cells[3].lower() in ('description', ''):
                            continue

                    num_m = re.match(r'^(\d+)\.?$', cells[0])

                    if not num_m:
                        # Dataset section divider or continuation
                        joined = ' '.join(c for c in cells if c)
                        ds = _infer_dataset(joined)
                        if ds:
                            current_ds = ds
                        elif current and len(cells) > 3 and cells[3]:
                            current['description'] += ' ' + cells[3]
                        elif current and cells[0] and not cells[0].startswith('N'):
                            current['description'] += ' ' + cells[0]
                        continue

                    # New numbered entry
                    if current:
                        entries.append(current)

                    # Detect NPC vs EPC format once
                    if npc_fmt is None:
                        npc_fmt = (len(row) >= 6 and str(row[4] or '').strip() in ('ACC','REJ','PEN','PEND',''))

                    num = num_m.group(1)

                    if npc_fmt and len(cells) >= 6:
                        section_ref = cells[1]
                        element_ref = cells[2]
                        description = cells[3]
                        status      = cells[4]
                        chg_type    = (cells[5] or 'CHAN').upper()
                    else:
                        section_ref = cells[1]
                        element_ref = cells[2]
                        description = cells[3]
                        status      = ''
                        chg_type    = (cells[4] if len(cells) > 4 else 'CHAN').upper()

                    if chg_type not in ('CHAN', 'CLAR', 'TYPO'):
                        chg_type = 'CHAN'

                    # Try to infer dataset from section / element
                    ds = _infer_dataset(section_ref + ' ' + element_ref) or current_ds

                    current = {
                        'number':      num,
                        'section_ref': section_ref,
                        'element_ref': element_ref,
                        'description': description,
                        'status':      status,
                        'change_type': chg_type,
                        'dataset':     ds,
                    }

        if current:
            entries.append(current)

    # Clean descriptions and generate business notes
    for e in entries:
        e['description']   = re.sub(r'\s+', ' ', e['description']).strip()
        e['business_note'] = _business_note(e['change_type'], e['element_ref'], e['description'])

    return meta, entries


# Column definition: (header, dict_key, width)
_COLS = [
    ('N°',                              'number',       6),
    ('Section\nRef',                    'section_ref', 14),
    ('Element /\nIndex Ref',            'element_ref', 18),
    ('Dataset / Message',               'dataset',     36),
    ('Change\nType',                    'change_type', 16),
    ('Status',                          'status',       9),
    ('Description of Change',           'description', 62),
    ('Business Note & Implementation Guidance', 'business_note', 70),
]


def _write_sheet(ws, entries, meta, label):
    total_cols  = len(_COLS)
    last_letter = get_column_letter(total_cols)

    # Row 1: banner
    c = ws.cell(row=1, column=1, value=(
        f"{meta.get('doc_number','') or label}  |  {meta.get('version','')}  |  "
        f"Issued: {meta.get('date_issued','')}  |  "
        f"Effective: {meta.get('date_effective','')}  |  "
        f"Scheme: {meta.get('scheme','')}"
    ))
    ws.merge_cells(f'A1:{last_letter}1')
    c.font      = _font(bold=True, size=12, color=_C['white'])
    c.fill      = _fill(_C['navy'])
    c.alignment = _align(h='left', v='center')
    ws.row_dimensions[1].height = 22

    # Row 2: stats boxes
    chan_n  = sum(1 for e in entries if e['change_type'] == 'CHAN')
    clar_n  = sum(1 for e in entries if e['change_type'] == 'CLAR')
    typo_n  = sum(1 for e in entries if e['change_type'] == 'TYPO')
    boxes   = [
        (f'{len(entries)}  Total', _C['blue'],    _C['white']),
        (f'{chan_n}  🟠 CHAN',      _C['chan_bg'], _C['chan_fg']),
        (f'{clar_n}  🔵 CLAR',     _C['clar_bg'], _C['clar_fg']),
        (f'{typo_n}  🟡 TYPO',     _C['typo_bg'], _C['typo_fg']),
    ]
    bw = max(1, total_cols // len(boxes))
    for i, (txt, bg, fg) in enumerate(boxes):
        cs = i * bw + 1
        ce = (i+1)*bw if i < len(boxes)-1 else total_cols
        c  = ws.cell(row=2, column=cs, value=txt)
        c.font      = _font(bold=True, size=11, color=fg)
        c.fill      = _fill(bg)
        c.alignment = _align(h='center', v='center', wrap=False)
        c.border    = _border_med()
        if cs != ce:
            ws.merge_cells(f'{get_column_letter(cs)}2:{get_column_letter(ce)}2')
    ws.row_dimensions[2].height = 30

    # Row 3: legend
    legend = ('🟠 CHAN = Rulebook / content change — may require implementation update     '
              '🔵 CLAR = Clarification — no functional change expected     '
              '🟡 TYPO = Typo or layout fix — no functional impact')
    c = ws.cell(row=3, column=1, value=legend)
    ws.merge_cells(f'A3:{last_letter}3')
    c.font      = _font(size=8, color='FF333333')
    c.fill      = _fill(_C['lt_gray'])
    c.alignment = _align(h='left', v='center', wrap=False)
    c.border    = _border()
    ws.row_dimensions[3].height = 14

    # Row 4: column headers
    for ci, (hdr, _, _w) in enumerate(_COLS, 1):
        c = ws.cell(row=4, column=ci, value=hdr)
        c.font      = _font(bold=True, size=9, color=_C['white'])
        c.fill      = _fill(_C['blue'])
        c.alignment = _align(h='center', v='center', wrap=True)
        c.border    = _border()
    ws.row_dimensions[4].height = 32
    ws.freeze_panes = 'A5'

    # Data rows
    prev_ds  = None
    data_row = 5

    for entry in entries:
        chg = entry['change_type']
        bg  = _TYPE_BG.get(chg, _C['white'])
        fg  = _TYPE_FG.get(chg, 'FF000000')

        # Dataset divider row
        if entry['dataset'] != prev_ds:
            prev_ds = entry['dataset']
            dc = ws.cell(row=data_row, column=1, value='▶  ' + entry['dataset'])
            dc.font      = _font(bold=True, size=9, color=_C['ds_fg'])
            dc.fill      = _fill(_C['ds_bg'])
            dc.alignment = _align(h='left', v='center', wrap=False)
            dc.border    = _border_med()
            ws.merge_cells(f'A{data_row}:{last_letter}{data_row}')
            ws.row_dimensions[data_row].height = 16
            data_row += 1

        for ci, (_, field, _w) in enumerate(_COLS, 1):
            val = entry.get(field, '')
            if field == 'change_type':
                val = _TYPE_LABEL.get(chg, chg)

            c = ws.cell(row=data_row, column=ci, value=val)
            c.border    = _border()
            c.alignment = _align(
                h='center' if field in ('number', 'change_type', 'status') else 'left',
                v='top', wrap=True
            )
            if field == 'business_note':
                c.font = _font(size=8, color='FF1A5C38')
                c.fill = _fill('FFEBF5ED')
            elif field == 'change_type':
                c.font = _font(bold=True, size=8, color=fg)
                c.fill = _fill(bg)
            elif field == 'description':
                c.font = _font(size=8, color='FF1A1A1A')
                c.fill = _fill(bg)
            else:
                c.font = _font(size=8, color='FF404040')
                c.fill = _fill(bg)

        ws.row_dimensions[data_row].height = 60
        data_row += 1

    # Column widths
    for ci, (_, _, w) in enumerate(_COLS, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w

    return len(entries)


def _write_summary(ws, doc_results):
    total_cols = 8
    last = get_column_letter(total_cols)

    c = ws.cell(row=1, column=1, value='Rulebook Change Tracker — Summary')
    ws.merge_cells(f'A1:{last}1')
    c.font      = _font(bold=True, size=14, color=_C['white'])
    c.fill      = _fill(_C['navy'])
    c.alignment = _align(h='left', v='center')
    ws.row_dimensions[1].height = 26

    hdrs = ['Document', 'Version', 'Date Issued', 'Date Effective',
            'Scheme', '🟠 CHAN', '🔵 CLAR', '🟡 TYPO']
    for ci, h in enumerate(hdrs, 1):
        c = ws.cell(row=2, column=ci, value=h)
        c.font      = _font(bold=True, size=10, color=_C['white'])
        c.fill      = _fill(_C['blue'])
        c.alignment = _align(h='center', v='center', wrap=True)
        c.border    = _border()
    ws.row_dimensions[2].height = 28
    ws.freeze_panes = 'A3'

    for ri, (meta, entries) in enumerate(doc_results, 3):
        chan_n = sum(1 for e in entries if e['change_type'] == 'CHAN')
        clar_n = sum(1 for e in entries if e['change_type'] == 'CLAR')
        typo_n = sum(1 for e in entries if e['change_type'] == 'TYPO')
        vals   = [meta.get('doc_number', meta.get('filename', '')),
                  meta.get('version', ''), meta.get('date_issued', ''),
                  meta.get('date_effective', ''), meta.get('scheme', ''),
                  chan_n, clar_n, typo_n]
        bg = _C['lt_gray'] if ri % 2 == 0 else _C['white']
        for ci, val in enumerate(vals, 1):
            c = ws.cell(row=ri, column=ci, value=val)
            c.border    = _border()
            c.alignment = _align(h='center' if ci > 5 else 'left', wrap=False)
            if   ci == 6 and isinstance(val, int) and val > 0:
                c.font = _font(bold=True, size=10, color=_C['chan_fg']); c.fill = _fill(_C['chan_bg'])
            elif ci == 7 and isinstance(val, int) and val > 0:
                c.font = _font(bold=True, size=10, color=_C['clar_fg']); c.fill = _fill(_C['clar_bg'])
            elif ci == 8 and isinstance(val, int) and val > 0:
                c.font = _font(bold=True, size=10, color=_C['typo_fg']); c.fill = _fill(_C['typo_bg'])
            else:
                c.font = _font(size=10, bold=(ci == 1)); c.fill = _fill(bg)
        ws.row_dimensions[ri].height = 22

    widths = [22, 20, 22, 22, 22, 10, 10, 10]
    for ci, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w


def track_changes(pdf_a: str, output_path: str, pdf_b: Optional[str] = None) -> dict:
    """
    Extract change lists from one or two IG PDFs into a structured Excel workbook.

    Parameters
    ----------
    pdf_a       : Path to first (or only) IG PDF
    output_path : Where to save the output Excel
    pdf_b       : Optional second PDF for comparison

    Returns
    -------
    dict: {total_changes, docs: [...], output}
    """
    sources = [(pdf_a, 'Doc A')]
    if pdf_b:
        sources.append((pdf_b, 'Doc B'))

    doc_results = []
    for pdf_path, fallback in sources:
        meta, entries = extract_changes(pdf_path)
        label = meta.get('doc_number') or fallback
        doc_results.append((meta, entries, label))

    wb = Workbook()
    wb.remove(wb.active)

    summary_ws = wb.create_sheet('Summary', 0)
    _write_summary(summary_ws, [(m, e) for m, e, _ in doc_results])

    for meta, entries, label in doc_results:
        ws_name = re.sub(r'[/\\*?:\[\]]', '-', label)[:31]
        ws = wb.create_sheet(ws_name)
        _write_sheet(ws, entries, meta, label)

    os.makedirs(os.path.dirname(os.path.abspath(output_path)), exist_ok=True)
    wb.save(output_path)

    return {
        'total_changes': sum(len(e) for _, e, _ in doc_results),
        'docs': [
            {'doc_number': meta.get('doc_number', lbl),
             'version':    meta.get('version', ''),
             'num_changes': len(entries),
             'chan': sum(1 for e in entries if e['change_type'] == 'CHAN'),
             'clar': sum(1 for e in entries if e['change_type'] == 'CLAR'),
             'typo': sum(1 for e in entries if e['change_type'] == 'TYPO')}
            for meta, entries, lbl in doc_results
        ],
        'output': output_path,
    }
