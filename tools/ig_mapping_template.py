"""
ig_mapping_template.py  —  IG to Mapping Template
Takes one or more IG Extractor Excel files and produces a pre-filled
implementation mapping workbook.

For each message section the workbook contains:
  • One mapping sheet  — all fields, colour-coded, ready to fill in
  • One summary sheet  — field count dashboard by status

The mapping sheet columns
--------------------------
  A  Status          🟡 Mandatory / ⬜ Optional / 🔴 Not Permitted
  B  Index           IG field number (1.1, 2.8 …)
  C  XML Tag         e.g. MsgId, CreDtTm
  D  Element Name    Human-readable name
  E  XPath           Full element path
  F  Multiplicity    0..1 / 1..1 / 0..n …
  G  Data Type       Max35Text, ISODateTime …
  H  Length          SEPA/NPC length constraint
  I  Usage Rule      SEPA/NPC usage rule text
  J  Rulebook Ref    AT-XXXX reference
  K  Format Rule     Format/pattern constraints
  L  Code Values     Allowed code set values
  M  ISO Definition  Official ISO 20022 field definition
  ── IMPLEMENTATION COLUMNS (to be filled by the team) ──────────────────
  N  Source System   Which source system provides this field
  O  Source Field    Source field name / column
  P  Transformation  Mapping rule / transformation logic
  Q  Default Value   Hardcoded or default value if no source
  R  Validation      Additional business validation rule
  S  Dev Owner       Developer responsible for this field
  T  Status          TODO / IN PROGRESS / DONE / N/A
  U  Notes           Free-text notes

Usage
-----
    from ig_mapping_template import generate_mapping
    result = generate_mapping('NPC_IG.xlsx', 'NPC_Mapping.xlsx',
                              scheme_label='NPC', version='2025 v1.1')
    # result = {total_fields, mandatory, optional, not_permitted, sheets, output_path}
"""

import os
import re
from typing import Optional

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# ── Colour palette ─────────────────────────────────────────────────────────────
_C = {
    'navy':          'FF1F3864',
    'blue':          'FF2E74B5',
    'white':         'FFFFFFFF',
    'lt_blue':       'FFD6E4F7',
    'lt_gray':       'FFF5F5F5',
    'gray':          'FFD9D9D9',
    'dark_gray':     'FF595959',

    # IG field status  (matching ig_extractor colours)
    'mandatory_bg':  'FFFFF2CC',   # yellow
    'mandatory_fg':  'FF7D4F00',
    'optional_bg':   'FFFFFFFF',   # white
    'optional_fg':   'FF333333',
    'notperm_bg':    'FFFFCCCC',   # red
    'notperm_fg':    'FF8B0000',

    # Implementation column header band
    'impl_hdr_bg':   'FF375623',   # dark green
    'impl_hdr_fg':   'FFFFFFFF',

    # IG info column header band
    'ig_hdr_bg':     'FF1F3864',   # navy
    'ig_hdr_fg':     'FFFFFFFF',

    # Status dropdown values
    'todo_bg':       'FFFCE8C3',
    'inprog_bg':     'FFD6E4F7',
    'done_bg':       'FFE2EFDA',
    'na_bg':         'FFF2F2F2',
}

# IG source columns (read from IG Extractor sheet)
_IG_READ_COLS = [
    'Index', 'Multiplicity', 'XPath', 'Element Name', 'ISO Name',
    'ISO Definition', 'XML Tag', 'Type',
    'ISO Length', 'SEPA/NPC Length',
    'SEPA/NPC Usage Rules', 'SEPA/NPC Rulebook',
    'SEPA/NPC Format Rules', 'SEPA/NPC FractDigits',
    'SEPA/NPC Inclusive', 'SEPA/NPC Code Restrictions',
]

# Output column definitions: (header, width, section)
# section: 'ig' = pre-filled from IG | 'impl' = blank for team to fill
_MAPPING_COLS = [
    # ── IG Info (pre-filled) ────────────────────────────────────────────────
    ('Status',           14, 'ig'),
    ('Index',             8, 'ig'),
    ('XML Tag',          16, 'ig'),
    ('Element Name',     30, 'ig'),
    ('XPath',            52, 'ig'),
    ('Multiplicity',     12, 'ig'),
    ('Data Type',        22, 'ig'),
    ('Length',           14, 'ig'),
    ('Usage Rule',       45, 'ig'),
    ('Rulebook Ref',     18, 'ig'),
    ('Format Rule',      32, 'ig'),
    ('Code Values',      30, 'ig'),
    ('ISO Definition',   55, 'ig'),
    # ── Implementation (blank) ──────────────────────────────────────────────
    ('Source System',    20, 'impl'),
    ('Source Field',     25, 'impl'),
    ('Transformation',   42, 'impl'),
    ('Default Value',    20, 'impl'),
    ('Validation Rule',  38, 'impl'),
    ('Dev Owner',        16, 'impl'),
    ('Impl Status',      14, 'impl'),
    ('Notes',            40, 'impl'),
]

_N_IG_COLS   = sum(1 for _, _, s in _MAPPING_COLS if s == 'ig')
_N_IMPL_COLS = sum(1 for _, _, s in _MAPPING_COLS if s == 'impl')
_TOTAL_COLS  = len(_MAPPING_COLS)

_IMPL_STATUS_OPTIONS = ['TODO', 'IN PROGRESS', 'DONE', 'N/A', 'BLOCKED']


# ── Style helpers ──────────────────────────────────────────────────────────────

def _fill(hex8: str) -> PatternFill:
    return PatternFill('solid', fgColor=hex8)

def _font(bold=False, color='FF000000', size=9, name='Arial') -> Font:
    return Font(name=name, bold=bold, size=size, color=color)

def _border(thin=True) -> Border:
    style = 'thin' if thin else 'medium'
    s = Side(style=style, color='FFD0D0D0')
    return Border(left=s, right=s, top=s, bottom=s)

def _align(h='left', v='top', wrap=True) -> Alignment:
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def _write(ws, row, col, value, bg=None, fg='FF000000', bold=False,
           size=9, h='left', v='top', wrap=True, border=True):
    c = ws.cell(row=row, column=col, value=value)
    c.font      = _font(bold=bold, color=fg, size=size)
    c.alignment = _align(h=h, v=v, wrap=wrap)
    if bg:
        c.fill = _fill(bg)
    if border:
        c.border = _border()
    return c


# ── Sample value library ───────────────────────────────────────────────────────

_SAMPLES = {
    'MsgId':           'MSG20240115-001',
    'InstrId':         'INSTR-20240115-001',
    'EndToEndId':      'E2E-20240115-001',
    'TxId':            'TX-20240115-001',
    'UETR':            'eb6305c9-1f7f-49de-aed0-16487c27b42d',
    'CreDtTm':         '2024-01-15T10:30:00',
    'IntrBkSttlmDt':   '2024-01-15',
    'NbOfTxs':         '1',
    'TtlIntrBkSttlmAmt': '1000.00',
    'IntrBkSttlmAmt':  '1000.00',
    'InstdAmt':        '1000.00',
    'ChrgBr':          'SLEV',
    'Nm':              'Acme Corporation',
    'IBAN':            'DE89370400440532013000',
    'BICFI':           'DEUTDEFFXXX',
    'BIC':             'DEUTDEFFXXX',
    'Ctry':            'DE',
    'Ccy':             'EUR',
    'StrtNm':          'Hauptstrasse',
    'BldgNb':          '42',
    'PstCd':           '10115',
    'TwnNm':           'Berlin',
    'Ustrd':           'Payment for Invoice INV-2024-001',
    'Cd':              'SALA',
    'Prtry':           'PROPRIETARY001',
    'RmtInf':          'Invoice INV-2024-001',
}

def _sample_value(xml_tag: str, data_type: str) -> str:
    """Return a contextual sample value for a field."""
    tag = (xml_tag or '').strip()
    dtype = (data_type or '').strip().lower()

    if tag in _SAMPLES:
        return _SAMPLES[tag]

    if 'datetime' in dtype or 'date' in dtype:
        return '2024-01-15T10:30:00' if 'time' in dtype else '2024-01-15'
    if 'iban' in dtype:
        return 'DE89370400440532013000'
    if 'bic' in dtype:
        return 'DEUTDEFFXXX'
    if 'amount' in dtype or 'decimal' in dtype:
        return '1000.00'
    if 'numeric' in dtype or 'integer' in dtype:
        return '1'
    if re.search(r'max(\d+)text', dtype):
        n = int(re.search(r'max(\d+)', dtype).group(1))
        return f'SAMPLE{min(n, 20)}'
    if 'indicator' in dtype or 'boolean' in dtype:
        return 'true'
    if 'code' in dtype:
        return 'SLEV'
    return ''


# ── IG data reader ─────────────────────────────────────────────────────────────

def _read_ig_sheet(ws) -> list:
    """
    Read one IG Extractor sheet into a list of field dicts.
    Each dict has keys matching _IG_READ_COLS plus '_status'.
    """
    # Find header row (row 3)
    headers = [ws.cell(row=3, column=c).value for c in range(1, ws.max_column + 1)]

    fields = []
    for r in range(4, ws.max_row + 1):
        row = {}
        for c, h in enumerate(headers, 1):
            if h:
                row[h] = ws.cell(row=r, column=c).value or ''

        # Determine status from row fill colour
        fill = ws.cell(row=r, column=1).fill
        if fill and fill.fill_type == 'solid' and fill.fgColor:
            rgb6 = (fill.fgColor.rgb or 'FFFFFFFF')[-6:].upper()
        else:
            rgb6 = 'FFFFFF'

        color_map = {
            'FFF2CC': 'Mandatory',
            'FFCCCC': 'Not Permitted',
        }
        row['_status'] = color_map.get(rgb6, 'Optional')
        row['_rgb6']   = rgb6

        xpath = str(row.get('XPath', '') or '').strip()
        if xpath:
            fields.append(row)

    return fields


# ── Mapping sheet writer ───────────────────────────────────────────────────────

def _write_mapping_sheet(ws, fields: list, message_id: str,
                          scheme_label: str, section_label: str,
                          filter_mode: str = 'all'):
    """
    Write one mapping sheet for a single message.

    filter_mode: 'all' | 'mandatory' | 'exclude_notperm'
    """

    # Filter fields
    if filter_mode == 'mandatory':
        fields = [f for f in fields if f['_status'] == 'Mandatory']
    elif filter_mode == 'exclude_notperm':
        fields = [f for f in fields if f['_status'] != 'Not Permitted']

    total_cols = _TOTAL_COLS

    # ── Title row ──────────────────────────────────────────────────────────────
    title = f"{message_id.upper()}  —  {section_label}  │  {scheme_label} Implementation Mapping"
    ws.cell(row=1, column=1, value=title)
    ws.merge_cells(f'A1:{get_column_letter(total_cols)}1')
    ws['A1'].font      = _font(bold=True, size=12, color=_C['white'])
    ws['A1'].fill      = _fill(_C['navy'])
    ws['A1'].alignment = _align(h='left', v='center', wrap=False)
    ws.row_dimensions[1].height = 24

    # ── Legend row ─────────────────────────────────────────────────────────────
    legend = [
        ('🟡 Mandatory — must be populated in every message',
         _C['mandatory_bg'], _C['mandatory_fg'], 5),
        ('⬜ Optional — include if available / applicable',
         _C['optional_bg'], 'FF666666', 4),
        ('🔴 Not Permitted — must NOT be included',
         _C['notperm_bg'], _C['notperm_fg'], 3),
        ('🟢 Implementation columns — fill in your mapping details',
         'FFE2EFDA', 'FF1A5632', total_cols - 12),
    ]
    col = 1
    for text, bg, fg, span in legend:
        end = col + span - 1
        ws.cell(row=2, column=col, value=text)
        if span > 1:
            ws.merge_cells(f'{get_column_letter(col)}2:{get_column_letter(end)}2')
        c = ws[f'{get_column_letter(col)}2']
        c.font      = _font(bold=True, size=8, color=fg)
        c.fill      = _fill(bg)
        c.alignment = _align(h='center', v='center', wrap=False)
        c.border    = _border()
        col = end + 1
    ws.row_dimensions[2].height = 14

    # ── Section headers row 3 ──────────────────────────────────────────────────
    # "IG Reference Data" spans IG columns | "Implementation Details" spans impl columns
    ig_start   = 1
    ig_end     = _N_IG_COLS
    impl_start = _N_IG_COLS + 1
    impl_end   = total_cols

    ws.cell(row=3, column=ig_start, value='IG Reference Data  (pre-filled from IG Extractor)')
    ws.merge_cells(f'{get_column_letter(ig_start)}3:{get_column_letter(ig_end)}3')
    ws[f'{get_column_letter(ig_start)}3'].font      = _font(bold=True, size=10, color=_C['ig_hdr_fg'])
    ws[f'{get_column_letter(ig_start)}3'].fill      = _fill(_C['ig_hdr_bg'])
    ws[f'{get_column_letter(ig_start)}3'].alignment = _align(h='center', v='center', wrap=False)
    ws[f'{get_column_letter(ig_start)}3'].border    = _border()

    ws.cell(row=3, column=impl_start,
            value='Implementation Mapping  (fill in by development team)')
    ws.merge_cells(f'{get_column_letter(impl_start)}3:{get_column_letter(impl_end)}3')
    ws[f'{get_column_letter(impl_start)}3'].font      = _font(bold=True, size=10, color=_C['impl_hdr_fg'])
    ws[f'{get_column_letter(impl_start)}3'].fill      = _fill(_C['impl_hdr_bg'])
    ws[f'{get_column_letter(impl_start)}3'].alignment = _align(h='center', v='center', wrap=False)
    ws[f'{get_column_letter(impl_start)}3'].border    = _border()
    ws.row_dimensions[3].height = 18

    # ── Column headers row 4 ──────────────────────────────────────────────────
    for ci, (col_name, _, section) in enumerate(_MAPPING_COLS, 1):
        bg = _C['ig_hdr_bg'] if section == 'ig' else _C['impl_hdr_bg']
        c = ws.cell(row=4, column=ci, value=col_name)
        c.font      = _font(bold=True, size=8, color='FFFFFFFF')
        c.fill      = _fill(bg)
        c.alignment = _align(h='center', v='center', wrap=True)
        c.border    = _border()
    ws.row_dimensions[4].height = 30
    ws.freeze_panes = 'A5'

    # ── Data rows ─────────────────────────────────────────────────────────────
    data_row = 5
    for field in fields:
        status = field.get('_status', 'Optional')

        # Row background based on status
        if status == 'Mandatory':
            row_bg = _C['mandatory_bg']
            row_fg = _C['mandatory_fg']
            status_label = '🟡 Mandatory'
        elif status == 'Not Permitted':
            row_bg = _C['notperm_bg']
            row_fg = _C['notperm_fg']
            status_label = '🔴 Not Permitted'
        else:
            row_bg = _C['optional_bg']
            row_fg = _C['optional_fg']
            status_label = '⬜ Optional'

        # Derive sample value
        xml_tag  = str(field.get('XML Tag', '') or '').strip()
        dtype    = str(field.get('Type', '') or '').strip()
        sample   = _sample_value(xml_tag, dtype)

        # IG columns (pre-filled)
        ig_values = [
            status_label,
            field.get('Index', ''),
            xml_tag,
            str(field.get('Element Name', '') or ''),
            str(field.get('XPath', '') or ''),
            str(field.get('Multiplicity', '') or ''),
            dtype,
            str(field.get('SEPA/NPC Length', '') or ''),
            str(field.get('SEPA/NPC Usage Rules', '') or ''),
            str(field.get('SEPA/NPC Rulebook', '') or ''),
            str(field.get('SEPA/NPC Format Rules', '') or ''),
            str(field.get('SEPA/NPC Code Restrictions', '') or ''),
            str(field.get('ISO Definition', '') or ''),
        ]

        # Implementation columns (pre-populated with hints)
        # Source System: guess from XML tag
        src_system = _guess_source_system(xml_tag, status)
        # Transformation: populate with sample where we have one
        transform  = sample if sample else ''
        # Default value: populate for known hardcoded fields
        default    = _guess_default(xml_tag, dtype)
        # Impl status: default based on field status
        impl_status = 'TODO' if status == 'Mandatory' else ('N/A' if status == 'Not Permitted' else 'TODO')

        impl_values = [
            src_system,   # Source System
            '',           # Source Field (blank — team fills)
            transform,    # Transformation (sample value as hint)
            default,      # Default Value
            '',           # Validation Rule
            '',           # Dev Owner
            impl_status,  # Impl Status
            '',           # Notes
        ]

        all_values = ig_values + impl_values
        for ci, (val, (col_name, _, section)) in enumerate(zip(all_values, _MAPPING_COLS), 1):
            if section == 'ig':
                # IG columns: use field status colour
                cell_bg = row_bg
                cell_fg = row_fg
                cell_bold = (ci == 1)  # bold the status cell
            else:
                # Implementation columns: neutral background
                if col_name == 'Impl Status':
                    status_bg_map = {
                        'TODO':        _C['todo_bg'],
                        'IN PROGRESS': _C['inprog_bg'],
                        'DONE':        _C['done_bg'],
                        'N/A':         _C['na_bg'],
                        'BLOCKED':     _C['notperm_bg'],
                    }
                    cell_bg = status_bg_map.get(str(val), _C['lt_gray'])
                else:
                    # Alternate stripe for readability
                    cell_bg = 'FFFFFEF2' if status == 'Mandatory' else _C['lt_gray']
                cell_fg = 'FF333333'
                cell_bold = False

            _write(ws, data_row, ci, val, bg=cell_bg, fg=cell_fg,
                   bold=cell_bold, h='left' if ci > 2 else 'center',
                   v='top', size=8, wrap=True)

        ws.row_dimensions[data_row].height = 50
        data_row += 1

    # ── Add data validation dropdown for Impl Status column ───────────────────
    impl_status_col = next(
        (i + 1 for i, (n, _, _) in enumerate(_MAPPING_COLS) if n == 'Impl Status'), None
    )
    if impl_status_col and data_row > 5:
        status_col_letter = get_column_letter(impl_status_col)
        dv = DataValidation(
            type='list',
            formula1='"' + ','.join(_IMPL_STATUS_OPTIONS) + '"',
            allow_blank=True
        )
        dv.sqref = f'{status_col_letter}5:{status_col_letter}{data_row}'
        ws.add_data_validation(dv)

    # ── Column widths ──────────────────────────────────────────────────────────
    for ci, (_, width, _) in enumerate(_MAPPING_COLS, 1):
        ws.column_dimensions[get_column_letter(ci)].width = width

    return data_row - 5  # number of field rows written


def _guess_source_system(xml_tag: str, status: str) -> str:
    """Heuristic guess at source system for common fields."""
    if status == 'Not Permitted':
        return ''
    tag_map = {
        'MsgId':    'Internal/Generated',
        'CreDtTm':  'Internal/Generated',
        'NbOfTxs':  'Internal/Generated',
        'InstrId':  'Internal/Generated',
        'UETR':     'Internal/Generated',
        'TxId':     'Internal/Generated',
        'EndToEndId': 'Originating System',
        'IntrBkSttlmAmt': 'Originating System',
        'IntrBkSttlmDt':  'Originating System',
        'InstdAmt': 'Originating System',
        'Nm':       'CRM / Customer DB',
        'IBAN':     'Account System',
        'BICFI':    'BIC Directory',
        'BIC':      'BIC Directory',
        'Ctry':     'CRM / Customer DB',
        'Ustrd':    'Originating System',
        'RmtInf':   'Originating System',
    }
    return tag_map.get(xml_tag, '')


def _guess_default(xml_tag: str, data_type: str) -> str:
    """Known default/hardcoded values for certain fields."""
    defaults = {
        'ChrgBr':   'SLEV',
        'Ccy':      'EUR',
        'SvcLvl':   'SEPA',
        'LclInstrm': 'INST',
        'CtgyPurp':  '',
    }
    if xml_tag in defaults:
        return defaults[xml_tag]
    if 'indicator' in (data_type or '').lower():
        return 'false'
    return ''


# ── Summary sheet writer ───────────────────────────────────────────────────────

def _write_summary_sheet(ws, sheet_stats: list, scheme_label: str,
                          source_file: str, version: str):
    """Write the workbook Summary sheet."""
    total_cols = 8

    # Title
    ws.cell(row=1, column=1,
            value=f'Mapping Template — {scheme_label}  {version}')
    ws.merge_cells(f'A1:{get_column_letter(total_cols)}1')
    ws['A1'].font      = _font(bold=True, size=14, color='FFFFFFFF')
    ws['A1'].fill      = _fill(_C['navy'])
    ws['A1'].alignment = _align(h='left', v='center', wrap=False)
    ws.row_dimensions[1].height = 28

    # Metadata
    meta_rows = [
        ('Scheme / Label', scheme_label),
        ('Source IG File', source_file),
        ('Version', version),
        ('Generated', __import__('datetime').datetime.now().strftime('%Y-%m-%d %H:%M')),
    ]
    for ri, (label, value) in enumerate(meta_rows, 2):
        _write(ws, ri, 1, label, bg=_C['lt_gray'], fg=_C['dark_gray'],
               bold=True, h='right', size=9)
        _write(ws, ri, 2, value, bg=_C['white'], fg='FF222222', size=9)
        ws.merge_cells(f'B{ri}:{get_column_letter(total_cols)}{ri}')
        ws.row_dimensions[ri].height = 16

    # Stats header
    start_row = len(meta_rows) + 3
    ws.cell(row=start_row, column=1, value='Field Statistics by Message')
    ws.merge_cells(f'A{start_row}:{get_column_letter(total_cols)}{start_row}')
    ws[f'A{start_row}'].font      = _font(bold=True, size=10, color='FFFFFFFF')
    ws[f'A{start_row}'].fill      = _fill(_C['blue'])
    ws[f'A{start_row}'].alignment = _align(h='center', v='center')
    ws.row_dimensions[start_row].height = 18

    # Column headers
    hdr_row = start_row + 1
    hdrs = ['Message', 'Sheet', 'Total Fields', '🟡 Mandatory',
            '⬜ Optional', '🔴 Not Permitted', '% Mandatory', 'Action Required']
    hdr_bgs = [_C['blue']] * len(hdrs)
    for ci, (h, bg) in enumerate(zip(hdrs, hdr_bgs), 1):
        _write(ws, hdr_row, ci, h, bg=bg, fg='FFFFFFFF', bold=True,
               h='center', size=9)
    ws.row_dimensions[hdr_row].height = 26

    # Data rows
    totals = {'total': 0, 'mand': 0, 'opt': 0, 'notperm': 0}
    for ri, s in enumerate(sheet_stats, hdr_row + 1):
        total = s['mandatory'] + s['optional'] + s['not_permitted']
        pct   = f"{s['mandatory'] / total * 100:.0f}%" if total else '0%'
        action = ('🔴 High priority — many mandatory fields' if s['mandatory'] > 50
                  else '🟡 Medium — review mandatory fields'  if s['mandatory'] > 10
                  else '⬜ Lower priority')

        row_vals = [s['message'], s['sheet'],
                    total, s['mandatory'], s['optional'], s['not_permitted'],
                    pct, action]
        row_bg = _C['lt_gray'] if ri % 2 == 0 else _C['white']

        for ci, val in enumerate(row_vals, 1):
            bg = row_bg
            fg = 'FF333333'
            if ci == 4 and isinstance(val, int) and val > 0:
                bg = _C['mandatory_bg']; fg = _C['mandatory_fg']
            elif ci == 6 and isinstance(val, int) and val > 0:
                bg = _C['notperm_bg'];   fg = _C['notperm_fg']
            _write(ws, ri, ci, val, bg=bg, fg=fg,
                   h='center' if ci > 2 else 'left', size=9)
        ws.row_dimensions[ri].height = 18

        totals['total']   += total
        totals['mand']    += s['mandatory']
        totals['opt']     += s['optional']
        totals['notperm'] += s['not_permitted']

    # Totals row
    tot_row = hdr_row + len(sheet_stats) + 1
    pct_tot = f"{totals['mand'] / totals['total'] * 100:.0f}%" if totals['total'] else '0%'
    tot_vals = ['TOTAL', '', totals['total'], totals['mand'],
                totals['opt'], totals['notperm'], pct_tot, '']
    for ci, val in enumerate(tot_vals, 1):
        _write(ws, tot_row, ci, val, bg=_C['lt_blue'], fg=_C['navy'],
               bold=True, h='center' if ci > 1 else 'left', size=9)
    ws.row_dimensions[tot_row].height = 20

    # How-to guide box
    guide_row = tot_row + 2
    ws.cell(row=guide_row, column=1, value='📖 How to Use This Mapping Template')
    ws.merge_cells(f'A{guide_row}:{get_column_letter(total_cols)}{guide_row}')
    ws[f'A{guide_row}'].font      = _font(bold=True, size=10, color='FFFFFFFF')
    ws[f'A{guide_row}'].fill      = _fill(_C['impl_hdr_bg'])
    ws[f'A{guide_row}'].alignment = _align(h='left', v='center')
    ws.row_dimensions[guide_row].height = 18

    guide_lines = [
        ('1. Start with Mandatory fields',
         'Work through 🟡 yellow rows first — these must be mapped for every message.'),
        ('2. Identify your source',
         'Fill in "Source System" and "Source Field" for each field. Use "Internal/Generated" for IDs and timestamps your system creates.'),
        ('3. Define transformations',
         'In "Transformation" describe any mapping logic: format conversion, lookup, concatenation, or hardcoded value.'),
        ('4. Set Default Values',
         'For fields with fixed values (e.g. ChrgBr=SLEV, Ccy=EUR) enter the default. These rarely need a source field.'),
        ('5. Track implementation',
         'Update "Impl Status" as work progresses: TODO → IN PROGRESS → DONE. Filter by status to track progress.'),
        ('6. Not Permitted fields',
         '🔴 Red rows must NOT be sent. Mark as N/A and add a note if your source system populates these fields.'),
        ('7. Columns are editable',
         'Add rows, columns, or sheets as needed. The IG Reference Data columns are for reference — do not delete them.'),
    ]

    for i, (heading, text) in enumerate(guide_lines):
        gr = guide_row + 1 + i
        _write(ws, gr, 1, heading, bg='FFE8F5E9', fg=_C['impl_hdr_bg'],
               bold=True, h='left', size=9)
        ws.merge_cells(f'B{gr}:{get_column_letter(total_cols)}{gr}')
        _write(ws, gr, 2, text, bg='FFFAFAFA', fg='FF333333', size=8)
        ws.row_dimensions[gr].height = 22

    # Column widths
    widths = [32, 28, 14, 14, 14, 16, 12, 30]
    for ci, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w


# ── Sheet-naming helper ────────────────────────────────────────────────────────

def _safe_sheet_name(name: str, existing: list) -> str:
    name = re.sub(r'[\\/*?:\[\]]', '_', name)[:31]
    base, counter = name, 2
    while name in existing:
        name = f'{base[:28]}_{counter}'; counter += 1
    return name


# ── Public API ─────────────────────────────────────────────────────────────────

def generate_mapping(
    ig_excel_path: str,
    output_path: str,
    scheme_label: str = '',
    version: str = '',
    filter_mode: str = 'all',
) -> dict:
    """
    Generate an implementation mapping template from an IG Extractor Excel.

    Parameters
    ----------
    ig_excel_path : Path to IG Extractor output (.xlsx)
    output_path   : Where to save the mapping workbook
    scheme_label  : Label for the scheme, e.g. 'EPC' or 'NPC'
    version       : Version string, e.g. '2025 v1.0'
    filter_mode   : 'all'           — include all fields (default)
                    'mandatory'     — mandatory only (smaller, focused)
                    'exclude_notperm' — all except Not Permitted fields

    Returns
    -------
    dict with keys: total_fields, mandatory, optional, not_permitted,
                    sheets, output_path
    """
    ig_wb = load_workbook(ig_excel_path, data_only=True)

    # Auto-detect scheme label and version from Summary sheet if not provided
    if not scheme_label or not version:
        if 'Summary' in ig_wb.sheetnames:
            ws_sum = ig_wb['Summary']
            # Row 2 of summary has column headers, row 3+ has data
            for r in range(3, ws_sum.max_row + 1):
                msg_val = ws_sum.cell(row=r, column=2).value or ''
                if msg_val:
                    if not scheme_label:
                        # Guess from filename
                        fname = os.path.basename(ig_excel_path).upper()
                        scheme_label = ('NPC' if 'NPC' in fname else
                                        'EPC' if 'EPC' in fname else 'SEPA')
                    break

    scheme_label = scheme_label or 'SEPA'
    version      = version or ''

    out_wb = Workbook()
    out_wb.remove(out_wb.active)

    # Summary sheet (written last, inserted at position 0)
    sheet_stats = []

    for sheet_name in ig_wb.sheetnames:
        if sheet_name == 'Summary':
            continue

        ws_ig = ig_wb[sheet_name]
        fields = _read_ig_sheet(ws_ig)
        if not fields:
            continue

        # Infer message ID and section label from sheet name and Summary
        message_id = sheet_name.replace('_', '.').rstrip('.')
        section_label = sheet_name.replace('_', ' ')

        # Try to get a better label from Summary sheet
        if 'Summary' in ig_wb.sheetnames:
            ws_sum = ig_wb['Summary']
            for r in range(3, ws_sum.max_row + 1):
                sn = str(ws_sum.cell(row=r, column=1).value or '')
                if sn == sheet_name:
                    msg_v   = ws_sum.cell(row=r, column=2).value or ''
                    label_v = ws_sum.cell(row=r, column=3).value or ''
                    if msg_v:    message_id    = str(msg_v)
                    if label_v:  section_label = str(label_v)
                    break

        # Write mapping sheet
        map_sheet_name = _safe_sheet_name(sheet_name, [s.title for s in out_wb.worksheets])
        ws_map = out_wb.create_sheet(map_sheet_name)
        rows_written = _write_mapping_sheet(
            ws_map, fields, message_id, scheme_label, section_label, filter_mode
        )

        # Tally stats
        mand    = sum(1 for f in fields if f['_status'] == 'Mandatory')
        opt     = sum(1 for f in fields if f['_status'] == 'Optional')
        notperm = sum(1 for f in fields if f['_status'] == 'Not Permitted')

        sheet_stats.append({
            'message':      message_id,
            'sheet':        map_sheet_name,
            'mandatory':    mand,
            'optional':     opt,
            'not_permitted': notperm,
        })

    # Write summary sheet at position 0
    sum_ws = out_wb.create_sheet('Summary', 0)
    _write_summary_sheet(
        sum_ws, sheet_stats, scheme_label,
        os.path.basename(ig_excel_path), version
    )

    os.makedirs(os.path.dirname(os.path.abspath(output_path)), exist_ok=True)
    out_wb.save(output_path)

    total_mand    = sum(s['mandatory']     for s in sheet_stats)
    total_opt     = sum(s['optional']      for s in sheet_stats)
    total_notperm = sum(s['not_permitted'] for s in sheet_stats)

    return {
        'total_fields':   total_mand + total_opt + total_notperm,
        'mandatory':      total_mand,
        'optional':       total_opt,
        'not_permitted':  total_notperm,
        'sheets':         sheet_stats,
        'output_path':    output_path,
    }
