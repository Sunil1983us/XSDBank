#!/usr/bin/env python3
"""
ISO 20022 Batch XML Validator
=============================
Validate multiple XML files against an XSD schema at once.

Features:
✅ Validate hundreds of XML files in one go
✅ Summary report: X passed, Y failed
✅ Detailed error report per file
✅ Export results to Excel
✅ HTML dashboard with statistics
✅ Progress tracking
"""

import xml.etree.ElementTree as ET
import json
import argparse
import zipfile
import os
from pathlib import Path
from datetime import datetime
from typing import Dict, List, Optional
from dataclasses import dataclass, asdict
from concurrent.futures import ThreadPoolExecutor, as_completed
import sys

try:
    from lxml import etree as lxml_etree
    HAS_LXML = True
except ImportError:
    HAS_LXML = False

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


@dataclass
class FileValidationResult:
    filename: str
    valid: bool
    error_count: int
    warning_count: int
    errors: List[Dict]
    processing_time_ms: float


class BatchXMLValidator:
    """Validate multiple XML files against an XSD schema"""
    
    def __init__(self, xsd_file: str):
        self.xsd_file = xsd_file
        self.schema = None
        self.results: List[FileValidationResult] = []
        
        # Load XSD schema
        if HAS_LXML:
            try:
                with open(xsd_file, 'rb') as f:
                    schema_doc = lxml_etree.parse(f)
                self.schema = lxml_etree.XMLSchema(schema_doc)
            except Exception as e:
                print(f"Warning: Could not load XSD schema: {e}")
    
    def validate_file(self, xml_file: str) -> FileValidationResult:
        """Validate a single XML file"""
        import time
        start_time = time.time()
        
        filename = os.path.basename(xml_file)
        errors = []
        
        # Basic XML parsing check
        try:
            tree = ET.parse(xml_file)
            root = tree.getroot()
        except ET.ParseError as e:
            return FileValidationResult(
                filename=filename,
                valid=False,
                error_count=1,
                warning_count=0,
                errors=[{
                    'type': 'PARSE_ERROR',
                    'message': f"XML parsing error: {str(e)}",
                    'line': getattr(e, 'position', (None,))[0]
                }],
                processing_time_ms=(time.time() - start_time) * 1000
            )
        
        # XSD validation (if lxml available)
        if self.schema and HAS_LXML:
            try:
                with open(xml_file, 'rb') as f:
                    xml_doc = lxml_etree.parse(f)
                
                if not self.schema.validate(xml_doc):
                    for error in self.schema.error_log:
                        errors.append({
                            'type': 'XSD_ERROR',
                            'message': str(error.message)[:200],
                            'line': error.line,
                            'path': error.path
                        })
            except Exception as e:
                errors.append({
                    'type': 'VALIDATION_ERROR',
                    'message': f"Validation error: {str(e)}"
                })
        
        processing_time = (time.time() - start_time) * 1000
        
        return FileValidationResult(
            filename=filename,
            valid=len(errors) == 0,
            error_count=len(errors),
            warning_count=0,
            errors=errors,
            processing_time_ms=processing_time
        )
    
    def validate_batch(self, xml_files: List[str], max_workers: int = 4, 
                       progress_callback=None) -> Dict:
        """Validate multiple XML files"""
        self.results = []
        total_files = len(xml_files)
        completed = 0
        
        with ThreadPoolExecutor(max_workers=max_workers) as executor:
            future_to_file = {executor.submit(self.validate_file, f): f 
                            for f in xml_files}
            
            for future in as_completed(future_to_file):
                result = future.result()
                self.results.append(result)
                completed += 1
                
                if progress_callback:
                    progress_callback(completed, total_files, result.filename)
        
        # Sort by filename
        self.results.sort(key=lambda x: x.filename)
        
        # Generate summary
        passed = sum(1 for r in self.results if r.valid)
        failed = sum(1 for r in self.results if not r.valid)
        total_errors = sum(r.error_count for r in self.results)
        avg_time = sum(r.processing_time_ms for r in self.results) / len(self.results) if self.results else 0
        
        return {
            'summary': {
                'total_files': total_files,
                'passed': passed,
                'failed': failed,
                'pass_rate': f"{(passed/total_files*100):.1f}%" if total_files > 0 else "N/A",
                'total_errors': total_errors,
                'avg_processing_time_ms': round(avg_time, 2)
            },
            'results': [asdict(r) for r in self.results]
        }
    
    def generate_excel_report(self, output_path: str):
        """Generate Excel report with validation results"""
        if not HAS_OPENPYXL:
            print("Warning: openpyxl not installed, skipping Excel report")
            return
        
        wb = Workbook()
        
        # Summary sheet
        ws_summary = wb.active
        ws_summary.title = "Summary"
        
        # Styles
        header_fill = PatternFill(start_color="1a1a2e", end_color="1a1a2e", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        pass_fill = PatternFill(start_color="10b981", end_color="10b981", fill_type="solid")
        fail_fill = PatternFill(start_color="ef4444", end_color="ef4444", fill_type="solid")
        
        # Summary stats
        passed = sum(1 for r in self.results if r.valid)
        failed = sum(1 for r in self.results if not r.valid)
        
        summary_data = [
            ["Batch XML Validation Report", ""],
            ["Generated", datetime.now().strftime('%Y-%m-%d %H:%M:%S')],
            ["XSD Schema", os.path.basename(self.xsd_file)],
            ["", ""],
            ["SUMMARY", ""],
            ["Total Files", len(self.results)],
            ["Passed ✅", passed],
            ["Failed ❌", failed],
            ["Pass Rate", f"{(passed/len(self.results)*100):.1f}%" if self.results else "N/A"],
            ["Total Errors", sum(r.error_count for r in self.results)]
        ]
        
        for row_idx, row_data in enumerate(summary_data, 1):
            for col_idx, value in enumerate(row_data, 1):
                cell = ws_summary.cell(row=row_idx, column=col_idx, value=value)
                if row_idx == 1 or row_idx == 5:
                    cell.font = Font(bold=True, size=14)
        
        ws_summary.column_dimensions['A'].width = 20
        ws_summary.column_dimensions['B'].width = 30
        
        # Results sheet
        ws_results = wb.create_sheet("Results")
        
        headers = ["#", "Filename", "Status", "Errors", "Processing Time (ms)"]
        for col_idx, header in enumerate(headers, 1):
            cell = ws_results.cell(row=1, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
        
        for row_idx, result in enumerate(self.results, 2):
            ws_results.cell(row=row_idx, column=1, value=row_idx-1)
            ws_results.cell(row=row_idx, column=2, value=result.filename)
            
            status_cell = ws_results.cell(row=row_idx, column=3, 
                                          value="✅ PASS" if result.valid else "❌ FAIL")
            status_cell.fill = pass_fill if result.valid else fail_fill
            status_cell.font = Font(color="FFFFFF", bold=True)
            
            ws_results.cell(row=row_idx, column=4, value=result.error_count)
            ws_results.cell(row=row_idx, column=5, value=round(result.processing_time_ms, 2))
        
        # Set column widths
        ws_results.column_dimensions['A'].width = 8
        ws_results.column_dimensions['B'].width = 50
        ws_results.column_dimensions['C'].width = 12
        ws_results.column_dimensions['D'].width = 10
        ws_results.column_dimensions['E'].width = 20
        
        # Errors sheet
        ws_errors = wb.create_sheet("Errors")
        
        error_headers = ["File", "Error Type", "Line", "Message"]
        for col_idx, header in enumerate(error_headers, 1):
            cell = ws_errors.cell(row=1, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
        
        error_row = 2
        for result in self.results:
            for error in result.errors:
                ws_errors.cell(row=error_row, column=1, value=result.filename)
                ws_errors.cell(row=error_row, column=2, value=error.get('type', 'ERROR'))
                ws_errors.cell(row=error_row, column=3, value=error.get('line', ''))
                ws_errors.cell(row=error_row, column=4, value=error.get('message', '')[:500])
                error_row += 1
        
        ws_errors.column_dimensions['A'].width = 40
        ws_errors.column_dimensions['B'].width = 15
        ws_errors.column_dimensions['C'].width = 8
        ws_errors.column_dimensions['D'].width = 100
        
        wb.save(output_path)
    
    def generate_html_report(self, output_path: str):
        """Generate HTML dashboard report"""
        passed = sum(1 for r in self.results if r.valid)
        failed = sum(1 for r in self.results if not r.valid)
        total = len(self.results)
        pass_rate = (passed/total*100) if total > 0 else 0
        
        # Generate results rows
        results_html = ""
        for i, result in enumerate(self.results, 1):
            status_class = "pass" if result.valid else "fail"
            status_text = "✅ PASS" if result.valid else "❌ FAIL"
            
            errors_html = ""
            if result.errors:
                errors_list = "<br>".join([f"Line {e.get('line', '?')}: {e.get('message', '')[:100]}" 
                                          for e in result.errors[:5]])
                if len(result.errors) > 5:
                    errors_list += f"<br>... and {len(result.errors) - 5} more"
                errors_html = f'<div class="errors-detail">{errors_list}</div>'
            
            results_html += f"""
            <tr class="{status_class}">
                <td>{i}</td>
                <td class="filename">{result.filename}</td>
                <td class="status {status_class}">{status_text}</td>
                <td>{result.error_count}</td>
                <td>{result.processing_time_ms:.1f}ms</td>
            </tr>
            """
        
        html = f"""<!DOCTYPE html>
<html>
<head>
    <title>Batch XML Validation Report</title>
    <style>
        body {{ font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, sans-serif; margin: 0; padding: 20px; background: #f5f5f5; }}
        .container {{ max-width: 1400px; margin: 0 auto; }}
        .header {{ background: linear-gradient(135deg, #1a1a2e 0%, #16213e 100%); color: white; padding: 30px; border-radius: 12px; margin-bottom: 20px; }}
        .header h1 {{ margin: 0 0 10px 0; }}
        .meta {{ color: #94a3b8; font-size: 14px; }}
        
        .dashboard {{ display: grid; grid-template-columns: repeat(auto-fit, minmax(200px, 1fr)); gap: 20px; margin-bottom: 20px; }}
        .card {{ background: white; padding: 25px; border-radius: 12px; box-shadow: 0 2px 10px rgba(0,0,0,0.1); text-align: center; }}
        .card .number {{ font-size: 48px; font-weight: bold; margin-bottom: 5px; }}
        .card .label {{ color: #666; font-size: 14px; text-transform: uppercase; }}
        .card.passed .number {{ color: #10b981; }}
        .card.failed .number {{ color: #ef4444; }}
        .card.rate .number {{ color: #3b82f6; }}
        
        .progress-bar {{ height: 30px; background: #e5e7eb; border-radius: 15px; overflow: hidden; margin-bottom: 20px; }}
        .progress-fill {{ height: 100%; display: flex; }}
        .progress-pass {{ background: #10b981; }}
        .progress-fail {{ background: #ef4444; }}
        
        .results {{ background: white; padding: 20px; border-radius: 12px; box-shadow: 0 2px 10px rgba(0,0,0,0.1); }}
        .results h2 {{ margin-top: 0; }}
        
        table {{ width: 100%; border-collapse: collapse; }}
        th {{ background: #1a1a2e; color: white; padding: 12px; text-align: left; }}
        td {{ padding: 10px 12px; border-bottom: 1px solid #e5e7eb; }}
        tr:hover {{ background: #f8f9fa; }}
        tr.pass {{ }}
        tr.fail {{ background: #fef2f2; }}
        
        .status {{ font-weight: bold; padding: 4px 8px; border-radius: 4px; }}
        .status.pass {{ background: #10b981; color: white; }}
        .status.fail {{ background: #ef4444; color: white; }}
        
        .filename {{ font-family: monospace; font-size: 13px; }}
        .errors-detail {{ font-size: 11px; color: #666; margin-top: 5px; }}
        
        .filter-bar {{ margin-bottom: 15px; }}
        .filter-bar button {{ padding: 8px 16px; margin-right: 10px; border: none; border-radius: 6px; cursor: pointer; }}
        .filter-bar button.active {{ background: #1a1a2e; color: white; }}
        .filter-bar button:not(.active) {{ background: #e5e7eb; }}
    </style>
</head>
<body>
    <div class="container">
        <div class="header">
            <h1>📦 Batch XML Validation Report</h1>
            <div class="meta">
                <div>📋 XSD: {os.path.basename(self.xsd_file)}</div>
                <div>🕐 Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}</div>
            </div>
        </div>
        
        <div class="dashboard">
            <div class="card">
                <div class="number">{total}</div>
                <div class="label">Total Files</div>
            </div>
            <div class="card passed">
                <div class="number">{passed}</div>
                <div class="label">Passed ✅</div>
            </div>
            <div class="card failed">
                <div class="number">{failed}</div>
                <div class="label">Failed ❌</div>
            </div>
            <div class="card rate">
                <div class="number">{pass_rate:.1f}%</div>
                <div class="label">Pass Rate</div>
            </div>
        </div>
        
        <div class="progress-bar">
            <div class="progress-fill">
                <div class="progress-pass" style="width: {pass_rate}%"></div>
                <div class="progress-fail" style="width: {100-pass_rate}%"></div>
            </div>
        </div>
        
        <div class="results">
            <h2>📋 Validation Results</h2>
            <div class="filter-bar">
                <button class="active" onclick="filterResults('all')">All ({total})</button>
                <button onclick="filterResults('pass')">Passed ({passed})</button>
                <button onclick="filterResults('fail')">Failed ({failed})</button>
            </div>
            <table>
                <thead>
                    <tr>
                        <th>#</th>
                        <th>Filename</th>
                        <th>Status</th>
                        <th>Errors</th>
                        <th>Time</th>
                    </tr>
                </thead>
                <tbody id="resultsBody">
                    {results_html}
                </tbody>
            </table>
        </div>
    </div>
    
    <script>
        function filterResults(filter) {{
            const rows = document.querySelectorAll('#resultsBody tr');
            const buttons = document.querySelectorAll('.filter-bar button');
            
            buttons.forEach(b => b.classList.remove('active'));
            event.target.classList.add('active');
            
            rows.forEach(row => {{
                if (filter === 'all') {{
                    row.style.display = '';
                }} else if (filter === 'pass') {{
                    row.style.display = row.classList.contains('pass') ? '' : 'none';
                }} else if (filter === 'fail') {{
                    row.style.display = row.classList.contains('fail') ? '' : 'none';
                }}
            }});
        }}
    </script>
</body>
</html>"""
        
        with open(output_path, 'w', encoding='utf-8') as f:
            f.write(html)


def extract_xml_from_zip(zip_file: str, output_dir: str) -> List[str]:
    """Extract XML files from a ZIP archive"""
    xml_files = []
    
    with zipfile.ZipFile(zip_file, 'r') as zf:
        for name in zf.namelist():
            if name.endswith('.xml') and not name.startswith('__MACOSX'):
                zf.extract(name, output_dir)
                xml_files.append(os.path.join(output_dir, name))
    
    return xml_files


def main():
    parser = argparse.ArgumentParser(
        description='ISO 20022 Batch XML Validator',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  # Validate all XML files in a folder
  python batch_validator.py schema.xsd folder/*.xml
  
  # Validate XML files from a ZIP
  python batch_validator.py schema.xsd messages.zip
  
  # Generate Excel report
  python batch_validator.py schema.xsd *.xml -o report.xlsx
  
  # Generate HTML dashboard
  python batch_validator.py schema.xsd *.xml -o report.html
        """
    )
    
    parser.add_argument('xsd_file', help='XSD schema file')
    parser.add_argument('xml_files', nargs='+', help='XML files, folder, or ZIP to validate')
    parser.add_argument('-o', '--output', help='Output file (HTML or Excel based on extension)')
    parser.add_argument('--json', action='store_true', help='Output as JSON')
    parser.add_argument('--workers', type=int, default=4, help='Number of parallel workers (default: 4)')
    
    args = parser.parse_args()
    
    # Validate XSD exists
    if not Path(args.xsd_file).exists():
        print(f"❌ Error: XSD file not found: {args.xsd_file}")
        return
    
    # Collect XML files
    xml_files = []
    
    for path in args.xml_files:
        if path.endswith('.zip'):
            import tempfile
            temp_dir = tempfile.mkdtemp()
            xml_files.extend(extract_xml_from_zip(path, temp_dir))
        elif os.path.isdir(path):
            xml_files.extend([os.path.join(path, f) for f in os.listdir(path) 
                            if f.endswith('.xml')])
        elif path.endswith('.xml') and os.path.exists(path):
            xml_files.append(path)
        else:
            # Try glob pattern
            import glob
            xml_files.extend(glob.glob(path))
    
    if not xml_files:
        print("❌ Error: No XML files found")
        return
    
    print(f"\n{'='*70}")
    print("ISO 20022 BATCH XML VALIDATOR")
    print(f"{'='*70}\n")
    print(f"📋 XSD: {args.xsd_file}")
    print(f"📁 Files to validate: {len(xml_files)}")
    
    if not HAS_LXML:
        print("\n⚠️  Warning: lxml not installed - using basic validation only")
    
    print(f"\n⏳ Validating...")
    
    # Create validator
    validator = BatchXMLValidator(args.xsd_file)
    
    # Progress callback
    def progress(completed, total, filename):
        pct = completed / total * 100
        bar_len = 40
        filled = int(bar_len * completed / total)
        bar = '█' * filled + '░' * (bar_len - filled)
        status = "✅" if validator.results[-1].valid else "❌"
        print(f"\r   [{bar}] {pct:5.1f}% ({completed}/{total}) {status} {filename[:40]}", end='', flush=True)
    
    # Run validation
    result = validator.validate_batch(xml_files, max_workers=args.workers, 
                                      progress_callback=progress)
    
    print("\n")
    
    # Output results
    if args.json:
        print(json.dumps(result, indent=2))
    else:
        print(f"{'='*70}")
        print("VALIDATION RESULTS")
        print(f"{'='*70}")
        print(f"\n📊 Summary:")
        print(f"   Total Files:  {result['summary']['total_files']}")
        print(f"   Passed:       {result['summary']['passed']} ✅")
        print(f"   Failed:       {result['summary']['failed']} ❌")
        print(f"   Pass Rate:    {result['summary']['pass_rate']}")
        print(f"   Total Errors: {result['summary']['total_errors']}")
        
        # Show failed files
        failed_results = [r for r in result['results'] if not r['valid']]
        if failed_results:
            print(f"\n❌ Failed Files:")
            for r in failed_results[:10]:
                print(f"   • {r['filename']} ({r['error_count']} errors)")
                for err in r['errors'][:2]:
                    print(f"     - {err.get('message', '')[:80]}")
            if len(failed_results) > 10:
                print(f"   ... and {len(failed_results) - 10} more")
    
    # Generate reports
    if args.output:
        if args.output.endswith('.xlsx'):
            validator.generate_excel_report(args.output)
            print(f"\n📁 Excel report saved to: {args.output}")
        elif args.output.endswith('.html'):
            validator.generate_html_report(args.output)
            print(f"\n📁 HTML report saved to: {args.output}")
        elif args.output.endswith('.json'):
            with open(args.output, 'w') as f:
                json.dump(result, f, indent=2)
            print(f"\n📁 JSON report saved to: {args.output}")
    
    print(f"\n{'='*70}")
    print("COMPLETE")
    print(f"{'='*70}\n")
    
    # Return exit code based on results
    sys.exit(0 if result['summary']['failed'] == 0 else 1)


if __name__ == '__main__':
    main()
