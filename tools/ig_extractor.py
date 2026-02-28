"""
Rulebook IG Extractor
=====================
Parses ISO 20022 Inter-PSP Implementation Guideline PDFs (EPC/NPC format)
and produces a structured Excel workbook.

One sheet per message section (pacs.008, pacs.002, camt.056, …).
Each field = ONE row with columns:
  #  |  Multiplicity  |  XPath  |  ISO Name  |  ISO Definition  |  XML Tag  |  Type
  |  ISO Length  |  SEPA/NPC Length  |  SEPA/NPC Usage Rules  |  SEPA/NPC Rulebook
  |  SEPA/NPC Format Rules  |  SEPA/NPC FractDigits  |  SEPA/NPC Inclusive
  |  SEPA/NPC Code Restrictions

Usage (CLI):
    python ig_extractor.py <input.pdf> [-o output.xlsx] [--messages pacs.008 pacs.002]
    
Usage (API):
    from ig_extractor import extract_ig
    result = extract_ig("path/to/ig.pdf", "output.xlsx")
"""

import re
import sys
from pathlib import Path
from typing import Optional

import pdfplumber
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ── Colours ──────────────────────────────────────────────────────────────────
NAVY        = "FF1F3864"
BLUE        = "FF2E74B5"
LIGHT_BLUE  = "D9E1F2"
ALT_ROW     = "EBF0FA"
WHITE       = "FFFFFFFF"
YELLOW_SEPA = "FFF2CC"   # yellow = SEPA core mandatory
WHITE_ROW   = "FFFFFF"   # white  = optional / AOS
RED_ROW     = "FFCCCC"   # red    = NOT to be used in SEPA

# PDF rect RGB values (rounded to 2dp for matching)
_PDF_YELLOW     = (0.98, 0.99, 0.37)   # EPC Core Mandatory
_PDF_RED        = (1.0,  0.0,  0.0 )   # EPC Not Permitted
_PDF_YELLOW_NPC = (1.00, 0.96, 0.80)   # NPC Core Mandatory (soft yellow)
_PDF_RED_NPC    = (0.97, 0.57, 0.63)   # NPC Not Permitted (pink-red)

def _color_match(nsc, target, tol=0.05):
    """Check if an RGB tuple matches a target within tolerance."""
    return (isinstance(nsc, tuple) and len(nsc) == 3 and
            all(abs(nsc[i] - target[i]) < tol for i in range(3)))


# ── Label patterns in col-4 (requirements column) ────────────────────────────
# Order: longest/most specific first to prevent partial matches
_LABEL_PATTERNS = [
    r'SEPA Usage Rule\(s\)',
    r'SEPA Format Rule\(s\)',
    r'SEPA Code Restrictions',
    r'SEPA FractDigits',
    r'SEPA Inclusive',
    r'SEPA Rulebook',
    r'SEPA Length',
    r'NPC Usage Rule\(s\)',
    r'NPC Format Rule\(s\)',
    r'NPC Code Restrictions',
    r'NPC FractDigits',
    r'NPC Inclusive',
    r'NPC Rulebook',
    r'NPC Length',
    r'ISO Definition',
    r'ISO Length',
    r'ISO Name',
    r'XML Tag',
    r'Type(?=\s+[A-Z])',        # "Type" followed by a capital = type name
]
_SPLIT_RE = re.compile('(' + '|'.join(_LABEL_PATTERNS) + r')\s+')

# Normalise label name (remove regex look-ahead etc.)
def _norm_label(raw: str) -> str:
    return raw.strip()


def _parse_col4(text: str) -> dict:
    """
    Split a col-4 requirements block into {label: value} pairs.
    Any text before the first recognised label goes into 'Rules' (free-form SEPA/NPC rules).
    Multi-line continuation text is already joined before calling this function.
    """
    if not text:
        return {}
    text = re.sub(r'\s+', ' ', text).strip()

    tokens = _SPLIT_RE.split(text)
    result = {}

    # tokens = [pre_text, label, value, label, value, ...]
    pre = (tokens[0] or '').strip()
    if pre:
        result['Rules'] = pre          # text before first label

    i = 1
    while i < len(tokens) - 1:
        raw_label = tokens[i]
        value = re.sub(r'\s+', ' ', (tokens[i + 1] or '')).strip()
        label = _norm_label(raw_label)
        # Merge if same label seen twice (e.g. multi-block SEPA Usage Rules)
        if label in result:
            result[label] = result[label] + ' ' + value
        else:
            result[label] = value
        i += 2

    return result


# ── Path builder ──────────────────────────────────────────────────────────────
def _build_xpath(col3_lines: list[str]) -> str:
    """
    Convert the multi-row col-3 path into a single XPath string.

    Col-3 uses +/++/+++ prefix to denote depth:
        FITo FICustomer Credit Transfer V08   → root (depth 0)
        +Group Header                          → depth 1
        ++Settlement Information               → depth 2
        +++Instructing Reimbursement Agent     → depth 3  (may wrap!)
        Account                                → continuation of depth 3 (no +)

    Continuation lines (no leading +, not the root line) are joined to the
    previous line that had a + prefix, fixing wrapped long element names.
    """
    # First pass: join continuation lines to their preceding + line
    merged: list[str] = []
    for line in col3_lines:
        if not line or not line.strip():
            continue
        stripped = line.strip()
        has_plus = stripped.startswith('+')
        is_root  = not stripped.startswith('+') and len(merged) == 0

        if has_plus or is_root:
            merged.append(stripped)
        else:
            # Continuation — join to the last line
            if merged:
                merged[-1] = merged[-1] + ' ' + stripped
            else:
                merged.append(stripped)

    if not merged:
        return ''

    # Second pass: build hierarchy from + counts
    path_parts = []   # (depth, name)
    for line in merged:
        depth = len(line) - len(line.lstrip('+'))
        name  = line.lstrip('+').strip()
        if name:
            path_parts.append((depth, name))

    if not path_parts:
        return ''

    # Build stack to reconstruct full path
    stack = []
    for depth, name in path_parts:
        while stack and stack[-1][0] >= depth:
            stack.pop()
        stack.append((depth, name))

    return '/'.join(name for _, name in stack)


# ── Section detector ──────────────────────────────────────────────────────────
_SECTION_RE = re.compile(
    r'((?:\d+\.)+\d+)[\.\s]+Use of\s+(.+?)\s*\(([a-z]{2,8}\.\d{3}\.\d{3}\.\d{2,3})\)',
    re.IGNORECASE,
)


def _find_sections(pdf: pdfplumber.PDF) -> list[dict]:
    """
    Detect all 'Use of <message> (pacs/camt)' sections and their page ranges.
    Returns list of dicts: {section_number, message, label, page_start, page_end, section_id, sheet_name}
    """
    found = []
    for i, page in enumerate(pdf.pages, start=1):
        text = page.extract_text() or ''
        for m in _SECTION_RE.finditer(text):
            surrounding = text[max(0, m.start()-10):m.end()+20]
            if re.search(r'\.{5,}', surrounding):
                continue   # skip TOC dotted entries
            found.append({
                'section_number': m.group(1).strip(),
                'label':          m.group(2).strip(),
                'message':        m.group(3).strip().lower(),
                'page_start':     i,
                'page_end':       None,
            })

    # Deduplicate by (message, page_start)
    seen = set()
    unique = []
    for s in found:
        key = (s['message'], s['page_start'])
        if key not in seen:
            seen.add(key)
            unique.append(s)
    found = unique

    # Assign page_end
    for idx in range(len(found)):
        if idx + 1 < len(found):
            found[idx]['page_end'] = found[idx + 1]['page_start'] - 1
        else:
            found[idx]['page_end'] = len(pdf.pages)

    # Build unique sheet names
    msg_count: dict[str, int] = {}
    for s in found:
        msg = s['message']
        msg_count[msg] = msg_count.get(msg, 0) + 1
        s['section_id'] = msg_count[msg]
        base = msg.replace('.', '_')
        s['sheet_name'] = base if msg_count[msg] == 1 else f"{base}_{msg_count[msg]}"

    return found


# ── Record builder ────────────────────────────────────────────────────────────
def _page_color_ranges(page) -> dict:
    """
    Return dict of color-name → list of (top, bottom) Y ranges on the page.
    Detects: 'yellow' (SEPA core mandatory) and 'red' (not permitted in SEPA).
    """
    ranges = {'yellow': [], 'red': []}
    for r in page.rects:
        nsc = r.get('non_stroking_color')
        if (_color_match(nsc, _PDF_YELLOW) or
                _color_match(nsc, _PDF_YELLOW_NPC, tol=0.04)):
            ranges['yellow'].append((r['top'], r['bottom']))
        elif (_color_match(nsc, _PDF_RED, tol=0.02) or
                _color_match(nsc, _PDF_RED_NPC, tol=0.04)):
            ranges['red'].append((r['top'], r['bottom']))
    return ranges


def _row_color(row_obj, color_ranges: dict) -> str:
    """
    Return 'yellow', 'red', or 'white' for a pdfplumber row object.
    """
    if not row_obj.cells or not row_obj.cells[0]:
        return 'white'
    cell0 = row_obj.cells[0]
    if cell0 is None:
        return 'white'
    row_mid = (cell0[1] + cell0[3]) / 2
    if any(yr[0] <= row_mid <= yr[1] for yr in color_ranges.get('red', [])):
        return 'red'
    if any(yr[0] <= row_mid <= yr[1] for yr in color_ranges.get('yellow', [])):
        return 'yellow'
    return 'white'


def _collect_records(pdf: pdfplumber.PDF, page_start: int, page_end: int) -> list[dict]:
    """
    Walk pages page_start..page_end, collect all 4-column IG table rows,
    group them into per-field records, and return a list of parsed dicts.
    Each record carries a 'row_color': 'yellow' | 'red' | 'white'.
    """
    # 1. Gather raw rows, tagged with color
    raw_rows = []  # each entry: (row_list_4cols, color_str)
    for pg_idx in range(page_start - 1, page_end):
        page = pdf.pages[pg_idx]
        color_ranges = _page_color_ranges(page)

        found_tables = page.find_tables()
        extracted    = page.extract_tables()

        for t_idx, tbl_obj in enumerate(found_tables):
            if t_idx >= len(extracted):
                continue
            raw_table = extracted[t_idx]
            if not raw_table or len(raw_table[0]) < 4:
                continue

            for r_idx, row_obj in enumerate(tbl_obj.rows):
                if r_idx >= len(raw_table):
                    continue
                row = raw_table[r_idx]
                c0 = (row[0] or '').strip()
                c1 = (row[1] or '').strip()
                if c0 == '#' or c1 == 'Mult':
                    continue
                row_col = _row_color(row_obj, color_ranges)
                raw_rows.append(([row[i] if i < len(row) else None for i in range(4)], row_col))

    # 2. Group rows into records.
    #    Boundary: col0 has a field index (e.g. "1.6").
    #    Color priority: red > yellow > white (if ANY row in the block is red, record = red)
    COLOR_PRIORITY = {'red': 2, 'yellow': 1, 'white': 0}

    records = []
    current = None

    for (row, row_col) in raw_rows:
        c0, c1, c2, c3 = [(row[i] or '').strip() for i in range(4)]

        is_new_record = bool(re.match(r'^\d+\.\d+', c0))
        is_choice_row = 'xs:choice' in (c3 or '')

        if is_new_record:
            if current is not None:
                records.append(current)
            current = {
                'index':      c0,
                'mult':       c1,
                'col3_lines': [c2] if c2 else [],
                'col4_text':  c3,
                'row_color':  row_col,
            }
        elif is_choice_row:
            if current is not None:
                records.append(current)
            records.append({
                'index':      '',
                'mult':       c1,
                'col3_lines': [c2] if c2 else [],
                'col4_text':  'XML Tag xs:choice',
                'row_color':  row_col,
                '_choice':    True,
            })
            current = None
        elif current is not None:
            if c2:
                current['col3_lines'].append(c2)
            if c3:
                current['col4_text'] = (current['col4_text'] or '') + ' ' + c3
            # Escalate colour: red overrides yellow overrides white
            if COLOR_PRIORITY.get(row_col, 0) > COLOR_PRIORITY.get(current['row_color'], 0):
                current['row_color'] = row_col

    if current is not None:
        records.append(current)

    # 3. Parse each record into a flat dict
    parsed = []
    for rec in records:
        xpath = _build_xpath(rec.get('col3_lines', []))
        col4  = _parse_col4(rec.get('col4_text', ''))

        element_name = xpath.split('/')[-1] if xpath else col4.get('ISO Name', '')

        row_dict = {
            'Index':              rec.get('index', ''),
            'Multiplicity':       rec.get('mult', ''),
            'XPath':              xpath,
            'Element Name':       element_name,
            'ISO Name':           col4.get('ISO Name', ''),
            'ISO Definition':     col4.get('ISO Definition', ''),
            'XML Tag':            col4.get('XML Tag', ''),
            'Type':               col4.get('Type', ''),
            'ISO Length':         col4.get('ISO Length', ''),
            'SEPA/NPC Length':    col4.get('SEPA Length', col4.get('NPC Length', '')),
            'SEPA/NPC Usage Rules':       col4.get('SEPA Usage Rule(s)', col4.get('NPC Usage Rule(s)', col4.get('Rules', ''))),
            'SEPA/NPC Rulebook':          col4.get('SEPA Rulebook', col4.get('NPC Rulebook', '')),
            'SEPA/NPC Format Rules':      col4.get('SEPA Format Rule(s)', col4.get('NPC Format Rule(s)', '')),
            'SEPA/NPC FractDigits':       col4.get('SEPA FractDigits', col4.get('NPC FractDigits', '')),
            'SEPA/NPC Inclusive':         col4.get('SEPA Inclusive', col4.get('NPC Inclusive', '')),
            'SEPA/NPC Code Restrictions': col4.get('SEPA Code Restrictions', col4.get('NPC Code Restrictions', '')),
            'row_color':   rec.get('row_color', 'white'),  # 'yellow'|'red'|'white'
            # keep backward compat
            'sepa_core':   rec.get('row_color', 'white') == 'yellow',
            '_is_choice':  rec.get('_choice', False),
        }
        parsed.append(row_dict)

    return parsed


# ── Excel writer ──────────────────────────────────────────────────────────────
COLUMNS = [
    'Index', 'Multiplicity', 'XPath', 'Element Name',
    'ISO Name', 'ISO Definition', 'XML Tag', 'Type',
    'ISO Length', 'SEPA/NPC Length',
    'SEPA/NPC Usage Rules', 'SEPA/NPC Rulebook',
    'SEPA/NPC Format Rules', 'SEPA/NPC FractDigits',
    'SEPA/NPC Inclusive', 'SEPA/NPC Code Restrictions',
]

COL_WIDTHS = {
    'Index': 8, 'Multiplicity': 12, 'XPath': 55, 'Element Name': 30,
    'ISO Name': 30, 'ISO Definition': 60, 'XML Tag': 28, 'Type': 32,
    'ISO Length': 12, 'SEPA/NPC Length': 14,
    'SEPA/NPC Usage Rules': 55, 'SEPA/NPC Rulebook': 45,
    'SEPA/NPC Format Rules': 40, 'SEPA/NPC FractDigits': 16,
    'SEPA/NPC Inclusive': 25, 'SEPA/NPC Code Restrictions': 45,
}

def _thin():
    s = Side(style='thin', color='BFC9E0')
    return Border(top=s, bottom=s, left=s, right=s)

def _thick():
    s = Side(style='medium', color=NAVY)
    return Border(top=s, bottom=s, left=s, right=s)


def _write_sheet(ws, section: dict, records: list[dict]):
    """Write one message section to a worksheet."""

    # ── Title row ─────────────────────────────────────────────────────────────
    ws.merge_cells(f"A1:{get_column_letter(len(COLUMNS))}1")
    title = ws['A1']
    title.value = f"{section['message'].upper()}  —  {section['label']}"
    title.font      = Font(name='Arial', bold=True, size=13, color=WHITE)
    title.fill      = PatternFill('solid', fgColor=NAVY)
    title.alignment = Alignment(horizontal='left', vertical='center')
    ws.row_dimensions[1].height = 22

    # ── Legend row ────────────────────────────────────────────────────────────
    total_cols = len(COLUMNS)
    col3 = get_column_letter(total_cols // 3)
    col6 = get_column_letter((total_cols * 2) // 3)
    
    ws.merge_cells(f"A2:{col3}2")
    leg_y = ws['A2']
    leg_y.value = '🟡 Yellow = SEPA/NPC Core Mandatory'
    leg_y.font  = Font(name='Arial', bold=True, size=9, color='7D6600')
    leg_y.fill  = PatternFill('solid', fgColor='FFFFF2CC')
    leg_y.alignment = Alignment(horizontal='left', vertical='center', indent=1)

    ws.merge_cells(f"{get_column_letter(total_cols//3+1)}2:{col6}2")
    leg_w = ws[f"{get_column_letter(total_cols//3+1)}2"]
    leg_w.value = '⬜ White = Optional / AOS'
    leg_w.font  = Font(name='Arial', bold=True, size=9, color='444444')
    leg_w.fill  = PatternFill('solid', fgColor='FFF2F2F2')
    leg_w.alignment = Alignment(horizontal='left', vertical='center', indent=1)

    ws.merge_cells(f"{get_column_letter(total_cols*2//3+1)}2:{get_column_letter(total_cols)}2")
    leg_r = ws[f"{get_column_letter(total_cols*2//3+1)}2"]
    leg_r.value = '🔴 Red = NOT permitted in SEPA/NPC payments'
    leg_r.font  = Font(name='Arial', bold=True, size=9, color='8B0000')
    leg_r.fill  = PatternFill('solid', fgColor='FFFFCCCC')
    leg_r.alignment = Alignment(horizontal='left', vertical='center', indent=1)
    ws.row_dimensions[2].height = 16

    # ── Column headers ─────────────────────────────────────────────────────────
    for col_i, col_name in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=3, column=col_i, value=col_name)
        cell.font      = Font(name='Arial', bold=True, size=10, color=WHITE)
        cell.fill      = PatternFill('solid', fgColor=BLUE)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border    = _thick()
    ws.row_dimensions[3].height = 24
    ws.freeze_panes = 'A4'

    # ── Data rows ──────────────────────────────────────────────────────────────
    data_row_num = 4
    for rec in records:
        if rec.get('_is_choice'):
            continue    # skip xs:choice structural markers

        # Row background mirrors PDF: yellow=core mandatory, red=not permitted, white=optional
        row_color = rec.get('row_color', 'white')
        if row_color == 'yellow':
            bg_hex = 'FFFFF2CC'  # warm yellow (FF = full opacity)
        elif row_color == 'red':
            bg_hex = 'FFFFCCCC'  # light red (FF = full opacity)
        else:
            bg_hex = 'FFFFFFFF'  # white (FF = full opacity)
        row_fill = PatternFill('solid', fgColor=bg_hex)

        for col_i, col_name in enumerate(COLUMNS, start=1):
            val = rec.get(col_name, '') or ''
            cell = ws.cell(row=data_row_num, column=col_i, value=val)
            cell.font      = Font(name='Arial', size=9)
            cell.border    = _thin()
            cell.alignment = Alignment(
                horizontal='left', vertical='top',
                wrap_text=True,
            )
            cell.fill = row_fill

        ws.row_dimensions[data_row_num].height = 45
        data_row_num += 1

    # ── Column widths ─────────────────────────────────────────────────────────
    for col_i, col_name in enumerate(COLUMNS, start=1):
        ws.column_dimensions[get_column_letter(col_i)].width = COL_WIDTHS.get(col_name, 20)


def _write_summary(ws, sections: list[dict], pdf_name: str, records_by_sheet: dict):
    """Write a Summary sheet listing all message sections."""
    ws.merge_cells("A1:G1")
    ws['A1'].value = f"ISO 20022 IG Extractor — {pdf_name}"
    ws['A1'].font      = Font(name='Arial', bold=True, size=14, color=WHITE)
    ws['A1'].fill      = PatternFill('solid', fgColor=NAVY)
    ws['A1'].alignment = Alignment(horizontal='left', vertical='center')
    ws.row_dimensions[1].height = 24

    # Legend
    ws.merge_cells("A2:D2")
    lc = ws['A2']
    lc.value = '🟡 Yellow rows = SEPA/NPC Core Mandatory    ⬜ White rows = Optional / AOS'
    lc.font  = Font(name='Arial', size=9, bold=True, color='7D6600')
    lc.fill  = PatternFill('solid', fgColor='FFFFF2CC')
    lc.alignment = Alignment(horizontal='left', vertical='center', indent=1)
    ws.row_dimensions[2].height = 14

    headers = ['Sheet', 'Message', 'Section Label', 'Total Fields', 'Core (Yellow)', 'Optional (White)', 'Not Permitted (Red)', 'Pages']
    for col_i, h in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=col_i, value=h)
        cell.font  = Font(name='Arial', bold=True, size=10, color=WHITE)
        cell.fill  = PatternFill('solid', fgColor=BLUE)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = _thick()
    ws.row_dimensions[4].height = 20

    for row_i, s in enumerate(sections, start=5):
        recs   = records_by_sheet.get(s['sheet_name'], [])
        total  = len(recs)
        core   = sum(1 for r in recs if r.get('row_color') == 'yellow')
        red    = sum(1 for r in recs if r.get('row_color') == 'red')
        opt    = total - core - red
        data = [
            s['sheet_name'], s['message'], s['label'],
            total, core, opt, red,
            f"{s['page_start']}–{s['page_end']}",
        ]
        row_fill = PatternFill('solid', fgColor='FFF9F9F9') if row_i % 2 == 0 else None
        for col_i, val in enumerate(data, start=1):
            cell = ws.cell(row=row_i, column=col_i, value=val)
            cell.font   = Font(name='Arial', size=9)
            cell.border = _thin()
            cell.alignment = Alignment(horizontal='left', vertical='center')
            if row_fill:
                cell.fill = row_fill
        ws.row_dimensions[row_i].height = 16

    for col_letter, width in zip('ABCDEFGH', [22, 22, 45, 14, 14, 16, 18, 12]):
        ws.column_dimensions[col_letter].width = width


# ── Main public API ──────────────────────────────────────────────────────────
def extract_ig(
    pdf_path: str,
    output_path: Optional[str] = None,
    filter_messages: Optional[list[str]] = None,
    filter_sections: Optional[list[str]] = None,
) -> dict:
    """
    Parse an ISO 20022 IG PDF and write a structured Excel workbook.

    Parameters
    ----------
    pdf_path        : path to the source PDF
    output_path     : destination .xlsx  (default: same stem + .xlsx)
    filter_messages : e.g. ['pacs.008.001.08']  – None = all sections
    filter_sections : e.g. ['2.1.1', '2.2.1']  – filter by section number

    Returns
    -------
    dict with keys: output_file, sections, total_fields
    """
    pdf_path = Path(pdf_path)
    if not pdf_path.exists():
        raise FileNotFoundError(f"PDF not found: {pdf_path}")

    if output_path is None:
        output_path = pdf_path.with_suffix('.xlsx')
    output_path = Path(output_path)

    wb = Workbook()
    summary_ws = wb.active
    summary_ws.title = 'Summary'

    all_sections = []
    records_by_sheet: dict[str, list] = {}

    with pdfplumber.open(str(pdf_path)) as pdf:
        sections = _find_sections(pdf)

        if filter_messages:
            # Support both full IDs (pacs.008.001.08) and prefix/partial (pacs.008)
            fm = [m.lower().strip() for m in filter_messages]
            sections = [
                s for s in sections
                if any(s['message'].lower() == f or s['message'].lower().startswith(f + '.') or f in s['message'].lower()
                       for f in fm)
            ]

        if filter_sections:
            sections = [s for s in sections if s.get('section_number', '') in filter_sections]

        for section in sections:
            records = _collect_records(pdf, section['page_start'], section['page_end'])
            data_records = [r for r in records if r.get('Index') or r.get('XML Tag') or r.get('ISO Name')]

            ws = wb.create_sheet(title=section['sheet_name'])
            _write_sheet(ws, section, data_records)

            records_by_sheet[section['sheet_name']] = data_records
            all_sections.append(section)

    _write_summary(summary_ws, all_sections, pdf_path.name, records_by_sheet)

    wb.save(output_path)

    total_fields = sum(len(v) for v in records_by_sheet.values())
    return {
        'output_file':   str(output_path),
        'sections':      [
            {
                'sheet':   s['sheet_name'],
                'section': s.get('section_number', ''),
                'message': s['message'],
                'label':   s['label'],
                'fields':  len(records_by_sheet.get(s['sheet_name'], [])),
                'pages':   f"{s['page_start']}-{s['page_end']}",
            }
            for s in all_sections
        ],
        'total_fields': total_fields,
    }


# ── Public section-detection helper ──────────────────────────────────────────
def detect_sections(pdf_path: str) -> list[dict]:
    """
    Open a PDF and return a list of detectable IG sections — suitable for
    driving the section-picker UI.

    Each item:
        id          – section_number e.g. '2.1.1'  (passed back as filter_sections)
        label       – human label   e.g. 'Use of pacs.008.001.08'
        title       – message ID    e.g. 'pacs.008.001.08'
        page_start  – first page of the section
        page_end    – last  page of the section
        field_count – 0 (fast scan; full count only available after extraction)
    """
    import pdfplumber
    with pdfplumber.open(pdf_path) as pdf:
        sections = _find_sections(pdf)

    return [
        {
            'id':          s['section_number'],
            'label':       s['label'],
            'title':       s['message'],
            'page_start':  s['page_start'],
            'page_end':    s['page_end'],
            'field_count': 0,          # lightweight scan — no record parsing
        }
        for s in sections
    ]


# ── CLI ───────────────────────────────────────────────────────────────────────
if __name__ == '__main__':
    import argparse, json

    parser = argparse.ArgumentParser(description='ISO 20022 IG PDF → Excel extractor')
    parser.add_argument('pdf',  help='Input PDF path')
    parser.add_argument('-o',   '--output', default=None, help='Output .xlsx path')
    parser.add_argument('--messages', nargs='*', default=None,
                        help='Filter specific messages e.g. pacs.008.001.08 camt.056.001.08')
    parser.add_argument('--sections', nargs='*', default=None,
                        help='Filter specific section numbers e.g. 2.1.1 2.2.1')
    args = parser.parse_args()

    result = extract_ig(args.pdf, args.output, args.messages, args.sections)
    print(json.dumps(result, indent=2))
