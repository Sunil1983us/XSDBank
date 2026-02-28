#!/usr/bin/env python3
"""
YAML API Schema Extractor
==========================
Parses OpenAPI / AsyncAPI YAML files and produces a structured Excel workbook.

One sheet per endpoint (path + method), with full deep schema expansion —
including nested $ref resolution, allOf/anyOf/oneOf merging, arrays, and
polymorphic types.

Each field = ONE row with columns:
  #  |  Path  |  Field Name  |  Full Path  |  Type  |  Format  |  Required
  |  Nullable  |  Description  |  Enum Values  |  Example  |  Default
  |  Min  |  Max  |  Pattern  |  Read Only  |  Write Only  |  Deprecated

Usage (CLI):
    python yaml_api_extractor.py <input.yaml> [-o output.xlsx] [--endpoints GET:/pets POST:/pets]

Usage (API):
    from yaml_api_extractor import extract_yaml_api, detect_endpoints
    endpoints = detect_endpoints("spec.yaml")
    result    = extract_yaml_api("spec.yaml", "output.xlsx", filter_endpoints=["GET:/pets"])
"""

from __future__ import annotations
import argparse
import copy
import re
import sys
from pathlib import Path
from typing import Optional
from datetime import datetime

try:
    import yaml
except ImportError:
    raise ImportError("PyYAML is required: pip install pyyaml")

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    raise ImportError("openpyxl is required: pip install openpyxl")


# ── Colour palette (matches toolkit style) ────────────────────────────────────
NAVY          = "FF1F3864"
BLUE          = "FF2E75B6"
TEAL          = "FF0E7490"
LIGHT_BLUE    = "FFDBEAFE"
LIGHT_TEAL    = "FFD0F4F1"
LIGHT_GREEN   = "FFD1FAE5"
LIGHT_YELLOW  = "FFFEF9C3"
LIGHT_RED     = "FFFEE2E2"
ALT_ROW       = "FFF0F7FF"
WHITE         = "FFFFFFFF"
GRAY_HEADER   = "FFF1F5F9"
ORANGE_LIGHT  = "FFFEF3C7"

METHOD_COLORS = {
    "GET":     ("FF16A34A", "FFD1FAE5"),   # green text, green bg
    "POST":    ("FF2563EB", "FFDBEAFE"),   # blue text, blue bg
    "PUT":     ("FFD97706", "FFFEF3C7"),   # amber text, amber bg
    "PATCH":   ("FF9333EA", "FFFAE8FF"),   # purple text, purple bg
    "DELETE":  ("FFDC2626", "FFFEE2E2"),   # red text, red bg
    "HEAD":    ("FF0891B2", "FFE0F7FA"),
    "OPTIONS": ("FF64748B", "FFF1F5F9"),
}

SCHEMA_COLS = [
    ("#",           5),
    ("Field Name",  22),
    ("Full Path",   40),
    ("Type",        14),
    ("Format",      14),
    ("Required",    10),
    ("Nullable",    10),
    ("Description", 50),
    ("Enum Values", 35),
    ("Example",     22),
    ("Default",     18),
    ("Min",         10),
    ("Max",         10),
    ("Pattern",     28),
    ("Read Only",    10),
    ("Write Only",   10),
    ("Deprecated",   10),
]

COL_NAMES  = [c[0] for c in SCHEMA_COLS]
COL_WIDTHS = [c[1] for c in SCHEMA_COLS]


# ── $ref resolver ──────────────────────────────────────────────────────────────
class RefResolver:
    """Resolve JSON-pointer $ref values within a single YAML document."""

    def __init__(self, root: dict):
        self._root = root

    def resolve(self, ref: str) -> dict:
        """Follow a #/components/... ref and return the target dict."""
        if not ref.startswith("#/"):
            return {}
        parts = ref.lstrip("#/").split("/")
        node = self._root
        for part in parts:
            part = part.replace("~1", "/").replace("~0", "~")
            if isinstance(node, dict):
                node = node.get(part, {})
            elif isinstance(node, list):
                try:
                    node = node[int(part)]
                except (ValueError, IndexError):
                    return {}
            else:
                return {}
        return node if isinstance(node, dict) else {}

    def deref(self, schema: dict, depth: int = 0, _seen: set | None = None, _seen_refs: set | None = None) -> dict:
        """
        Fully dereference a schema dict, merging allOf/anyOf/oneOf,
        resolving $ref, and returning a flat merged dict.
        Depth-limited to avoid infinite recursion on circular schemas.
        """
        if _seen is None:
            _seen = _seen_refs if _seen_refs else set()
        if depth > 20:
            return schema

        schema = copy.deepcopy(schema)

        # Resolve top-level $ref
        ref = schema.get("$ref")
        if ref and ref not in _seen:
            _seen = _seen | {ref}
            resolved = self.deref(self.resolve(ref), depth + 1, _seen)
            # Merge resolved with any sibling keys (description override, etc.)
            resolved.update({k: v for k, v in schema.items() if k != "$ref"})
            schema = resolved

        # Merge allOf
        for kw in ("allOf", "anyOf", "oneOf"):
            if kw in schema:
                parts = schema.pop(kw)
                for part in (parts or []):
                    merged = self.deref(part, depth + 1, _seen)
                    # Merge properties
                    if "properties" in merged:
                        schema.setdefault("properties", {}).update(merged["properties"])
                    if "required" in merged:
                        existing = schema.get("required", [])
                        schema["required"] = list(set(existing) | set(merged.get("required", [])))
                    # Copy other keys only if not already present
                    for k, v in merged.items():
                        if k not in ("properties", "required") and k not in schema:
                            schema[k] = v

        return schema


# ── Schema walker ──────────────────────────────────────────────────────────────
class SchemaWalker:
    """Walk a (possibly deeply nested) JSON Schema and produce flat row dicts."""

    def __init__(self, resolver: RefResolver):
        self.resolver = resolver

    def walk(
        self,
        schema: dict,
        parent_path: str = "",
        required_set: set | None = None,
        depth: int = 0,
        _seen_refs: set | None = None,
    ) -> list[dict]:
        """
        Recursively walk schema and return list of row dicts.
        """
        rows: list[dict] = []
        if _seen_refs is None:
            _seen_refs = set()
        if depth > 30:
            return rows

        schema = self.resolver.deref(schema, _seen_refs=_seen_refs)
        if not schema:
            return rows

        schema_type = schema.get("type", "")
        props = schema.get("properties", {})

        if props:
            req_here = set(schema.get("required", []))
            for field_name, field_schema in props.items():
                field_schema = self.resolver.deref(field_schema, depth + 1, _seen_refs)
                full_path = f"{parent_path}.{field_name}" if parent_path else field_name
                is_required = field_name in (required_set or set()) or field_name in req_here

                row = self._make_row(field_name, full_path, field_schema, is_required, depth)
                rows.append(row)

                # Recurse into objects
                child_type = field_schema.get("type", "")
                if child_type == "object" or field_schema.get("properties"):
                    rows.extend(self.walk(
                        field_schema, full_path,
                        required_set=set(field_schema.get("required", [])),
                        depth=depth + 1,
                        _seen_refs=_seen_refs,
                    ))
                elif child_type == "array":
                    items = field_schema.get("items", {})
                    if items:
                        items = self.resolver.deref(items, depth + 1, _seen_refs)
                        items_type = items.get("type", "")
                        # Add a "[]" row to indicate array items
                        arr_path = full_path + "[]"
                        arr_row = self._make_row(
                            field_name + "[]", arr_path, items, False, depth + 1
                        )
                        arr_row["_is_array_item"] = True
                        rows.append(arr_row)
                        if items_type == "object" or items.get("properties"):
                            rows.extend(self.walk(
                                items, arr_path,
                                required_set=set(items.get("required", [])),
                                depth=depth + 2,
                                _seen_refs=_seen_refs,
                            ))

        elif schema_type == "array":
            items = schema.get("items", {})
            if items:
                items = self.resolver.deref(items, depth + 1, _seen_refs)
                arr_path = (parent_path + "[]") if parent_path else "[]"
                arr_row = self._make_row(
                    (parent_path or "items") + "[]", arr_path, items, False, depth
                )
                arr_row["_is_array_item"] = True
                rows.append(arr_row)
                rows.extend(self.walk(
                    items, arr_path,
                    required_set=set(items.get("required", [])),
                    depth=depth + 1,
                    _seen_refs=_seen_refs,
                ))

        return rows

    def _make_row(
        self,
        field_name: str,
        full_path: str,
        schema: dict,
        is_required: bool,
        depth: int,
    ) -> dict:
        enum_vals = schema.get("enum", [])
        enum_str = " | ".join(str(e) for e in enum_vals) if enum_vals else ""

        # Min / Max — handle both number ranges and string length
        minimum = schema.get("minimum", schema.get("minLength", ""))
        maximum = schema.get("maximum", schema.get("maxLength", ""))

        example = schema.get("example", schema.get("x-example", ""))
        if isinstance(example, (dict, list)):
            import json
            example = json.dumps(example, ensure_ascii=False)

        return {
            "field_name":  field_name,
            "full_path":   full_path,
            "type":        schema.get("type", "object" if schema.get("properties") else ""),
            "format":      schema.get("format", ""),
            "required":    "Yes" if is_required else "No",
            "nullable":    "Yes" if schema.get("nullable", False) else "No",
            "description": schema.get("description", schema.get("title", "")),
            "enum":        enum_str,
            "example":     str(example) if example != "" else "",
            "default":     str(schema.get("default", "")) if schema.get("default") is not None else "",
            "minimum":     str(minimum) if minimum != "" else "",
            "maximum":     str(maximum) if maximum != "" else "",
            "pattern":     schema.get("pattern", ""),
            "readOnly":    "Yes" if schema.get("readOnly", False) else "No",
            "writeOnly":   "Yes" if schema.get("writeOnly", False) else "No",
            "deprecated":  "Yes" if schema.get("deprecated", False) else "No",
            "_depth":      depth,
            "_is_array_item": False,
        }


# ── Endpoint detection ─────────────────────────────────────────────────────────
def detect_endpoints(yaml_path: str) -> list[dict]:
    """
    Parse YAML and return list of endpoint dicts for the UI picker.

    Each dict:
        id         – "METHOD:/path"  e.g. "GET:/pets"
        method     – "GET"
        path       – "/pets"
        label      – "GET /pets — List all pets"
        summary    – summary text from spec
        tag        – first tag (if any)
    """
    spec = _load_yaml(yaml_path)
    if not spec:
        return []

    endpoints = []
    paths = spec.get("paths", {})
    for path, path_item in paths.items():
        if not isinstance(path_item, dict):
            continue
        for method in ("get", "post", "put", "patch", "delete", "head", "options"):
            op = path_item.get(method)
            if not isinstance(op, dict):
                continue
            m_upper = method.upper()
            summary  = op.get("summary", op.get("description", ""))
            tags     = op.get("tags", [])
            ep_id    = f"{m_upper}:{path}"
            label    = f"{m_upper}  {path}"
            if summary:
                label += f"  —  {summary[:60]}"
            endpoints.append({
                "id":      ep_id,
                "method":  m_upper,
                "path":    path,
                "label":   label,
                "summary": summary,
                "tag":     tags[0] if tags else "",
            })

    return endpoints


# ── Main extraction ────────────────────────────────────────────────────────────
def extract_yaml_api(
    yaml_path: str,
    output_path: Optional[str] = None,
    filter_endpoints: Optional[list[str]] = None,
) -> dict:
    """
    Parse YAML spec and produce Excel workbook.

    Parameters
    ----------
    yaml_path        : path to input YAML file
    output_path      : path to output .xlsx (auto-generated if None)
    filter_endpoints : list of "METHOD:/path" strings; None = extract all

    Returns
    -------
    dict with keys: success, output_path, message, sheets, total_fields
    """
    spec = _load_yaml(yaml_path)
    if not spec:
        return {"success": False, "error": "Could not parse YAML file"}

    resolver = RefResolver(spec)
    walker   = SchemaWalker(resolver)

    # Determine output path
    if not output_path:
        stem = Path(yaml_path).stem
        output_path = str(Path(yaml_path).parent / f"{stem}_api_schema.xlsx")

    wb = Workbook()
    wb.remove(wb.active)   # remove default sheet

    # ── Summary sheet ──────────────────────────────────────────────────────────
    info = spec.get("info", {})
    api_title   = info.get("title", Path(yaml_path).stem)
    api_version = info.get("version", "")

    all_endpoints = detect_endpoints(yaml_path)
    if filter_endpoints:
        all_endpoints = [e for e in all_endpoints if e["id"] in filter_endpoints]

    if not all_endpoints:
        return {"success": False, "error": "No endpoints found (or none matched filter)"}

    sheets_created = []
    total_fields   = 0

    # Create Summary sheet first
    ws_sum = wb.create_sheet("📋 Summary")
    _write_summary_sheet(ws_sum, spec, all_endpoints, yaml_path)

    # ── One sheet per endpoint ─────────────────────────────────────────────────
    for ep in all_endpoints:
        method = ep["method"]
        path   = ep["path"]
        op     = spec.get("paths", {}).get(path, {}).get(method.lower(), {})
        if not isinstance(op, dict):
            continue

        sheet_name = _safe_sheet_name(f"{method} {path}")
        ws = wb.create_sheet(sheet_name)

        rows = _collect_endpoint_rows(op, resolver, walker)
        total_fields += len(rows)

        _write_endpoint_sheet(ws, ep, op, rows)
        sheets_created.append(sheet_name)

    wb.save(output_path)

    return {
        "success":      True,
        "output_path":  output_path,
        "message":      (
            f"Extracted {len(sheets_created)} endpoint(s) with {total_fields} total fields "
            f"from \"{api_title}\" (v{api_version})"
        ),
        "sheets":        sheets_created,
        "total_fields":  total_fields,
        "api_title":     api_title,
        "api_version":   api_version,
    }


# ── Row collection for one endpoint ───────────────────────────────────────────
def _collect_endpoint_rows(op: dict, resolver: RefResolver, walker: SchemaWalker) -> list[dict]:
    """Gather all parameter + request body + response rows for one operation."""
    rows: list[dict] = []

    # 1. Parameters (path, query, header, cookie)
    params = op.get("parameters", [])
    for param in params:
        param = resolver.deref(param)
        schema = resolver.deref(param.get("schema", {}))
        location = param.get("in", "")
        row = walker._make_row(
            param.get("name", ""),
            f"[{location}] {param.get('name','')}",
            schema,
            param.get("required", False),
            0,
        )
        row["description"] = param.get("description", row["description"])
        row["_section"]    = f"Parameter ({location})"
        rows.append(row)

    # 2. Request body
    req_body = op.get("requestBody", {})
    if req_body:
        req_body = resolver.deref(req_body)
        content  = req_body.get("content", {})
        for media_type, media_obj in content.items():
            schema = resolver.deref(media_obj.get("schema", {}))
            body_rows = walker.walk(schema, parent_path="", required_set=set(schema.get("required", [])))
            for r in body_rows:
                r["_section"] = f"Request Body ({media_type})"
            rows.extend(body_rows)
            break   # use first content type only

    # 3. Responses
    responses = op.get("responses", {})
    for status_code, resp_obj in sorted(responses.items(), key=lambda x: str(x[0])):
        resp_obj = resolver.deref(resp_obj)
        content  = resp_obj.get("content", {})
        if not content:
            # Response with no body — add placeholder row
            rows.append({
                "field_name":  f"[{status_code}]",
                "full_path":   f"Response {status_code}",
                "type":        "",
                "format":      "",
                "required":    "",
                "nullable":    "",
                "description": resp_obj.get("description", ""),
                "enum":        "",
                "example":     "",
                "default":     "",
                "minimum":     "",
                "maximum":     "",
                "pattern":     "",
                "readOnly":    "",
                "writeOnly":   "",
                "deprecated":  "",
                "_depth":      0,
                "_is_array_item": False,
                "_section":    f"Response {status_code}",
            })
            continue
        for media_type, media_obj in content.items():
            schema = resolver.deref(media_obj.get("schema", {}))
            resp_rows = walker.walk(schema, parent_path="", required_set=set(schema.get("required", [])))
            for r in resp_rows:
                r["_section"] = f"Response {status_code} ({media_type})"
            rows.extend(resp_rows)
            break   # use first content type

    return rows


# ── Excel writers ──────────────────────────────────────────────────────────────
def _write_summary_sheet(ws, spec: dict, endpoints: list[dict], yaml_path: str):
    info        = spec.get("info", {})
    api_title   = info.get("title", Path(yaml_path).stem)
    api_version = info.get("version", "")
    description = info.get("description", "")
    base_url    = ""
    for server in spec.get("servers", []):
        base_url = server.get("url", "")
        break

    # Title banner
    ws.merge_cells("A1:H1")
    c = ws["A1"]
    c.value = f"📄 {api_title}"
    c.font  = Font(name="Calibri", size=18, bold=True, color=WHITE)
    c.fill  = PatternFill("solid", fgColor=TEAL)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[1].height = 40

    # Meta rows
    meta = [
        ("Version",    api_version),
        ("Base URL",   base_url),
        ("Description", description),
        ("Source File", Path(yaml_path).name),
        ("Generated",  datetime.now().strftime("%Y-%m-%d %H:%M")),
        ("Endpoints",  str(len(endpoints))),
    ]
    for i, (k, v) in enumerate(meta, start=2):
        ws.cell(row=i, column=1, value=k).font  = Font(bold=True, color=NAVY)
        ws.cell(row=i, column=2, value=v)
        ws.row_dimensions[i].height = 18

    # Endpoint table header
    header_row = len(meta) + 3
    for ci, label in enumerate(["Method", "Path", "Summary", "Tag", "Sheet Name"], start=1):
        cell = ws.cell(row=header_row, column=ci, value=label)
        cell.font  = Font(bold=True, color=WHITE, name="Calibri", size=11)
        cell.fill  = PatternFill("solid", fgColor=NAVY)
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[header_row].height = 22

    for i, ep in enumerate(endpoints, start=header_row + 1):
        method = ep["method"]
        txt_c, bg_c = METHOD_COLORS.get(method, ("FF374151", "FFF9FAFB"))
        mc = ws.cell(row=i, column=1, value=method)
        mc.font      = Font(bold=True, color=txt_c)
        mc.fill      = PatternFill("solid", fgColor=bg_c)
        mc.alignment = Alignment(horizontal="center")
        ws.cell(row=i, column=2, value=ep["path"])
        ws.cell(row=i, column=3, value=ep["summary"])
        ws.cell(row=i, column=4, value=ep["tag"])
        ws.cell(row=i, column=5, value=_safe_sheet_name(f"{ep['method']} {ep['path']}"))
        ws.row_dimensions[i].height = 18

    # Column widths for summary
    for ci, w in enumerate([12, 35, 55, 18, 30], start=1):
        ws.column_dimensions[get_column_letter(ci)].width = w

    ws.freeze_panes = "A2"


def _write_endpoint_sheet(ws, ep: dict, op: dict, rows: list[dict]):
    method  = ep["method"]
    path    = ep["path"]
    summary = op.get("summary", op.get("description", ""))
    txt_c, bg_c = METHOD_COLORS.get(method, ("FF374151", "FFF9FAFB"))

    # ── Title banner ──────────────────────────────────────────────────────────
    ws.merge_cells(f"A1:{get_column_letter(len(SCHEMA_COLS))}1")
    c = ws["A1"]
    c.value = f"{method}  {path}"
    c.font  = Font(name="Calibri", size=15, bold=True, color=WHITE)
    c.fill  = PatternFill("solid", fgColor=TEAL)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[1].height = 36

    if summary:
        ws.merge_cells(f"A2:{get_column_letter(len(SCHEMA_COLS))}2")
        sc = ws["A2"]
        sc.value = summary
        sc.font  = Font(italic=True, color="FF374151", size=10)
        sc.fill  = PatternFill("solid", fgColor=LIGHT_TEAL)
        sc.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws.row_dimensions[2].height = 18
        header_row = 3
    else:
        header_row = 2

    # ── Column headers ─────────────────────────────────────────────────────────
    _thin = Side(style="thin", color="FFCBD5E1")
    _border = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)

    for ci, col_name in enumerate(COL_NAMES, start=1):
        cell = ws.cell(row=header_row, column=ci, value=col_name)
        cell.font      = Font(bold=True, color=WHITE, name="Calibri", size=10)
        cell.fill      = PatternFill("solid", fgColor=NAVY)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = _border
    ws.row_dimensions[header_row].height = 24

    # ── Data rows ─────────────────────────────────────────────────────────────
    current_section = None
    data_row = header_row + 1

    for idx, r in enumerate(rows, start=1):
        # Section separator row
        section = r.get("_section", "")
        if section and section != current_section:
            current_section = section
            sec_cell = ws.cell(row=data_row, column=1, value=f"▸  {section}")
            ws.merge_cells(
                start_row=data_row, start_column=1,
                end_row=data_row, end_column=len(SCHEMA_COLS)
            )
            sec_cell.font  = Font(bold=True, color=TEAL, size=10)
            sec_cell.fill  = PatternFill("solid", fgColor=LIGHT_TEAL)
            sec_cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
            ws.row_dimensions[data_row].height = 20
            data_row += 1

        depth = r.get("_depth", 0)
        is_arr = r.get("_is_array_item", False)

        # Row background
        if r.get("deprecated") == "Yes":
            fill_col = LIGHT_RED
        elif r.get("required") == "Yes":
            fill_col = LIGHT_BLUE if (data_row % 2 == 0) else "FFE0EDFF"
        elif is_arr:
            fill_col = LIGHT_YELLOW
        elif data_row % 2 == 0:
            fill_col = ALT_ROW
        else:
            fill_col = WHITE

        row_fill = PatternFill("solid", fgColor=fill_col)

        # Indent field name to show hierarchy
        indent_str = "  " * depth
        field_display = indent_str + r["field_name"]

        values = [
            idx,
            method,   # Path column shows method for easy reading
            field_display,
            r["full_path"],
            r["type"],
            r["format"],
            r["required"],
            r["nullable"],
            r["description"],
            r["enum"],
            r["example"],
            r["default"],
            r["minimum"],
            r["maximum"],
            r["pattern"],
            r["readOnly"],
            r["writeOnly"],
            r["deprecated"],
        ]
        # Map to SCHEMA_COLS:
        # #, Field Name, Full Path, Type, Format, Required, Nullable,
        # Description, Enum, Example, Default, Min, Max, Pattern, RO, WO, Dep
        mapped = [
            idx,
            field_display,
            r["full_path"],
            r["type"],
            r["format"],
            r["required"],
            r["nullable"],
            r["description"],
            r["enum"],
            r["example"],
            r["default"],
            r["minimum"],
            r["maximum"],
            r["pattern"],
            r["readOnly"],
            r["writeOnly"],
            r["deprecated"],
        ]

        for ci, val in enumerate(mapped, start=1):
            cell = ws.cell(row=data_row, column=ci, value=val)
            cell.fill      = row_fill
            cell.border    = _border
            cell.alignment = Alignment(vertical="center", wrap_text=(ci == 8))  # wrap desc
            cell.font      = Font(size=9)

            # Special formatting
            if ci == 1:   # #
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.font = Font(size=9, color="FF64748B")
            elif ci == 2:  # Field Name — bold if required
                cell.font = Font(
                    size=9,
                    bold=(r["required"] == "Yes"),
                    color="FF1E3A5F" if depth == 0 else "FF374151"
                )
            elif ci == 6:  # Required
                if val == "Yes":
                    cell.font = Font(size=9, bold=True, color="FF16A34A")
                else:
                    cell.font = Font(size=9, color="FF9CA3AF")
            elif ci == 17:  # Deprecated
                if val == "Yes":
                    cell.font = Font(size=9, bold=True, color="FFDC2626")

        ws.row_dimensions[data_row].height = 17
        data_row += 1

    # ── Column widths ─────────────────────────────────────────────────────────
    for ci, width in enumerate(COL_WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(ci)].width = width

    # Freeze panes below header + summary
    ws.freeze_panes = f"C{header_row + 1}"


# ── Helpers ───────────────────────────────────────────────────────────────────
def _load_yaml(path: str) -> dict:
    """Load a YAML or JSON file (OpenAPI supports both)."""
    p = Path(path)
    if not p.exists():
        return {}
    text = p.read_text(encoding="utf-8", errors="replace")
    try:
        data = yaml.safe_load(text)
        return data if isinstance(data, dict) else {}
    except yaml.YAMLError as e:
        raise ValueError(f"YAML parse error in {p.name}: {e}") from e


def _safe_sheet_name(name: str) -> str:
    """Make a valid Excel sheet name (≤31 chars, no special chars)."""
    name = re.sub(r"[/\\?*\[\]:]", " ", name)
    name = re.sub(r"\s+", " ", name).strip()
    return name[:31]


# ── CLI ───────────────────────────────────────────────────────────────────────
if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="OpenAPI/AsyncAPI YAML → Excel schema extractor")
    parser.add_argument("yaml",   help="Input YAML file path")
    parser.add_argument("-o",     "--output", default=None, help="Output .xlsx path")
    parser.add_argument("--endpoints", nargs="*", default=None,
                        help="Filter endpoints e.g. GET:/pets POST:/pets/{id}")
    parser.add_argument("--list", action="store_true",
                        help="List detected endpoints and exit")
    args = parser.parse_args()

    if args.list:
        eps = detect_endpoints(args.yaml)
        print(f"Found {len(eps)} endpoints:")
        for ep in eps:
            print(f"  {ep['id']:40s} {ep['summary']}")
        sys.exit(0)

    result = extract_yaml_api(args.yaml, args.output, filter_endpoints=args.endpoints)
    if result["success"]:
        print(f"✅ {result['message']}")
        print(f"   Output: {result['output_path']}")
    else:
        print(f"❌ Error: {result.get('error', 'Unknown')}")
        sys.exit(1)
