#!/usr/bin/env python3
"""
XSD Explorer v2 – Professional XMLSpy-like XSD Analyser
═══════════════════════════════════════════════════════
Features:
  • Interactive collapsible tree with content model indicators (sequence/choice/all)
  • Schema diagram – visual node-link graph (D3.js)
  • Type inheritance chain panel
  • XPath expression per node (copy to clipboard)
  • Validation / constraints summary panel
  • Statistics panel – cardinality heatmap + type usage counts
  • Split view: tree left | raw XSD source snippet right
  • Excel report (unchanged quality)
"""

import sys, os, json, re
from lxml import etree
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

XSD_NS  = "http://www.w3.org/2001/XMLSchema"
XSD_PRE = "xs"

# ──────────────────────────────────────────────────────────────────────────────
# Helpers
# ──────────────────────────────────────────────────────────────────────────────

def qtag(name):
    return f"{{{XSD_NS}}}{name}"

def local(el):
    try:
        return etree.QName(el.tag).localname
    except Exception:
        return None

def strip_ns(v):
    if v and ":" in v:
        return v.split(":", 1)[1]
    return v or ""

def get_doc(el):
    d = el.find(f".//{qtag('documentation')}")
    return (d.text or "").strip() if d is not None and d.text else ""

# ──────────────────────────────────────────────────────────────────────────────
# Deep XSD Parser – captures content model, facets, inheritance, raw source
# ──────────────────────────────────────────────────────────────────────────────

def parse_xsd(xsd_path):
    raw_text = open(xsd_path, encoding="utf-8").read()
    tree = etree.parse(xsd_path)
    root = tree.getroot()

    # ── index all named top-level definitions ──────────────────────────────
    type_map   = {}   # name -> element
    source_map = {}   # name -> raw XSD snippet
    for child in root:
        name = child.get("name")
        if name:
            type_map[name] = child
            try:
                src = etree.tostring(child, pretty_print=True).decode()
                source_map[name] = src
            except Exception:
                source_map[name] = ""

    # ── collect type usage counts ──────────────────────────────────────────
    type_usage = {}
    for el in root.iter(qtag("element")):
        t = strip_ns(el.get("type", ""))
        if t:
            type_usage[t] = type_usage.get(t, 0) + 1

    # ── build inheritance chains ───────────────────────────────────────────
    def get_inheritance(type_name, visited=None):
        if visited is None: visited = set()
        if type_name in visited or type_name not in type_map:
            return []
        visited.add(type_name)
        el = type_map[type_name]
        chain = []
        for ext in el.iter(qtag("extension")):
            base = strip_ns(ext.get("base", ""))
            if base:
                chain.append({"type": type_name, "via": "extension", "base": base})
                chain += get_inheritance(base, visited)
                break
        for rst in el.iter(qtag("restriction")):
            base = strip_ns(rst.get("base", ""))
            if base:
                chain.append({"type": type_name, "via": "restriction", "base": base})
                chain += get_inheritance(base, visited)
                break
        return chain

    # ── collect facets ─────────────────────────────────────────────────────
    FACET_TAGS = {"minLength","maxLength","length","pattern","enumeration",
                  "minInclusive","maxInclusive","minExclusive","maxExclusive",
                  "fractionDigits","totalDigits","whiteSpace"}

    def get_facets(el):
        facets = {}
        enums  = []
        for child in el.iter():
            ln = local(child)
            if ln in FACET_TAGS:
                if ln == "enumeration":
                    enums.append(child.get("value",""))
                else:
                    facets[ln] = child.get("value","")
        if enums:
            facets["enumeration"] = enums
        return facets

    # ── content-model group kind ───────────────────────────────────────────
    def group_kind(el):
        for child in el:
            ln = local(child)
            if ln in ("sequence","choice","all","group"):
                return ln
        return "sequence"

    # ── node builder ───────────────────────────────────────────────────────
    node_id_counter = [0]

    def new_id():
        node_id_counter[0] += 1
        return f"n{node_id_counter[0]}"

    def build_element(el, path="/", visited=None):
        if visited is None: visited = set()

        name    = el.get("name") or strip_ns(el.get("ref","(ref)"))
        ref     = el.get("ref")

        # resolve ref
        if ref:
            ref_local = strip_ns(ref)
            if ref_local in type_map:
                el   = type_map[ref_local]
                name = el.get("name", ref_local)
            else:
                return mk_node(new_id(), name, "ref", path, {}, [], "", [], {})

        type_attr = strip_ns(el.get("type",""))
        props = {}
        for k,v in el.attrib.items():
            if k != "name":
                props[strip_ns(k)] = strip_ns(v)

        doc = get_doc(el)
        xpath = path + name

        node_key = (name, type_attr)
        circular = node_key in visited
        visited2 = visited | {node_key}

        children   = []
        content_model = "sequence"
        facets     = {}
        raw_src    = source_map.get(type_attr, "") or source_map.get(name, "")

        if not circular:
            # inline complexType
            ict = el.find(qtag("complexType"))
            if ict is not None:
                content_model = group_kind(ict)
                children = collect_children(ict, xpath+"/", visited2)
            # inline simpleType
            ist = el.find(qtag("simpleType"))
            if ist is not None:
                facets = get_facets(ist)
                for rst in ist.iter(qtag("restriction")):
                    props.setdefault("base", strip_ns(rst.get("base","")))
            # referenced type
            elif type_attr and type_attr in type_map:
                type_el = type_map[type_attr]
                lt = local(type_el)
                if lt == "complexType":
                    content_model = group_kind(type_el)
                    children = collect_children(type_el, xpath+"/", visited2)
                elif lt == "simpleType":
                    facets = get_facets(type_el)
                    for rst in type_el.iter(qtag("restriction")):
                        props.setdefault("base", strip_ns(rst.get("base","")))

        if circular:
            props["_circular"] = "true"

        inheritance = get_inheritance(type_attr) if type_attr else []

        return mk_node(new_id(), name, "element", xpath, props, children,
                       doc, inheritance, facets, content_model, raw_src,
                       type_usage.get(type_attr, 0))

    def collect_children(el, path, visited):
        kids = []
        for child in el:
            ln = local(child)
            if ln is None: continue
            if ln == "element":
                kids.append(build_element(child, path, visited))
            elif ln in ("sequence","choice","all","group",
                        "complexContent","simpleContent",
                        "extension","restriction","complexType","simpleType"):
                kids.extend(collect_children(child, path, visited))
            elif ln == "attribute":
                a_name  = child.get("name","(attr)")
                a_props = {strip_ns(k):strip_ns(v)
                           for k,v in child.attrib.items() if k!="name"}
                a_doc   = get_doc(child)
                a_facets = {}
                a_type  = strip_ns(child.get("type",""))
                if a_type in type_map:
                    a_facets = get_facets(type_map[a_type])
                kids.append(mk_node(new_id(), a_name, "attribute",
                                    path+f"@{a_name}", a_props, [], a_doc,
                                    [], a_facets, "n/a",
                                    source_map.get(a_type,""), 0))
        return kids

    def mk_node(nid, name, kind, xpath, props, children,
                doc, inheritance, facets, content_model="sequence",
                raw_src="", type_uses=0):
        return {
            "id": nid,
            "name": name,
            "kind": kind,
            "xpath": xpath,
            "props": props,
            "children": children,
            "doc": doc,
            "inheritance": inheritance,
            "facets": facets,
            "content_model": content_model,
            "raw_src": raw_src,
            "type_uses": type_uses,
        }

    # ── walk top-level elements ────────────────────────────────────────────
    roots = []
    for child in root:
        if local(child) == "element":
            roots.append(build_element(child, "/"))

    # ── global stats ───────────────────────────────────────────────────────
    def flatten(nodes):
        for n in nodes:
            yield n
            yield from flatten(n["children"])

    all_nodes = list(flatten(roots))
    stats = {
        "total_elements"  : sum(1 for n in all_nodes if n["kind"]=="element"),
        "total_attributes": sum(1 for n in all_nodes if n["kind"]=="attribute"),
        "required"        : sum(1 for n in all_nodes if n["props"].get("minOccurs","1") not in ("0","")),
        "optional"        : sum(1 for n in all_nodes if n["props"].get("minOccurs","1") == "0"),
        "unbounded"       : sum(1 for n in all_nodes if n["props"].get("maxOccurs","") == "unbounded"),
        "type_usage"      : sorted(type_usage.items(), key=lambda x:-x[1])[:20],
        "max_depth"       : max((n["xpath"].count("/")-1 for n in all_nodes), default=0),
        "total_types"     : len(type_map),
        "type_names"      : list(type_map.keys()),
    }

    # graph edges for diagram
    edges = []
    for n in all_nodes:
        t = strip_ns(n["props"].get("type",""))
        if t and t in type_map:
            edges.append({"from": n["xpath"], "to": t, "label": n["name"]})

    return roots, stats, edges, raw_text

# ──────────────────────────────────────────────────────────────────────────────
# HTML Generator
# ──────────────────────────────────────────────────────────────────────────────

CM_ICON = {"sequence":"⟶","choice":"⊕","all":"∀","group":"◈","n/a":""}
CM_COLOR= {"sequence":"#3b82f6","choice":"#f59e0b","all":"#10b981","group":"#8b5cf6","n/a":"#6b7280"}

def props_badges(props, facets):
    skip = {"_circular","name"}
    html = []
    for k,v in props.items():
        if k in skip: continue
        v2 = str(v)[:55]+"…" if len(str(v))>55 else str(v)
        html.append(f'<span class="badge b-{k}" title="{k}: {v}">{k}: <b>{v2}</b></span>')
    for k,v in facets.items():
        if k=="enumeration":
            v2 = ", ".join(v[:5])
            if len(v)>5: v2 += f"… (+{len(v)-5})"
        else:
            v2 = str(v)
        html.append(f'<span class="badge b-facet" title="{k}: {v2}">{k}: <b>{v2}</b></span>')
    return "".join(html)

def render_node(node, depth=0):
    nid      = node["id"]
    name     = node["name"]
    kind     = node["kind"]
    kids     = node["children"]
    props    = node["props"]
    facets   = node["facets"]
    cm       = node.get("content_model","sequence")
    circular = props.get("_circular")
    has_kids = bool(kids)

    icon  = "◆" if kind=="attribute" else ("▣" if has_kids else "▢")
    cm_badge = (f'<span class="cm-badge" style="color:{CM_COLOR.get(cm,"#888")}" '
                f'title="content model: {cm}">{CM_ICON.get(cm,"")} {cm}</span>') if has_kids else ""
    circ  = '<span class="circ">↩ circular</span>' if circular else ""
    badges = props_badges(props, facets)
    xpath_val = node["xpath"].replace('"','&quot;')

    data = json.dumps({
        "id"          : nid,
        "name"        : name,
        "kind"        : kind,
        "xpath"       : node["xpath"],
        "props"       : props,
        "facets"      : facets,
        "doc"         : node["doc"],
        "inheritance" : node["inheritance"],
        "content_model": cm,
        "raw_src"     : node["raw_src"],
        "type_uses"   : node["type_uses"],
    }, ensure_ascii=False).replace("'","&#39;")

    toggle = '<span class="tog">▶</span>' if has_kids else '<span class="tog leaf">·</span>'

    h = (f'<li class="node {"has-ch" if has_kids else "leaf"}" id="node-{nid}" data-depth="{depth}">'
         f'<div class="nrow" data-nid="{nid}" data-json=\'{data}\' '
         f'{"onclick=\"toggleNode(this)\"" if has_kids else "onclick=\"selectNode(this)\""}'
         f'data-xpath="{xpath_val}">'
         f'{toggle} <span class="icon">{icon}</span>'
         f'<span class="nname">{name}</span>'
         f'<span class="nkind k-{kind}">{kind}</span>'
         f'{cm_badge}{badges}{circ}'
         f'</div>')
    if has_kids:
        h += '<ul class="kids" style="display:none">'
        for ch in kids:
            h += render_node(ch, depth+1)
        h += '</ul>'
    h += '</li>'
    return h

def generate_html(roots, stats, edges, raw_xsd, xsd_name, out_path):
    tree_html = "".join(render_node(n) for n in roots)
    stats_j   = json.dumps(stats, ensure_ascii=False)

    # flatten all nodes for diagram
    def flatten(nodes, parent_id=None):
        out = []
        for n in nodes:
            out.append({"id": n["id"], "name": n["name"], "kind": n["kind"],
                        "parent": parent_id, "cm": n.get("content_model","sequence"),
                        "has_children": bool(n["children"])})
            out.extend(flatten(n["children"], n["id"]))
        return out
    graph_nodes = flatten(roots)

    HTML = f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<title>XSD Explorer v2 – {xsd_name}</title>
<script src="https://cdnjs.cloudflare.com/ajax/libs/d3/7.8.5/d3.min.js"></script>
<style>
@import url('https://fonts.googleapis.com/css2?family=JetBrains+Mono:wght@400;600&family=Sora:wght@300;400;600;700&display=swap');

:root {{
  --bg0:#0a0e1a; --bg1:#0f1526; --bg2:#151d35; --bg3:#1c2645;
  --border:#243060; --border2:#2e3d7a;
  --txt:#c8d4f0; --txt2:#8a9bc8; --txt3:#4a5a8a;
  --blue:#4f8ef7; --cyan:#38c4e8; --green:#34d997;
  --purple:#a78bfa; --orange:#fb923c; --yellow:#fbbf24; --red:#f87171;
  --seq:#4f8ef7; --choice:#fbbf24; --all:#34d997; --group:#a78bfa;
  --card:#111827; --card2:#0d1420;
}}
*{{box-sizing:border-box;margin:0;padding:0}}
html,body{{height:100%;overflow:hidden;font-family:'Sora',sans-serif;
          background:var(--bg0);color:var(--txt);font-size:13px}}

/* ── Layout ── */
#app{{display:flex;flex-direction:column;height:100vh}}
#header{{
  background:linear-gradient(135deg,var(--bg2),var(--bg1));
  border-bottom:1px solid var(--border);
  padding:10px 20px;display:flex;align-items:center;gap:16px;flex-shrink:0
}}
#header h1{{font-size:1.1rem;font-weight:700;color:var(--blue);letter-spacing:.03em}}
#header .sub{{color:var(--txt2);font-size:.8rem;font-family:'JetBrains Mono',monospace}}
.hbadge{{
  background:var(--bg3);border:1px solid var(--border);
  border-radius:6px;padding:3px 10px;font-size:.75rem;color:var(--txt2)
}}
.hbadge b{{color:var(--cyan)}}

/* ── Tab Bar ── */
#tabs{{
  display:flex;background:var(--bg1);border-bottom:1px solid var(--border);
  flex-shrink:0;gap:2px;padding:0 16px
}}
.tab{{
  padding:8px 18px;cursor:pointer;font-size:.8rem;font-weight:600;
  color:var(--txt3);border-bottom:2px solid transparent;transition:all .15s;
  letter-spacing:.04em;text-transform:uppercase
}}
.tab:hover{{color:var(--txt)}}
.tab.active{{color:var(--blue);border-bottom-color:var(--blue)}}

/* ── Toolbar ── */
#toolbar{{
  background:var(--bg2);border-bottom:1px solid var(--border);
  padding:8px 16px;display:flex;gap:8px;align-items:center;flex-shrink:0;flex-wrap:wrap
}}
.tbtn{{
  background:var(--bg3);color:var(--txt);border:1px solid var(--border);
  padding:5px 12px;border-radius:6px;cursor:pointer;font-size:.76rem;
  font-family:'Sora',sans-serif;transition:all .15s
}}
.tbtn:hover{{border-color:var(--blue);color:var(--blue);background:rgba(79,142,247,.08)}}
#search{{
  background:var(--bg0);color:var(--txt);border:1px solid var(--border);
  padding:5px 12px;border-radius:6px;font-size:.76rem;width:240px;outline:none;
  font-family:'Sora',sans-serif
}}
#search:focus{{border-color:var(--blue)}}
#searchCnt{{color:var(--txt3);font-size:.75rem}}
.sep{{width:1px;height:20px;background:var(--border);margin:0 4px}}

/* ── Main panels ── */
#views{{flex:1;overflow:hidden;display:flex}}

/* Tree + Detail split */
#tree-view{{display:flex;width:100%;height:100%}}
#tree-panel{{flex:1;overflow:auto;padding:12px 6px 12px 14px;min-width:0}}
#split-panel{{
  width:400px;border-left:1px solid var(--border);display:flex;
  flex-direction:column;background:var(--bg1);flex-shrink:0
}}
#split-tabs{{display:flex;border-bottom:1px solid var(--border)}}
.stab{{
  padding:7px 14px;cursor:pointer;font-size:.74rem;font-weight:600;
  color:var(--txt3);border-bottom:2px solid transparent;
  text-transform:uppercase;letter-spacing:.04em;transition:all .15s
}}
.stab:hover{{color:var(--txt)}}
.stab.active{{color:var(--cyan);border-bottom-color:var(--cyan)}}
#split-content{{flex:1;overflow:auto;padding:14px}}

/* ── Tree ── */
ul{{list-style:none;padding-left:18px}}
ul.root-ul{{padding-left:0}}
.node{{margin:1px 0}}
.nrow{{
  display:flex;align-items:flex-start;gap:5px;padding:3px 8px;
  border-radius:5px;cursor:pointer;flex-wrap:wrap;transition:background .1s;
  position:relative
}}
.nrow:hover{{background:var(--bg2)}}
.nrow.selected{{background:rgba(79,142,247,.12);border-left:2px solid var(--blue);padding-left:6px}}
.nrow.hl{{background:rgba(251,191,36,.08);border-left:2px solid var(--yellow)}}

.tog{{width:12px;flex-shrink:0;color:var(--txt3);font-size:.7rem;margin-top:3px;user-select:none}}
.tog.leaf{{color:var(--bg3)}}
.icon{{font-size:.8rem;color:var(--txt3);flex-shrink:0;margin-top:2px}}
.nname{{color:#79c0ff;font-weight:600;font-size:.83rem;font-family:'JetBrains Mono',monospace}}
.nkind{{
  font-size:.67rem;padding:1px 6px;border-radius:4px;font-weight:700;
  align-self:center;letter-spacing:.04em;flex-shrink:0
}}
.k-element{{background:rgba(79,142,247,.15);color:var(--blue)}}
.k-attribute{{background:rgba(167,139,250,.15);color:var(--purple)}}
.k-ref{{background:rgba(52,217,151,.15);color:var(--green)}}

.cm-badge{{font-size:.68rem;font-weight:700;letter-spacing:.05em;flex-shrink:0;align-self:center;padding:0 4px}}

.badge{{
  font-size:.67rem;padding:1px 7px;border-radius:10px;white-space:nowrap;
  flex-shrink:0;align-self:center
}}
.b-type{{background:rgba(56,196,232,.1);color:var(--cyan)}}
.b-minOccurs,.b-maxOccurs{{background:rgba(52,217,151,.1);color:var(--green)}}
.b-use{{background:rgba(248,113,113,.1);color:var(--red)}}
.b-default,.b-fixed{{background:rgba(251,191,36,.1);color:var(--yellow)}}
.b-nillable{{background:rgba(251,146,60,.1);color:var(--orange)}}
.b-ref{{background:rgba(52,217,151,.1);color:var(--green)}}
.b-facet{{background:rgba(167,139,250,.1);color:var(--purple)}}
.circ{{color:var(--orange);font-size:.67rem}}
.kids{{border-left:1px dashed var(--border2);margin-left:8px}}

/* ── Detail / Split panels ── */
.dp-section{{margin-bottom:16px}}
.dp-label{{
  font-size:.67rem;text-transform:uppercase;letter-spacing:.08em;
  color:var(--txt3);margin-bottom:6px;font-weight:700
}}
.dp-name{{font-size:1.05rem;font-weight:700;color:var(--blue);
          font-family:'JetBrains Mono',monospace;margin-bottom:4px}}
.dp-kind-chip{{
  display:inline-block;padding:2px 10px;border-radius:10px;font-size:.72rem;
  font-weight:700;margin-bottom:10px
}}
.dp-xpath{{
  font-family:'JetBrains Mono',monospace;font-size:.72rem;color:var(--cyan);
  background:var(--bg0);padding:6px 10px;border-radius:5px;
  border:1px solid var(--border);word-break:break-all;cursor:pointer;
  position:relative
}}
.dp-xpath:hover::after{{
  content:'📋 copied!';position:absolute;right:8px;top:6px;
  font-size:.65rem;color:var(--green)
}}
.prop-grid{{display:grid;grid-template-columns:1fr 1fr;gap:6px}}
.prop-card{{
  background:var(--bg0);border:1px solid var(--border);border-radius:6px;
  padding:7px 10px
}}
.pk{{font-size:.65rem;color:var(--txt3);text-transform:uppercase;letter-spacing:.06em}}
.pv{{font-size:.8rem;color:var(--txt);font-family:'JetBrains Mono',monospace;word-break:break-all}}
.inherit-chain{{display:flex;flex-direction:column;gap:4px}}
.ic-item{{
  display:flex;align-items:center;gap:6px;
  background:var(--bg0);border:1px solid var(--border);
  border-radius:6px;padding:5px 10px;font-size:.76rem
}}
.ic-via{{
  font-size:.65rem;padding:1px 7px;border-radius:8px;font-weight:700
}}
.ic-ext{{background:rgba(52,217,151,.12);color:var(--green)}}
.ic-rst{{background:rgba(248,113,113,.12);color:var(--red)}}
.ic-type{{color:var(--cyan);font-family:'JetBrains Mono',monospace}}
.facet-grid{{display:grid;grid-template-columns:1fr 1fr;gap:5px}}
.fcard{{background:var(--bg0);border:1px solid var(--border);border-radius:5px;padding:5px 8px}}
.fk{{font-size:.63rem;color:var(--purple);text-transform:uppercase;letter-spacing:.06em}}
.fv{{font-size:.75rem;color:var(--txt);word-break:break-all}}

/* Constraint summary */
.constraint-ok{{color:var(--green)}}
.constraint-warn{{color:var(--yellow)}}
.constraint-info{{color:var(--cyan)}}

/* Source panel */
pre.src-code{{
  font-family:'JetBrains Mono',monospace;font-size:.72rem;
  color:#abb2bf;background:var(--bg0);padding:12px;border-radius:6px;
  border:1px solid var(--border);overflow:auto;white-space:pre;
  max-height:100%;line-height:1.6
}}

/* ── Stats panel ── */
#stats-view{{width:100%;height:100%;overflow:auto;padding:20px;display:none}}
.stat-grid{{display:grid;grid-template-columns:repeat(auto-fill,minmax(160px,1fr));gap:12px;margin-bottom:24px}}
.stat-card{{
  background:var(--bg2);border:1px solid var(--border);border-radius:10px;
  padding:16px;text-align:center
}}
.stat-big{{font-size:2rem;font-weight:700;color:var(--blue)}}
.stat-lbl{{font-size:.75rem;color:var(--txt2);margin-top:4px;text-transform:uppercase;letter-spacing:.06em}}
.heatmap-row{{
  display:flex;align-items:center;gap:10px;margin-bottom:6px;font-size:.8rem
}}
.hm-bar{{height:14px;border-radius:4px;background:var(--blue);min-width:4px;transition:width .3s}}
.hm-name{{font-family:'JetBrains Mono',monospace;color:var(--cyan);width:220px;overflow:hidden;text-overflow:ellipsis;white-space:nowrap;flex-shrink:0}}
.hm-count{{color:var(--txt3);width:30px;text-align:right;flex-shrink:0}}
.section-title{{font-size:.9rem;font-weight:700;color:var(--txt);margin-bottom:12px;
                letter-spacing:.04em;border-bottom:1px solid var(--border);padding-bottom:8px}}

/* ── Diagram panel ── */
#diag-view{{width:100%;height:100%;display:none;position:relative}}
#diag-svg{{width:100%;height:100%}}
.diag-toolbar{{position:absolute;top:10px;right:10px;display:flex;gap:6px}}
.dtbtn{{
  background:var(--bg2);border:1px solid var(--border);color:var(--txt2);
  padding:5px 12px;border-radius:6px;cursor:pointer;font-size:.75rem
}}

/* ── Scrollbar ── */
::-webkit-scrollbar{{width:7px;height:7px}}
::-webkit-scrollbar-track{{background:var(--bg0)}}
::-webkit-scrollbar-thumb{{background:var(--border2);border-radius:4px}}
</style>
</head>
<body>
<div id="app">

<!-- Header -->
<div id="header">
  <div>
    <h1>⬡ XSD Explorer <span style="color:var(--txt2);font-size:.8rem">v2</span></h1>
    <div class="sub">{xsd_name}</div>
  </div>
  <span class="hbadge">Elements: <b id="hdr-elem">–</b></span>
  <span class="hbadge">Types: <b id="hdr-types">–</b></span>
  <span class="hbadge">Max depth: <b id="hdr-depth">–</b></span>
</div>

<!-- Tab bar -->
<div id="tabs">
  <div class="tab active" onclick="switchTab('tree')">🌲 Tree</div>
  <div class="tab" onclick="switchTab('diag')">⬡ Diagram</div>
  <div class="tab" onclick="switchTab('stats')">📊 Statistics</div>
</div>

<!-- Toolbar (tree only) -->
<div id="toolbar">
  <button class="tbtn" onclick="expandAll()">⊞ Expand All</button>
  <button class="tbtn" onclick="collapseAll()">⊟ Collapse All</button>
  <button class="tbtn" onclick="expandLevel(1)">L1</button>
  <button class="tbtn" onclick="expandLevel(2)">L2</button>
  <button class="tbtn" onclick="expandLevel(3)">L3</button>
  <div class="sep"></div>
  <button class="tbtn" onclick="filterKind('all')">All</button>
  <button class="tbtn" onclick="filterKind('element')">Elements only</button>
  <button class="tbtn" onclick="filterKind('attribute')">Attributes only</button>
  <button class="tbtn" onclick="filterCardinality('optional')">Optional only</button>
  <button class="tbtn" onclick="filterCardinality('unbounded')">Unbounded only</button>
  <div class="sep"></div>
  <input type="text" id="search" placeholder="🔍  Search name / type / path…" oninput="doSearch(this.value)">
  <span id="searchCnt"></span>
</div>

<!-- Views -->
<div id="views">

  <!-- TREE VIEW -->
  <div id="tree-view">
    <div id="tree-panel">
      <ul class="root-ul">{tree_html}</ul>
    </div>
    <!-- Split right panel -->
    <div id="split-panel">
      <div id="split-tabs">
        <div class="stab active" onclick="switchSTab('detail')">Detail</div>
        <div class="stab" onclick="switchSTab('source')">Source</div>
        <div class="stab" onclick="switchSTab('inherit')">Inheritance</div>
        <div class="stab" onclick="switchSTab('constraints')">Constraints</div>
      </div>
      <div id="split-content">
        <div id="stab-detail"><p style="color:var(--txt3);font-size:.82rem">Click any node to inspect it.</p></div>
        <div id="stab-source" style="display:none"><p style="color:var(--txt3);font-size:.82rem">Select a node to see its XSD source.</p></div>
        <div id="stab-inherit" style="display:none"><p style="color:var(--txt3);font-size:.82rem">Select a typed element to see its inheritance chain.</p></div>
        <div id="stab-constraints" style="display:none"><p style="color:var(--txt3);font-size:.82rem">Select a node to see constraints summary.</p></div>
      </div>
    </div>
  </div>

  <!-- DIAGRAM VIEW -->
  <div id="diag-view">
    <svg id="diag-svg"></svg>
    <div class="diag-toolbar">
      <button class="dtbtn" onclick="diagZoomIn()">＋</button>
      <button class="dtbtn" onclick="diagZoomOut()">－</button>
      <button class="dtbtn" onclick="diagReset()">⊙ Reset</button>
    </div>
  </div>

  <!-- STATS VIEW -->
  <div id="stats-view">
    <div class="stat-grid" id="stat-cards"></div>
    <div class="section-title">Type Usage Heatmap (Top 20)</div>
    <div id="heatmap"></div>
    <br>
    <div class="section-title">Cardinality Distribution</div>
    <div id="cardinality-bars"></div>
  </div>

</div><!-- /views -->
</div><!-- /app -->

<script>
// ── Data ──────────────────────────────────────────────────────────────────
const STATS = {stats_j};
const GRAPH_NODES = {json.dumps(graph_nodes, ensure_ascii=False)};

// ── Boot ──────────────────────────────────────────────────────────────────
document.getElementById('hdr-elem').textContent  = STATS.total_elements;
document.getElementById('hdr-types').textContent = STATS.total_types;
document.getElementById('hdr-depth').textContent = STATS.max_depth;

// ── Tab switching ─────────────────────────────────────────────────────────
let currentTab  = 'tree';
let diagInited  = false;
let statsInited = false;

function switchTab(name) {{
  currentTab = name;
  document.querySelectorAll('.tab').forEach((t,i)=>{{
    t.classList.toggle('active', ['tree','diag','stats'][i]===name);
  }});
  document.getElementById('toolbar').style.display = name==='tree' ? 'flex' : 'none';
  document.getElementById('tree-view').style.display  = name==='tree'  ? 'flex'  : 'none';
  document.getElementById('diag-view').style.display  = name==='diag'  ? 'block' : 'none';
  document.getElementById('stats-view').style.display = name==='stats' ? 'block' : 'none';
  if (name==='diag'  && !diagInited)  {{ initDiagram(); diagInited=true; }}
  if (name==='stats' && !statsInited) {{ initStats();   statsInited=true; }}
}}

// ── Split sub-tab ─────────────────────────────────────────────────────────
function switchSTab(name) {{
  document.querySelectorAll('.stab').forEach((t,i)=>{{
    t.classList.toggle('active',['detail','source','inherit','constraints'][i]===name);
  }});
  ['detail','source','inherit','constraints'].forEach(n=>{{
    document.getElementById('stab-'+n).style.display = n===name ? 'block' : 'none';
  }});
}}

// ── Node toggle ───────────────────────────────────────────────────────────
function toggleNode(row) {{
  const li  = row.parentElement;
  const ul  = li.querySelector(':scope > ul');
  const tog = row.querySelector('.tog');
  if (!ul) return;
  const open = ul.style.display !== 'none' && ul.style.display !== '';
  ul.style.display = open ? 'none' : 'block';
  tog.textContent  = open ? '▶' : '▼';
  selectNode(row);
}}

function selectNode(row) {{
  document.querySelectorAll('.nrow.selected').forEach(r=>r.classList.remove('selected'));
  row.classList.add('selected');
  try {{
    const nd = JSON.parse(row.dataset.json);
    renderDetail(nd);
  }} catch(e) {{}}
}}

// ── Detail panel ──────────────────────────────────────────────────────────
function renderDetail(nd) {{
  const kind    = nd.kind;
  const kindCol = kind==='attribute' ? 'var(--purple)' : kind==='ref' ? 'var(--green)' : 'var(--blue)';
  const props   = nd.props || {{}};
  const facets  = nd.facets || {{}};

  // --- Detail tab ---
  let d = `
    <div class="dp-section">
      <div class="dp-name">${{nd.name}}</div>
      <span class="dp-kind-chip" style="background:${{kindCol}}22;color:${{kindCol}}">${{kind}}</span>
    </div>
    <div class="dp-section">
      <div class="dp-label">XPath</div>
      <div class="dp-xpath" onclick="copyXPath(this,'${{nd.xpath.replace(/'/g,"&apos;")}}')"
           title="Click to copy">${{nd.xpath}}</div>
    </div>`;

  const propEntries = Object.entries(props).filter(([k])=>k!=='_circular');
  if (propEntries.length) {{
    d += `<div class="dp-section"><div class="dp-label">Properties</div><div class="prop-grid">`;
    propEntries.forEach(([k,v])=>{{
      d += `<div class="prop-card"><div class="pk">${{k}}</div><div class="pv">${{v}}</div></div>`;
    }});
    d += `</div></div>`;
  }}
  if (Object.keys(facets).length) {{
    d += `<div class="dp-section"><div class="dp-label">Facets / Constraints</div><div class="facet-grid">`;
    Object.entries(facets).forEach(([k,v])=>{{
      const vd = Array.isArray(v) ? v.join(', ') : v;
      d += `<div class="fcard"><div class="fk">${{k}}</div><div class="fv">${{vd}}</div></div>`;
    }});
    d += `</div></div>`;
  }}
  if (nd.doc) {{
    d += `<div class="dp-section"><div class="dp-label">Documentation</div>
          <div style="font-size:.8rem;color:var(--txt2);line-height:1.6">${{nd.doc}}</div></div>`;
  }}
  document.getElementById('stab-detail').innerHTML = d;

  // --- Source tab ---
  const src = nd.raw_src || '(no source snippet available)';
  document.getElementById('stab-source').innerHTML =
    `<pre class="src-code">${{escHtml(src)}}</pre>`;

  // --- Inheritance tab ---
  const chain = nd.inheritance || [];
  let ih = '';
  if (!chain.length) {{
    ih = `<p style="color:var(--txt3);font-size:.82rem">No inheritance chain for this node.</p>`;
  }} else {{
    ih = `<div class="dp-label">Inheritance Chain</div><div class="inherit-chain">`;
    chain.forEach(item=>{{
      const cls  = item.via==='extension' ? 'ic-ext' : 'ic-rst';
      ih += `<div class="ic-item">
        <span class="ic-type">${{item.type}}</span>
        <span class="ic-via ${{cls}}">${{item.via}}</span>
        <span style="color:var(--txt3)">→</span>
        <span class="ic-type">${{item.base}}</span>
      </div>`;
    }});
    ih += `</div>`;
  }}
  document.getElementById('stab-inherit').innerHTML = ih;

  // --- Constraints tab ---
  const minO = props.minOccurs ?? '1';
  const maxO = props.maxOccurs ?? '1';
  const use  = props.use || '';
  const type = props.type || '';
  let cs = `<div class="dp-label">Constraint Summary</div>`;

  const checks = [];
  // Cardinality
  if (minO==='0' && maxO==='1')      checks.push(['constraint-info','Optional (0..1)']);
  else if (minO==='1' && maxO==='1') checks.push(['constraint-ok','Required (1..1)']);
  else if (maxO==='unbounded')       checks.push(['constraint-warn',`Repeatable (${{minO}}..∞)`]);
  else if (minO===maxO)              checks.push(['constraint-ok',`Fixed count (${{minO}})`]);
  else                               checks.push(['constraint-info',`Cardinality ${{minO}}..${{maxO}}`]);

  // Attribute use
  if (use==='required')  checks.push(['constraint-ok','Attribute: required']);
  if (use==='optional')  checks.push(['constraint-info','Attribute: optional']);
  if (use==='prohibited')checks.push(['constraint-warn','Attribute: prohibited']);

  // Nillable
  if (props.nillable==='true') checks.push(['constraint-warn','Nillable: can have xsi:nil']);

  // Fixed / Default
  if (props.fixed  !== undefined) checks.push(['constraint-ok', `Fixed value: ${{props.fixed}}`]);
  if (props.default!== undefined) checks.push(['constraint-info',`Default: ${{props.default}}`]);

  // Facets
  if (facets.minLength || facets.maxLength) {{
    const mn = facets.minLength||'0', mx = facets.maxLength||'∞';
    checks.push(['constraint-info',`Length: ${{mn}}..${{mx}} chars`]);
  }}
  if (facets.pattern) checks.push(['constraint-info',`Pattern: ${{facets.pattern}}`]);
  if (facets.enumeration) {{
    const enums = Array.isArray(facets.enumeration) ? facets.enumeration : [facets.enumeration];
    checks.push(['constraint-ok',`Enumeration (${{enums.length}} values): ${{enums.slice(0,5).join(', ')}}${{enums.length>5?'…':''}}`]);
  }}
  if (facets.minInclusive || facets.maxInclusive)
    checks.push(['constraint-info',`Range: [${{facets.minInclusive??'–'}}, ${{facets.maxInclusive??'–'}}]`]);

  if (nd.type_uses > 0)
    checks.push(['constraint-info',`Type reused ${{nd.type_uses}} time(s)`]);

  if (props._circular)
    checks.push(['constraint-warn','⚠ Circular type reference detected']);

  if (!checks.length) checks.push(['constraint-info','No specific constraints']);

  cs += checks.map(([cls,msg])=>`<div style="padding:6px 10px;margin-bottom:5px;border-radius:6px;
    background:var(--bg0);border:1px solid var(--border);font-size:.8rem"
    class="${{cls}}">● ${{msg}}</div>`).join('');

  document.getElementById('stab-constraints').innerHTML = cs;
}}

function copyXPath(el, xpath) {{
  navigator.clipboard.writeText(xpath).catch(()=>{{}});
  el.style.background='rgba(52,217,151,.1)';
  setTimeout(()=>el.style.background='',1200);
}}

function escHtml(s) {{
  return s.replace(/&/g,'&amp;').replace(/</g,'&lt;').replace(/>/g,'&gt;');
}}

// ── Tree controls ─────────────────────────────────────────────────────────
function expandAll() {{
  document.querySelectorAll('.kids').forEach(u=>u.style.display='block');
  document.querySelectorAll('.tog:not(.leaf)').forEach(t=>t.textContent='▼');
}}
function collapseAll() {{
  document.querySelectorAll('.kids').forEach(u=>u.style.display='none');
  document.querySelectorAll('.tog:not(.leaf)').forEach(t=>t.textContent='▶');
}}
function expandLevel(max) {{
  collapseAll();
  document.querySelectorAll('.node[data-depth]').forEach(li=>{{
    const d = parseInt(li.dataset.depth);
    if (d < max) {{
      const ul = li.querySelector(':scope > ul');
      const t  = li.querySelector('.nrow .tog');
      if (ul) {{ ul.style.display='block'; if(t&&!t.classList.contains('leaf'))t.textContent='▼'; }}
    }}
  }});
}}

// ── Search ────────────────────────────────────────────────────────────────
function doSearch(q) {{
  document.querySelectorAll('.nrow').forEach(r=>r.classList.remove('hl'));
  if (!q) {{ document.getElementById('searchCnt').textContent=''; return; }}
  const lq = q.toLowerCase();
  let n=0;
  document.querySelectorAll('.nrow').forEach(row=>{{
    if (row.textContent.toLowerCase().includes(lq)) {{
      row.classList.add('hl');
      // reveal ancestors
      let el=row.parentElement;
      while(el){{
        if(el.classList&&el.classList.contains('kids')){{
          el.style.display='block';
          const t=el.previousElementSibling?.querySelector('.tog');
          if(t&&!t.classList.contains('leaf'))t.textContent='▼';
        }}
        el=el.parentElement;
      }}
      n++;
    }}
  }});
  document.getElementById('searchCnt').textContent=`${{n}} match${{n!==1?'es':''}}`;
}}

// ── Filter by kind ────────────────────────────────────────────────────────
function filterKind(kind) {{
  document.querySelectorAll('.node').forEach(li=>{{
    if (kind==='all') {{ li.style.display=''; return; }}
    // Use :scope so we only read the kind badge of THIS node's own row, not a descendant's
    const k = li.querySelector(':scope > div.nrow .nkind')?.textContent||'';
    li.style.display = k===kind ? '' : 'none';
  }});
  if (kind!=='all') _revealMatchedAncestors();
}}
function filterCardinality(type) {{
  document.querySelectorAll('.node').forEach(li=>{{
    // Scope to the direct row so nested badge text doesn't pollute the match
    const badges = li.querySelector(':scope > div.nrow')?.textContent||'';
    if (type==='optional')   li.style.display = badges.includes('minOccurs: 0')         ? '' : 'none';
    if (type==='unbounded')  li.style.display = badges.includes('maxOccurs: unbounded')  ? '' : 'none';
  }});
  _revealMatchedAncestors();
}}

// Walk up from every visible node and ensure all ancestor ul.kids and li.node
// elements are also visible so matched items are actually reachable in the tree.
function _revealMatchedAncestors() {{
  document.querySelectorAll('.node').forEach(li=>{{
    if (li.style.display==='none') return;
    let el = li.parentElement;
    while (el) {{
      if (el.tagName==='UL' && el.classList.contains('kids')) {{
        el.style.display = 'block';
        const tog = el.previousElementSibling?.querySelector('.tog');
        if (tog && !tog.classList.contains('leaf')) tog.textContent='▼';
      }}
      if (el.tagName==='LI' && el.classList.contains('node')) {{
        el.style.display = '';
      }}
      el = el.parentElement;
    }}
  }});
}}

// ── Statistics ────────────────────────────────────────────────────────────
function initStats() {{
  const s = STATS;
  const cards = [
    ['Total Elements',   s.total_elements,   'var(--blue)'],
    ['Total Attributes', s.total_attributes, 'var(--purple)'],
    ['Named Types',      s.total_types,      'var(--cyan)'],
    ['Required',         s.required,         'var(--green)'],
    ['Optional',         s.optional,         'var(--yellow)'],
    ['Unbounded',        s.unbounded,        'var(--orange)'],
    ['Max Depth',        s.max_depth,        'var(--red)'],
  ];
  const cg = document.getElementById('stat-cards');
  cards.forEach(([lbl,val,col])=>{{
    cg.innerHTML += `<div class="stat-card">
      <div class="stat-big" style="color:${{col}}">${{val}}</div>
      <div class="stat-lbl">${{lbl}}</div></div>`;
  }});

  // Heatmap
  const hm = document.getElementById('heatmap');
  const maxU = s.type_usage[0]?.[1]||1;
  s.type_usage.forEach(([name,cnt])=>{{
    const pct = Math.round(cnt/maxU*200);
    hm.innerHTML += `<div class="heatmap-row">
      <div class="hm-name" title="${{name}}">${{name}}</div>
      <div class="hm-bar" style="width:${{pct}}px;background:linear-gradient(90deg,var(--blue),var(--cyan))"></div>
      <div class="hm-count">${{cnt}}</div>
    </div>`;
  }});

  // Cardinality bars
  const cb = document.getElementById('cardinality-bars');
  const total = s.total_elements||1;
  [
    ['Required (minOccurs≥1)', s.required,   'var(--green)'],
    ['Optional (minOccurs=0)', s.optional,   'var(--yellow)'],
    ['Unbounded',              s.unbounded,  'var(--orange)'],
  ].forEach(([lbl,cnt,col])=>{{
    const pct = Math.round(cnt/total*100);
    cb.innerHTML += `<div class="heatmap-row">
      <div class="hm-name">${{lbl}}</div>
      <div class="hm-bar" style="width:${{Math.round(cnt/total*300)}}px;background:${{col}}"></div>
      <div class="hm-count">${{cnt}} (${{pct}}%)</div>
    </div>`;
  }});
}}

// ── Diagram (D3 collapsible tree) ─────────────────────────────────────────
let diagramZoom;
function initDiagram() {{
  const svg    = d3.select('#diag-svg');
  const w      = document.getElementById('diag-view').clientWidth;
  const h      = document.getElementById('diag-view').clientHeight;
  const margin = {{top:40,right:120,bottom:40,left:60}};

  // Convert flat GRAPH_NODES to hierarchy
  const nodeMap = {{}};
  GRAPH_NODES.forEach(n=>nodeMap[n.id]=n);

  const roots = GRAPH_NODES.filter(n=>!n.parent);
  if (!roots.length) {{ svg.append('text').attr('x',w/2).attr('y',h/2)
    .attr('text-anchor','middle').attr('fill','#888').text('No diagram data'); return; }}

  function buildHierarchy(node) {{
    const kids = GRAPH_NODES.filter(n=>n.parent===node.id);
    return {{ ...node, children: kids.length ? kids.map(buildHierarchy) : null }};
  }}
  const hierData = buildHierarchy(roots[0]);

  const g = svg.append('g');
  diagramZoom = d3.zoom().scaleExtent([.05,3]).on('zoom',e=>g.attr('transform',e.transform));
  svg.call(diagramZoom);

  const dx = 26, dy = 200;
  const treeLayout = d3.tree().nodeSize([dx,dy]);

  let root2 = d3.hierarchy(hierData);
  root2.x0 = h/2; root2.y0 = 0;

  // Collapse after depth 2
  root2.descendants().forEach((d,i)=>{{
    if (d.depth>1 && d.children) {{ d._children=d.children; d.children=null; }}
  }});

  const cmColors = {{"sequence":"#4f8ef7","choice":"#fbbf24","all":"#34d997","group":"#a78bfa","n/a":"#6b7280"}};

  function update(source) {{
    treeLayout(root2);
    const nodes = root2.descendants();
    const links = root2.links();

    // Links
    const link = g.selectAll('.d-link').data(links, d=>d.target.data.id);
    link.enter().append('path').attr('class','d-link')
      .attr('fill','none').attr('stroke','#243060').attr('stroke-width',1.5)
      .attr('d', d3.linkHorizontal().x(d=>d.y).y(d=>d.x))
      .merge(link).transition().duration(300)
      .attr('d', d3.linkHorizontal().x(d=>d.y).y(d=>d.x));
    link.exit().remove();

    // Nodes
    const node = g.selectAll('.d-node').data(nodes, d=>d.data.id);
    const ne = node.enter().append('g').attr('class','d-node')
      .attr('transform', d=>`translate(${{source.y0||0}},${{source.x0||0}})`)
      .style('cursor','pointer')
      .on('click', (e,d)=>{{
        if (d.children) {{ d._children=d.children; d.children=null; }}
        else if (d._children) {{ d.children=d._children; d._children=null; }}
        update(d);
      }});

    ne.append('rect')
      .attr('x',-5).attr('y',-11).attr('width',10).attr('height',10)
      .attr('rx',2)
      .attr('fill', d=>cmColors[d.data.cm]||'#4f8ef7')
      .attr('stroke', d=>d._children?'#fff':'none').attr('stroke-width',1.5);

    ne.append('text').attr('dy','0').attr('x',d=>d.children||d._children?-14:14)
      .attr('text-anchor',d=>d.children||d._children?'end':'start')
      .attr('fill','#79c0ff').attr('font-size',11)
      .attr('font-family','JetBrains Mono,monospace')
      .text(d=>d.data.name);

    node.merge(ne).transition().duration(300)
      .attr('transform',d=>`translate(${{d.y}},${{d.x}})`);
    node.exit().remove();

    nodes.forEach(d=>{{d.x0=d.x;d.y0=d.y;}});
  }}

  update(root2);
  // Center
  svg.call(diagramZoom.transform,
    d3.zoomIdentity.translate(margin.left, h/2).scale(0.8));
}}

function diagZoomIn()  {{ if(diagramZoom) d3.select('#diag-svg').call(diagramZoom.scaleBy,1.3); }}
function diagZoomOut() {{ if(diagramZoom) d3.select('#diag-svg').call(diagramZoom.scaleBy,0.77); }}
function diagReset()   {{
  if(diagramZoom) d3.select('#diag-svg').call(
    diagramZoom.transform,
    d3.zoomIdentity.translate(60,document.getElementById('diag-view').clientHeight/2).scale(0.8));
}}
</script>
</body>
</html>"""

    with open(out_path,"w",encoding="utf-8") as f:
        f.write(HTML)
    print(f"✅ HTML v2 written: {out_path}")


# ──────────────────────────────────────────────────────────────────────────────
# Excel (enhanced with facets + xpath columns)
# ──────────────────────────────────────────────────────────────────────────────

def flatten_nodes(nodes, path=""):
    rows = []
    for n in nodes:
        name  = n["name"]
        fpath = f"{path}/{name}" if path else name
        row   = {"path": fpath, "name": name, "kind": n["kind"],
                 "depth": fpath.count("/"), "xpath": n["xpath"],
                 "doc": n["doc"], **n["props"]}
        for k,v in n["facets"].items():
            row[f"facet_{k}"] = ", ".join(v) if isinstance(v,list) else v
        if n["inheritance"]:
            row["inheritance"] = " → ".join(f"{i['type']}({i['via']})" for i in n["inheritance"])
        rows.append(row)
        rows.extend(flatten_nodes(n["children"], fpath))
    return rows

def generate_excel(nodes, xsd_name, out_path):
    rows = flatten_nodes(nodes)
    wb   = Workbook()
    ws   = wb.active
    ws.title = "XSD Structure"

    prop_keys = ["type","minOccurs","maxOccurs","use","default","fixed",
                 "base","nillable","facet_minLength","facet_maxLength",
                 "facet_pattern","facet_enumeration","inheritance","doc"]
    cols = ["Full XPath","Element / Attribute","Kind","Depth"] + prop_keys

    thin   = Side(style="thin",color="D0D7DE")
    border = Border(left=thin,right=thin,top=thin,bottom=thin)

    hf = Font(name="Segoe UI",bold=True,color="FFFFFF",size=10)
    hfill = PatternFill("solid",fgColor="1B2A5E")
    ha = Alignment(horizontal="center",vertical="center",wrap_text=True)

    ws.append(cols)
    for c,_ in enumerate(cols,1):
        cell = ws.cell(1,c)
        cell.font=hf; cell.fill=hfill; cell.alignment=ha; cell.border=border
    ws.row_dimensions[1].height = 28
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(cols))}1"

    widths = {"Full XPath":55,"Element / Attribute":30,"Kind":10,"Depth":6,
              "type":18,"minOccurs":9,"maxOccurs":9,"use":8,"default":10,"fixed":10,
              "base":14,"nillable":8,"facet_minLength":10,"facet_maxLength":10,
              "facet_pattern":25,"facet_enumeration":35,"inheritance":30,"doc":50}
    for c,k in enumerate(cols,1):
        ws.column_dimensions[get_column_letter(c)].width = widths.get(k,14)

    alt  = PatternFill("solid",fgColor="EEF3FF")
    attr = PatternFill("solid",fgColor="F5F0FF")
    path_font  = Font(name="Consolas",size=9,color="1F5C8B")
    elem_font  = Font(name="Consolas",size=9)
    attr_font  = Font(name="Consolas",size=9,color="6A3E9E")

    for ri,row in enumerate(rows,2):
        depth = row.get("depth",0)
        kind  = row.get("kind","element")
        vals  = [row.get("xpath",""), ("  "*depth)+row.get("name",""),
                 kind, depth] + [row.get(k,"") for k in prop_keys]
        ws.append(vals)
        for ci,_ in enumerate(vals,1):
            cell = ws.cell(ri,ci)
            cell.border = border
            cell.alignment = Alignment(vertical="center",wrap_text=(ci==len(cols)))
            cell.font = path_font if ci==1 else (attr_font if kind=="attribute" else elem_font)
            if kind=="attribute": cell.fill=attr
            elif ri%2==0:         cell.fill=alt
        ws.row_dimensions[ri].height = 16

    # Summary sheet
    ws2 = wb.create_sheet("Summary")
    ws2.column_dimensions["A"].width = 32
    ws2.column_dimensions["B"].width = 22
    ws2.append(["XSD Explorer v2 – Summary",""])
    ws2["A1"].font = Font(name="Segoe UI",bold=True,size=13,color="1F3864")
    ws2.merge_cells("A1:B1")
    for i,(k,v) in enumerate([
        ("XSD File",xsd_name),("Total Rows",len(rows)),
        ("Elements",sum(1 for r in rows if r.get("kind")=="element")),
        ("Attributes",sum(1 for r in rows if r.get("kind")=="attribute")),
        ("Max Depth",max((r.get("depth",0) for r in rows),default=0)),
    ],3):
        ws2.cell(i,1,k).font = Font(name="Segoe UI",bold=True,size=10)
        ws2.cell(i,2,v).font = Font(name="Segoe UI",size=10)

    wb.save(out_path)
    print(f"✅ Excel v2 written: {out_path}")


# ──────────────────────────────────────────────────────────────────────────────
# Main
# ──────────────────────────────────────────────────────────────────────────────
def main():
    if len(sys.argv) < 2:
        print("Usage: python xsd_explorer_v2.py <file.xsd> [output_dir]")
        sys.exit(1)
    xsd_path = sys.argv[1]
    out_dir  = sys.argv[2] if len(sys.argv)>2 else os.path.dirname(xsd_path) or "."
    os.makedirs(out_dir, exist_ok=True)
    xsd_name = os.path.basename(xsd_path)
    base     = os.path.splitext(xsd_name)[0]

    print(f"📂 Parsing: {xsd_path}")
    roots, stats, edges, raw = parse_xsd(xsd_path)
    print(f"   Root elements: {len(roots)}  |  Named types: {stats['total_types']}")

    generate_html(roots, stats, edges, raw, xsd_name,
                  os.path.join(out_dir, f"{base}_explorer_v2.html"))
    generate_excel(roots, xsd_name,
                   os.path.join(out_dir, f"{base}_structure_v2.xlsx"))
    print(f"\n🎉 Done!\n   HTML : {out_dir}/{base}_explorer_v2.html\n   Excel: {out_dir}/{base}_structure_v2.xlsx")

if __name__ == "__main__":
    main()
