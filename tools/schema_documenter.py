#!/usr/bin/env python3
"""
XSD to Excel Converter - FULLY ENHANCED VERSION
Includes: Choice indicators, Sample values, Validation rules, Business mapping, ISO metadata
"""

import xml.etree.ElementTree as ET
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import argparse
from pathlib import Path
import re
from datetime import datetime
import random


class SampleValueGenerator:
    def _classify_field(self, element, elem_name='', min_occurs='1', annotation=None):
        """Read Yellow/White from XSD annotations - ISO 20022 Spec"""
        # Try to get element if passed as string
        if isinstance(element, str):
            elem_name = element
            element = None
        
        # Check XSD annotation first
        if element is not None:
            ns = getattr(self, 'ns', {'xs': 'http://www.w3.org/2001/XMLSchema'})
            annotation_elem = element.find('xs:annotation', ns)
            if annotation_elem is not None:
                docs = annotation_elem.findall('xs:documentation', ns)
                for doc in docs:
                    source = doc.get('source', '').strip()
                    if source == 'Yellow Field':
                        return '🟡 Yellow (ISO 20022 Spec)'
                    elif source == 'White Field':
                        return '⚪ White (ISO 20022 Spec)'
        
        # NO INFERENCE - Only use XSD annotations
        # If not in XSD, return NA
        
        return '⚫ NA (Not in XSD)'

    """Generate sample values based on XSD types and restrictions"""
    
    @staticmethod
    def generate(element_name, elem_type, restrictions, path):
        """Generate a sample value for an element"""
        
        # Parse restrictions
        restrictions_dict = {}
        if restrictions:
            for restriction in restrictions.split('|'):
                restriction = restriction.strip()
                if ':' in restriction:
                    key, value = restriction.split(':', 1)
                    restrictions_dict[key.strip()] = value.strip()
        
        # Check for pattern
        if 'Pattern' in restrictions_dict:
            return SampleValueGenerator._generate_from_pattern(
                restrictions_dict['Pattern'], element_name
            )
        
        # Check for enumeration
        if 'Enum' in restrictions_dict:
            enum_values = restrictions_dict['Enum'].split(',')
            return enum_values[0].strip().split('(')[0].strip()
        
        # Generate based on type
        type_lower = elem_type.lower() if elem_type else ''
        
        # Common ISO 20022 types
        if 'iban' in type_lower or 'IBAN' in element_name:
            return "DE89370400440532013000"
        
        if 'bic' in type_lower or 'BIC' in element_name:
            return "DEUTDEFF"
        
        if 'currency' in type_lower or element_name in ['Ccy']:
            return "EUR"
        
        if 'datetime' in type_lower or 'DateTime' in elem_type:
            return datetime.now().strftime("%Y-%m-%dT%H:%M:%S.000Z")
        
        if 'date' in type_lower and 'time' not in type_lower:
            return datetime.now().strftime("%Y-%m-%d")
        
        if 'time' in type_lower and 'date' not in type_lower:
            return "10:30:00"
        
        if 'boolean' in type_lower or 'bool' in type_lower:
            return "true"
        
        if 'decimal' in type_lower or 'Amount' in element_name:
            min_val = restrictions_dict.get('Min', '0.01')
            return f"{min_val}"
        
        if 'integer' in type_lower or 'int' in type_lower or element_name in ['NbOfTxs']:
            return "1"
        
        # String types
        if 'string' in type_lower or 'Text' in elem_type:
            max_len = restrictions_dict.get('MaxLength', restrictions_dict.get('maxLength', '35'))
            try:
                max_len_int = int(max_len)
            except:
                max_len_int = 35
            
            # Generate meaningful sample based on element name
            if 'Id' in element_name or 'msgid' in element_name.lower():
                return f"{element_name.upper()}001"
            elif 'Nm' in element_name or 'Name' in element_name:
                return f"Sample {element_name}"
            elif 'Adr' in element_name or 'Address' in element_name:
                return "123 Main Street"
            elif 'Ctry' in element_name or 'Country' in element_name:
                return "DE"
            elif 'TwnNm' in element_name or 'Town' in element_name:
                return "Frankfurt"
            else:
                sample = f"Sample{element_name}"
                return sample[:max_len_int]
        
        return f"<{elem_type}>"
    
    @staticmethod
    def _generate_from_pattern(pattern, element_name):
        """Generate sample value from regex pattern"""
        
        # Common patterns
        if pattern == "[A-Z]{3,3}":
            return "EUR"
        
        if pattern == "[A-Z]{2}[0-9]{2}[A-Z0-9]+":
            return "DE89370400440532013000"
        
        if "[A-Z]" in pattern and "{2}" in pattern:
            return "DE"
        
        if pattern.startswith("[0-9]"):
            # Count digits
            if "{" in pattern:
                try:
                    count = int(re.search(r'\{(\d+)', pattern).group(1))
                    return "1" * count
                except:
                    return "123"
            return "123"
        
        # Default
        return f"<matches: {pattern[:30]}>"


class BusinessEntityMapper:
    """Map technical fields to business entities"""
    
    ENTITY_MAP = {
        'Dbtr': 'Debtor (Payer/Originator)',
        'Cdtr': 'Creditor (Payee/Beneficiary)',
        'DbtrAgt': 'Debtor Agent (Originating Bank)',
        'CdtrAgt': 'Creditor Agent (Beneficiary Bank)',
        'IntrmyAgt': 'Intermediary Agent',
        'GrpHdr': 'Group Header (Message-level info)',
        'MsgId': 'Message Identifier',
        'CreDtTm': 'Creation Date Time',
        'NbOfTxs': 'Number of Transactions',
        'IntrBkSttlmAmt': 'Interbank Settlement Amount',
        'IntrBkSttlmDt': 'Interbank Settlement Date',
        'SttlmInf': 'Settlement Information',
        'PmtTpInf': 'Payment Type Information',
        'CdtTrfTxInf': 'Credit Transfer Transaction',
        'PmtId': 'Payment Identification',
        'InstrId': 'Instruction Identification',
        'EndToEndId': 'End to End Identification',
        'IBAN': 'International Bank Account Number',
        'BIC': 'Bank Identifier Code (SWIFT)',
        'Nm': 'Name',
        'PstlAdr': 'Postal Address',
        'Ctry': 'Country Code',
        'RmtInf': 'Remittance Information',
        'Ustrd': 'Unstructured (Free text)',
        'Strd': 'Structured (Formatted data)',
    }
    
    @staticmethod
    def get_entity(element_name, path):
        """Get business entity for an element"""
        
        # Direct match
        if element_name in BusinessEntityMapper.ENTITY_MAP:
            return BusinessEntityMapper.ENTITY_MAP[element_name]
        
        # Check path components
        for key, value in BusinessEntityMapper.ENTITY_MAP.items():
            if key in path:
                return value
        
        # Pattern matching
        if 'Agt' in element_name:
            return 'Agent (Financial Institution)'
        
        if 'Acct' in element_name:
            return 'Account Information'
        
        if 'Amt' in element_name:
            return 'Amount'
        
        if 'Dt' in element_name and 'Tm' not in element_name:
            return 'Date'
        
        if 'Id' in element_name:
            return 'Identifier'
        
        return ''


class XSDParser:
    """Enhanced parser with all features"""
    
    NAMESPACES = {
        'xs': 'http://www.w3.org/2001/XMLSchema',
        'xsd': 'http://www.w3.org/2001/XMLSchema'
    }
    
    def __init__(self, xsd_file, schema_name=None):
        self.xsd_file = xsd_file
        self.schema_name = schema_name or Path(xsd_file).stem
        self.tree = ET.parse(xsd_file)
        self.root = self.tree.getroot()
        self.ns_prefix = self._detect_namespace()
        self.type_cache = {}
        self.choice_groups = {}  # Track choice groups
        self._build_type_cache()
        self._extract_metadata()
        
    def _detect_namespace(self):
        tag = self.root.tag
        if '{http://www.w3.org/2001/XMLSchema}' in tag:
            return '{http://www.w3.org/2001/XMLSchema}'
        return ''
    
    def _extract_metadata(self):
        """Extract ISO 20022 metadata from schema"""
        self.metadata = {
            'target_namespace': self.root.get('targetNamespace', ''),
            'element_form_default': self.root.get('elementFormDefault', ''),
        }
        
        # Extract message type from namespace
        ns = self.metadata['target_namespace']
        if 'iso:std:iso:20022' in ns:
            # Extract message type (e.g., pacs.008.001.08)
            parts = ns.split(':xsd:')
            if len(parts) > 1:
                self.metadata['message_type'] = parts[1]
                
                # Parse message components
                msg_parts = parts[1].split('.')
                if len(msg_parts) >= 4:
                    self.metadata['business_area'] = msg_parts[0]
                    self.metadata['message_functionality'] = msg_parts[1]
                    self.metadata['variant'] = msg_parts[2]
                    self.metadata['version'] = msg_parts[3]
        
        # Get root element name (usually message name)
        root_elem = self.root.find(f'{self.ns_prefix}element', self.NAMESPACES)
        if root_elem is not None:
            root_type = root_elem.get('type', '')
            self.metadata['root_element'] = root_elem.get('name', '')
            self.metadata['root_type'] = root_type
            
            # Parse scheme info from type name
            if 'NPC' in root_type or 'SEPA' in root_type:
                self.metadata['scheme'] = root_type.split('_', 1)[-1] if '_' in root_type else ''

    
    def _classify_field_from_xsd(self, element):
        """Read Yellow/White ONLY from XSD annotations - NO INFERENCE"""
        if element is None:
            return '⚫ NA (Not in XSD)'
            
        annotation = element.find(f'{self.ns_prefix}annotation', self.NAMESPACES)
        if annotation is not None:
            docs = annotation.findall(f'{self.ns_prefix}documentation', self.NAMESPACES)
            for doc in docs:
                source = doc.get('source', '').strip()
                if source == 'Yellow Field':
                    return '🟡 Yellow (ISO 20022 Spec)'
                elif source == 'White Field':
                    return '⚪ White (ISO 20022 Spec)'
        
        return '⚫ NA (Not in XSD)'
    
    def _get_tag_name(self, element):
        tag = element.tag
        if '}' in tag:
            return tag.split('}')[1]
        return tag
    
    def _get_attribute(self, element, attr_name):
        return element.get(attr_name, '')
    
    def _get_annotation(self, element):
        """Extract annotation/documentation"""
        annotation_parts = []
        annotation = element.find(f'{self.ns_prefix}annotation', self.NAMESPACES)
        if annotation is not None:
            for doc in annotation.findall(f'{self.ns_prefix}documentation', self.NAMESPACES):
                if doc.text and doc.text.strip():
                    text = doc.text.strip()
                    source = doc.get('source', '')
                    
                    if source and source not in ['Name', 'Definition']:
                        annotation_parts.append(f"[{source}] {text}")
                    else:
                        annotation_parts.append(text)
        
        return ' | '.join(annotation_parts) if annotation_parts else ''
    
    def _build_type_cache(self):
        """Build cache of all named types"""
        for complex_type in self.root.findall(f'{self.ns_prefix}complexType', self.NAMESPACES):
            type_name = self._get_attribute(complex_type, 'name')
            if type_name:
                self.type_cache[type_name] = complex_type
        
        for simple_type in self.root.findall(f'{self.ns_prefix}simpleType', self.NAMESPACES):
            type_name = self._get_attribute(simple_type, 'name')
            if type_name:
                self.type_cache[type_name] = simple_type
    
    def _get_type_definition(self, type_name):
        """Get the definition of a named type"""
        if ':' in type_name:
            type_name = type_name.split(':')[1]
        return self.type_cache.get(type_name)
    
    def _parse_simple_type_restrictions(self, simple_type):
        """Parse restrictions from simple type"""
        restriction_info = {'base_type': '', 'restrictions': '', 'validation_rules': []}
        restrictions = []
        
        restriction = simple_type.find(f'{self.ns_prefix}restriction', self.NAMESPACES)
        if restriction is not None:
            base = self._get_attribute(restriction, 'base')
            restriction_info['base_type'] = base or 'string'
            
            # Enumerations
            enums = []
            for enum in restriction.findall(f'{self.ns_prefix}enumeration', self.NAMESPACES):
                value = self._get_attribute(enum, 'value')
                if value:
                    enums.append(value)
            if enums:
                restrictions.append(f"Enum: {', '.join(enums[:10])}{'...' if len(enums) > 10 else ''}")
                restriction_info['validation_rules'].append(f"✓ Must be one of: {', '.join(enums[:5])}{'...' if len(enums) > 5 else ''}")
            
            # Pattern
            pattern = restriction.find(f'{self.ns_prefix}pattern', self.NAMESPACES)
            if pattern is not None:
                pattern_value = self._get_attribute(pattern, 'value')
                if pattern_value:
                    restrictions.append(f"Pattern: {pattern_value}")
                    restriction_info['validation_rules'].append(f"✓ Must match pattern: {pattern_value}")
            
            # Length constraints
            min_length = restriction.find(f'{self.ns_prefix}minLength', self.NAMESPACES)
            max_length = restriction.find(f'{self.ns_prefix}maxLength', self.NAMESPACES)
            length = restriction.find(f'{self.ns_prefix}length', self.NAMESPACES)
            
            if length is not None:
                len_val = self._get_attribute(length, 'value')
                restrictions.append(f"Length: {len_val}")
                restriction_info['validation_rules'].append(f"✓ Must be exactly {len_val} characters")
            else:
                if min_length is not None:
                    min_val = self._get_attribute(min_length, 'value')
                    restrictions.append(f"MinLength: {min_val}")
                if max_length is not None:
                    max_val = self._get_attribute(max_length, 'value')
                    restrictions.append(f"MaxLength: {max_val}")
                if min_length is not None or max_length is not None:
                    min_v = self._get_attribute(min_length, 'value') if min_length is not None else ''
                    max_v = self._get_attribute(max_length, 'value') if max_length is not None else ''
                    if min_v and max_v:
                        restriction_info['validation_rules'].append(f"✓ Length must be {min_v}-{max_v} characters")
                    elif max_v:
                        restriction_info['validation_rules'].append(f"✓ Maximum {max_v} characters")
                    elif min_v:
                        restriction_info['validation_rules'].append(f"✓ Minimum {min_v} characters")
            
            # Numeric constraints
            min_inc = restriction.find(f'{self.ns_prefix}minInclusive', self.NAMESPACES)
            max_inc = restriction.find(f'{self.ns_prefix}maxInclusive', self.NAMESPACES)
            
            if min_inc is not None or max_inc is not None:
                min_v = self._get_attribute(min_inc, 'value') if min_inc is not None else None
                max_v = self._get_attribute(max_inc, 'value') if max_inc is not None else None
                
                if min_v:
                    restrictions.append(f"Min: {min_v}")
                if max_v:
                    restrictions.append(f"Max: {max_v}")
                
                if min_v and max_v:
                    restriction_info['validation_rules'].append(f"✓ Value must be between {min_v} and {max_v}")
                elif min_v:
                    restriction_info['validation_rules'].append(f"✓ Minimum value: {min_v}")
                elif max_v:
                    restriction_info['validation_rules'].append(f"✓ Maximum value: {max_v}")
            
            # Fraction digits
            frac_digits = restriction.find(f'{self.ns_prefix}fractionDigits', self.NAMESPACES)
            if frac_digits is not None:
                frac_val = self._get_attribute(frac_digits, 'value')
                restrictions.append(f"FractionDigits: {frac_val}")
                restriction_info['validation_rules'].append(f"✓ Up to {frac_val} decimal places")
        
        restriction_info['restrictions'] = ' | '.join(restrictions) if restrictions else ''
        return restriction_info
    
    def parse(self):
        """Parse XSD and build actual XML message structure"""
        elements = []
        sequence = {'count': 0}
        
        # Find root elements
        for elem in self.root.findall(f'{self.ns_prefix}element', self.NAMESPACES):
            element_name = self._get_attribute(elem, 'name')
            element_type = self._get_attribute(elem, 'type')
            
            element_data = self._expand_element(
                elem, element_name, element_type,
                path=element_name, level=0, sequence=sequence,
                parent_choice_info=None
            )
            elements.append(element_data)
        
        return elements
    
    def _expand_element(self, element_node, element_name, element_type, path, level, sequence, parent_choice_info):
        """Expand an element following type references"""
        sequence['count'] += 1
        
        min_occurs = self._get_attribute(element_node, 'minOccurs') or '1'
        max_occurs = self._get_attribute(element_node, 'maxOccurs') or '1'
        default = self._get_attribute(element_node, 'default')
        fixed = self._get_attribute(element_node, 'fixed')
        annotation = self._get_annotation(element_node)
        
        element_data = {
            'sequence': sequence['count'],
            'level': level,
            'full_path': path,
            'name': element_name,
            'type': element_type,
            'min_occurs': min_occurs,
            'max_occurs': max_occurs,
            'default': default,
            'fixed': fixed,
            'annotation': annotation,
            'field_class': self._classify_field_from_xsd(element_node),
            'restrictions': '',
            'validation_rules': [],
            'node_type': 'element',
            'choice_info': parent_choice_info or '',
            'business_entity': BusinessEntityMapper.get_entity(element_name, path),
            'sample_value': '',
            'children': []
        }
        
        # Check for inline type
        inline_complex = element_node.find(f'{self.ns_prefix}complexType', self.NAMESPACES)
        inline_simple = element_node.find(f'{self.ns_prefix}simpleType', self.NAMESPACES)
        
        if inline_complex is not None:
            children = self._expand_complex_type(inline_complex, path, level + 1, sequence)
            element_data['children'] = children
        elif inline_simple is not None:
            restriction_info = self._parse_simple_type_restrictions(inline_simple)
            element_data['type'] = restriction_info['base_type']
            element_data['restrictions'] = restriction_info['restrictions']
            element_data['validation_rules'] = restriction_info['validation_rules']
        elif element_type:
            type_def = self._get_type_definition(element_type)
            if type_def is not None:
                tag_name = self._get_tag_name(type_def)
                if tag_name == 'complexType':
                    children = self._expand_complex_type(type_def, path, level + 1, sequence)
                    element_data['children'] = children
                elif tag_name == 'simpleType':
                    restriction_info = self._parse_simple_type_restrictions(type_def)
                    element_data['type'] = restriction_info['base_type']
                    element_data['restrictions'] = restriction_info['restrictions']
                    element_data['validation_rules'] = restriction_info['validation_rules']
        
        # Generate sample value
        element_data['sample_value'] = SampleValueGenerator.generate(
            element_name, element_data['type'], 
            element_data['restrictions'], path
        )
        
        return element_data
    
    def _expand_complex_type(self, complex_type, parent_path, level, sequence):
        """Expand complex type"""
        children = []
        
        # Handle complexContent
        complex_content = complex_type.find(f'{self.ns_prefix}complexContent', self.NAMESPACES)
        if complex_content is not None:
            restriction = complex_content.find(f'{self.ns_prefix}restriction', self.NAMESPACES)
            extension = complex_content.find(f'{self.ns_prefix}extension', self.NAMESPACES)
            
            target = restriction if restriction is not None else extension
            if target is not None:
                base = self._get_attribute(target, 'base')
                base_type_def = self._get_type_definition(base)
                if base_type_def is not None:
                    children.extend(self._expand_complex_type(base_type_def, parent_path, level, sequence))
                children.extend(self._parse_type_content(target, parent_path, level, sequence))
        else:
            children.extend(self._parse_type_content(complex_type, parent_path, level, sequence))
        
        return children
    
    def _parse_type_content(self, type_node, parent_path, level, sequence):
        """Parse type content with CHOICE detection"""
        children = []
        
        # Handle sequence
        for seq in type_node.findall(f'{self.ns_prefix}sequence', self.NAMESPACES):
            for child_elem in seq.findall(f'{self.ns_prefix}element', self.NAMESPACES):
                child_name = self._get_attribute(child_elem, 'name')
                child_type = self._get_attribute(child_elem, 'type')
                child_path = f"{parent_path}/{child_name}"
                
                child_data = self._expand_element(
                    child_elem, child_name, child_type,
                    child_path, level, sequence, None
                )
                children.append(child_data)
        
        # Handle CHOICE - mark alternatives
        for choice in type_node.findall(f'{self.ns_prefix}choice', self.NAMESPACES):
            choice_elements = choice.findall(f'{self.ns_prefix}element', self.NAMESPACES)
            total_choices = len(choice_elements)
            
            for idx, child_elem in enumerate(choice_elements, 1):
                child_name = self._get_attribute(child_elem, 'name')
                child_type = self._get_attribute(child_elem, 'type')
                child_path = f"{parent_path}/{child_name}"
                
                choice_info = f"[CHOICE {idx} of {total_choices}]"
                
                child_data = self._expand_element(
                    child_elem, child_name, child_type,
                    child_path, level, sequence, choice_info
                )
                children.append(child_data)
        
        # Handle all
        for all_elem in type_node.findall(f'{self.ns_prefix}all', self.NAMESPACES):
            for child_elem in all_elem.findall(f'{self.ns_prefix}element', self.NAMESPACES):
                child_name = self._get_attribute(child_elem, 'name')
                child_type = self._get_attribute(child_elem, 'type')
                child_path = f"{parent_path}/{child_name}"
                
                child_data = self._expand_element(
                    child_elem, child_name, child_type,
                    child_path, level, sequence, None
                )
                children.append(child_data)
        
        # Handle attributes
        for attr in type_node.findall(f'{self.ns_prefix}attribute', self.NAMESPACES):
            sequence['count'] += 1
            attr_name = self._get_attribute(attr, 'name')
            attr_type = self._get_attribute(attr, 'type')
            use = self._get_attribute(attr, 'use') or 'optional'
            default = self._get_attribute(attr, 'default')
            fixed = self._get_attribute(attr, 'fixed')
            annotation = self._get_annotation(attr)
            
            attr_path = f"{parent_path}/@{attr_name}"
            
            attr_data = {
                'sequence': sequence['count'],
                'level': level,
                'full_path': attr_path,
                'name': f"@{attr_name}",
                'type': attr_type or 'string',
                'min_occurs': '1' if use == 'required' else '0',
                'max_occurs': '1',
                'default': default,
                'fixed': fixed,
                'annotation': annotation,
                'restrictions': '',
                'validation_rules': [],
                'node_type': 'attribute',
                'choice_info': '',
                'business_entity': BusinessEntityMapper.get_entity(attr_name, attr_path),
                'sample_value': '',
                'children': []
            }
            
            # Check for inline simple type
            inline_simple = attr.find(f'{self.ns_prefix}simpleType', self.NAMESPACES)
            if inline_simple is not None:
                restriction_info = self._parse_simple_type_restrictions(inline_simple)
                attr_data['type'] = restriction_info['base_type']
                attr_data['restrictions'] = restriction_info['restrictions']
                attr_data['validation_rules'] = restriction_info['validation_rules']
            elif attr_type:
                type_def = self._get_type_definition(attr_type)
                if type_def is not None and self._get_tag_name(type_def) == 'simpleType':
                    restriction_info = self._parse_simple_type_restrictions(type_def)
                    attr_data['type'] = restriction_info['base_type']
                    attr_data['restrictions'] = restriction_info['restrictions']
                    attr_data['validation_rules'] = restriction_info['validation_rules']
            
            # Generate sample value
            attr_data['sample_value'] = SampleValueGenerator.generate(
                attr_name, attr_data['type'],
                attr_data['restrictions'], attr_path
            )
            
            children.append(attr_data)
        
        return children
    
    def flatten_tree(self, elements):
        """Flatten tree structure"""
        flat_list = []
        
        def flatten_recursive(element):
            flat_list.append(element)
            for child in element.get('children', []):
                flatten_recursive(child)
        
        for elem in elements:
            flatten_recursive(elem)
        
        return flat_list


class ExcelExporter:
    """Export to Excel with all enhancements"""
    
    def __init__(self, output_file, metadata):
        self.output_file = output_file
        self.metadata = metadata
        self.wb = Workbook()
    
    def _classify_field(self, element_name, min_occurs, annotation):
        """
        Classify field - NO INFERENCE
        Only mark as Yellow/White if we have explicit data
        Since this tool works with flat elements (no XSD access), mark all as NA
        """
        # This tool doesn't have access to XSD annotations
        # All classification happens in the parser
        return '⚫ NA (Not in XSD)'

    def export(self, flat_elements):
        """Export with multiple sheets"""
        # Remove default sheet
        if 'Sheet' in self.wb.sheetnames:
            self.wb.remove(self.wb['Sheet'])
        
        # Create sheets
        self._create_metadata_sheet()
        self._create_structure_sheet(flat_elements)
        self._create_quick_reference_sheet(flat_elements)
        
        self.wb.save(self.output_file)
        print(f"\n✅ Excel file saved: {self.output_file}")
        print(f"   📊 Total elements: {len(flat_elements)}")
        print(f"   📄 Sheets created: 3 (Metadata, Full Structure, Quick Reference)")
    
    def _create_metadata_sheet(self):
        """Create sheet with ISO 20022 metadata"""
        ws = self.wb.create_sheet("Message Metadata", 0)
        
        # Title
        ws['A1'] = "ISO 20022 Message Metadata"
        ws['A1'].font = Font(size=16, bold=True, color='FFFFFF')
        ws['A1'].fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        ws.merge_cells('A1:B1')
        
        row = 3
        metadata_items = [
            ('Message Type', self.metadata.get('message_type', 'N/A')),
            ('Root Element', self.metadata.get('root_element', 'N/A')),
            ('Business Area', self.metadata.get('business_area', 'N/A')),
            ('Message Functionality', self.metadata.get('message_functionality', 'N/A')),
            ('Version', self.metadata.get('version', 'N/A')),
            ('Scheme', self.metadata.get('scheme', 'N/A')),
            ('Target Namespace', self.metadata.get('target_namespace', 'N/A')),
            ('Element Form Default', self.metadata.get('element_form_default', 'N/A')),
        ]
        
        for label, value in metadata_items:
            ws[f'A{row}'] = label
            ws[f'B{row}'] = value
            ws[f'A{row}'].font = Font(bold=True)
            row += 1
        
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 60
    
    def _create_structure_sheet(self, flat_elements):
        """Create main structure sheet"""
        ws = self.wb.create_sheet("XML Structure")
        
        # Headers - ADD Field Classification
        headers = ['Seq', 'Lvl', 'Full XML Path', 'Element', 'Choice', 'Type', 
                   'Min', 'Max', 'Field Class', 'Sample Value', 'Business Entity', 'Validation Rules', 'Documentation']
        ws.append(headers)
        
        # Style headers
        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF')
        
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Sort by sequence
        flat_elements_sorted = sorted(flat_elements, key=lambda x: x.get('sequence', 0))
        
        # Add data
        for elem in flat_elements_sorted:
            level = elem['level']
            indent = '  ' * level
            
            # Classify field (yellow/white)
            field_class = elem.get('field_class', '⚫ NA (Not in XSD)')
            
            # Format validation rules
            validation_rules = elem.get('validation_rules', [])
            validation_text = '\n'.join(validation_rules) if validation_rules else ''
            
            row = [
                elem.get('sequence', ''),
                level,
                elem.get('full_path', ''),
                indent + elem['name'],
                elem.get('choice_info', ''),
                elem.get('type', ''),
                elem.get('min_occurs', ''),
                elem.get('max_occurs', ''),
                field_class,  # NEW: Field Classification
                elem.get('sample_value', ''),
                elem.get('business_entity', ''),
                validation_text,
                elem.get('annotation', '')
            ]
            ws.append(row)
            
            row_num = ws.max_row
            
            # Highlight Yellow fields
            if '🟡' in field_class:
                for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M']:
                    ws[f'{col}{row_num}'].fill = PatternFill(start_color='FFF8DC', end_color='FFF8DC', fill_type='solid')
            
            # Highlight choices
            if elem.get('choice_info'):
                ws[f'E{row_num}'].fill = PatternFill(start_color='FFE6CC', end_color='FFE6CC', fill_type='solid')
                ws[f'E{row_num}'].font = Font(bold=True, color='FF6600')
            
            # Color attributes
            if elem.get('node_type') == 'attribute':
                ws[f'D{row_num}'].font = Font(color='FF6600')
            
            # Wrap text
            ws[f'L{row_num}'].alignment = Alignment(wrap_text=True, vertical='top')
            ws[f'M{row_num}'].alignment = Alignment(wrap_text=True, vertical='top')
        
        # Column widths - UPDATED for new column
        ws.column_dimensions['A'].width = 6
        ws.column_dimensions['B'].width = 5
        ws.column_dimensions['C'].width = 55
        ws.column_dimensions['D'].width = 30
        ws.column_dimensions['E'].width = 18
        ws.column_dimensions['F'].width = 25
        ws.column_dimensions['G'].width = 5
        ws.column_dimensions['H'].width = 5
        ws.column_dimensions['I'].width = 25  # Field Classification
        ws.column_dimensions['J'].width = 30  # Sample Value
        ws.column_dimensions['K'].width = 35  # Business Entity
        ws.column_dimensions['L'].width = 40  # Validation Rules
        ws.column_dimensions['M'].width = 50  # Documentation
        
        # Freeze panes
        ws.freeze_panes = 'D2'
    
    def _create_quick_reference_sheet(self, flat_elements):
        """Create quick reference sheet with key fields only"""
        ws = self.wb.create_sheet("Quick Reference")
        
        # Headers - ADD Field Classification
        headers = ['Full Path', 'Element', 'Required', 'Field Class', 'Sample Value', 'Business Entity']
        ws.append(headers)
        
        # Style headers
        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF')
        
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Filter to show only elements (not attributes) at certain levels
        key_elements = [e for e in flat_elements 
                       if e.get('node_type') == 'element' and e.get('level', 0) <= 5]
        key_elements_sorted = sorted(key_elements, key=lambda x: x.get('sequence', 0))
        
        for elem in key_elements_sorted:
            is_required = 'Yes' if elem.get('min_occurs', '0') != '0' else 'No'
            field_class = self._classify_field(elem.get('name', ''), elem.get('min_occurs', '1'), elem.get('annotation', ''))
            
            row = [
                elem.get('full_path', ''),
                elem.get('name', ''),
                is_required,
                field_class,  # NEW
                elem.get('sample_value', ''),
                elem.get('business_entity', '')
            ]
            ws.append(row)
            
            # Highlight Yellow fields
            row_num = ws.max_row
            if '🟡' in field_class:
                for col in ['A', 'B', 'C', 'D', 'E', 'F']:
                    ws[f'{col}{row_num}'].fill = PatternFill(start_color='FFF8DC', end_color='FFF8DC', fill_type='solid')
        
        # Column widths - UPDATED
        ws.column_dimensions['A'].width = 60
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 10
        ws.column_dimensions['D'].width = 25  # Field Classification
        ws.column_dimensions['E'].width = 30
        ws.column_dimensions['F'].width = 35
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 10
        ws.column_dimensions['D'].width = 30
        ws.column_dimensions['E'].width = 35
        
        # Freeze panes
        ws.freeze_panes = 'A2'


def main():
    parser = argparse.ArgumentParser(
        description='XSD to Excel - ENHANCED with Choice Indicators, Sample Values, Business Mapping'
    )
    parser.add_argument('xsd_file', help='Path to XSD file')
    parser.add_argument('-o', '--output', help='Output Excel file', default='xml_structure_enhanced.xlsx')
    
    args = parser.parse_args()
    
    if not Path(args.xsd_file).exists():
        print(f"Error: File '{args.xsd_file}' not found")
        return
    
    print(f"\n{'='*70}")
    print("ISO 20022 SCHEMA DOCUMENTATION - FULLY ENHANCED VERSION")
    print(f"{'='*70}")
    print(f"\n📂 Parsing: {args.xsd_file}")
    print("🔧 Features: Choice indicators, Sample values, Validation rules, Business mapping")
    print("\n⏳ Processing...")
    
    # Parse
    xsd_parser = XSDParser(args.xsd_file)
    elements = xsd_parser.parse()
    flat_elements = xsd_parser.flatten_tree(elements)
    
    # Export
    exporter = ExcelExporter(args.output, xsd_parser.metadata)
    exporter.export(flat_elements)
    
    print(f"\n{'='*70}")
    print("✅ COMPLETE!")
    print(f"{'='*70}\n")


if __name__ == '__main__':
    main()

    def _classify_field(self, elem_name, min_occurs='1', annotation=''):
        """
        Classify field from XSD - NO INFERENCE, only real annotations
        Note: This tool doesn't have access to element objects, so can't read XSD directly
        Returns NA for all since we don't parse XSD annotations here
        """
        # This tool processes flat elements after parsing
        # Real classification happens in the parser
        # For now, return NA since we can't access XSD annotations from flat elements
        return '⚫ NA (Not in XSD)'

    def _classify_field_from_xsd(self, element):
        """Read Yellow/White from XSD annotations - NO INFERENCE"""
        # Check XSD annotation
        annotation = element.find('xs:annotation', self.ns)
        if annotation is not None:
            docs = annotation.findall('xs:documentation', self.ns)
            for doc in docs:
                source = doc.get('source', '').strip()
                if source == 'Yellow Field':
                    return '🟡 Yellow (ISO 20022 Spec)'
                elif source == 'White Field':
                    return '⚪ White (ISO 20022 Spec)'
        
        return '⚫ NA (Not in XSD)'
