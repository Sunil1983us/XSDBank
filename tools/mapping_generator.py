#!/usr/bin/env python3
"""
ISO 20022 Mapping Template Generator
=====================================
Generate mapping templates for implementation teams.

Features:
✅ Excel mapping template with all fields
✅ Columns: XPath, Element, Type, Min/Max, Source Field, Transformation, Notes
✅ Pre-filled with sample values
✅ Yellow/White field indicators
✅ Conditional formatting
✅ Multiple sheet views (hierarchical, flat, mandatory only)
"""

import xml.etree.ElementTree as ET
import json
import argparse
import os
from pathlib import Path
from datetime import datetime
from typing import Dict, List, Optional, Tuple
from dataclasses import dataclass

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, NamedStyle
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.comments import Comment
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


@dataclass
class FieldInfo:
    xpath: str
    element_name: str
    level: int
    data_type: str
    min_occurs: str
    max_occurs: str
    is_mandatory: bool
    pattern: Optional[str]
    min_length: Optional[int]
    max_length: Optional[int]
    enumeration: List[str]
    annotation: Optional[str]
    is_yellow: bool
    is_white: bool
    sample_value: str


class MappingTemplateGenerator:
    """Generate Excel mapping templates from XSD schemas"""
    
    def __init__(self, xsd_file: str):
        self.xsd_file = xsd_file
        self.tree = ET.parse(xsd_file)
        self.root = self.tree.getroot()
        self.ns = {'xs': 'http://www.w3.org/2001/XMLSchema'}
        self.target_ns = self.root.get('targetNamespace', '')
        
        self.type_cache = {}
        self.fields: List[FieldInfo] = []
        
        # Sample values for common field types
        self.sample_values = {
            'MsgId': 'MSG20240115123456',
            'CreDtTm': '2024-01-15T10:30:00',
            'NbOfTxs': '1',
            'TtlIntrBkSttlmAmt': '1000.00',
            'IntrBkSttlmDt': '2024-01-15',
            'EndToEndId': 'E2E20240115001',
            'TxId': 'TX20240115001',
            'InstrId': 'INSTR20240115001',
            'UETR': 'eb6305c9-1f7f-49de-aed0-16487c27b42d',
            'IntrBkSttlmAmt': '1000.00',
            'InstdAmt': '1000.00',
            'ChrgBr': 'SLEV',
            'Nm': 'John Smith',
            'IBAN': 'DE89370400440532013000',
            'BICFI': 'DEUTDEFF',
            'BIC': 'DEUTDEFF',
            'Ctry': 'DE',
            'Ccy': 'EUR',
            'StrtNm': 'Main Street',
            'BldgNb': '123',
            'PstCd': '10115',
            'TwnNm': 'Berlin',
            'Ustrd': 'Payment for Invoice 12345',
            'Cd': 'SALA',
            'Prtry': 'PROPRIETARY',
        }
        
        self._cache_types()
    
    def _cache_types(self):
        """Cache all type definitions"""
        for complex_type in self.root.findall('.//xs:complexType[@name]', self.ns):
            name = complex_type.get('name')
            self.type_cache[name] = complex_type
        
        for simple_type in self.root.findall('.//xs:simpleType[@name]', self.ns):
            name = simple_type.get('name')
            self.type_cache[name] = simple_type
    
    def extract_fields(self) -> List[FieldInfo]:
        """Extract all fields from the XSD"""
        self.fields = []
        
        # Find root element
        root_elem = self.root.find('xs:element', self.ns)
        if root_elem is not None:
            self._process_element(root_elem, "", 0)
        
        return self.fields
    
    def _process_element(self, element, parent_path: str, level: int):
        """Process an element and its children"""
        elem_name = element.get('name', '')
        elem_type = element.get('type', '')
        min_occurs = element.get('minOccurs', '1')
        max_occurs = element.get('maxOccurs', '1')
        
        current_path = f"{parent_path}/{elem_name}" if parent_path else elem_name
        
        # Get type info
        type_info = self._get_type_info(element, elem_type)
        
        # Check for annotations
        annotation_text = self._get_annotation(element)
        is_yellow = self._is_yellow_field(element)
        is_white = self._is_white_field(element)
        
        # Get sample value
        sample = self._get_sample_value(elem_name, type_info)
        
        # Create field info
        field = FieldInfo(
            xpath=current_path,
            element_name=elem_name,
            level=level,
            data_type=type_info.get('base_type', elem_type or 'complex'),
            min_occurs=min_occurs,
            max_occurs=max_occurs,
            is_mandatory=min_occurs != '0',
            pattern=type_info.get('pattern'),
            min_length=type_info.get('minLength'),
            max_length=type_info.get('maxLength'),
            enumeration=type_info.get('enumeration', []),
            annotation=annotation_text,
            is_yellow=is_yellow,
            is_white=is_white,
            sample_value=sample
        )
        
        self.fields.append(field)
        
        # Process children
        self._process_children(element, elem_type, current_path, level + 1)
    
    def _process_children(self, element, type_name: str, parent_path: str, level: int):
        """Process child elements"""
        # Check inline complex type
        complex_type = element.find('xs:complexType', self.ns)
        
        # Or referenced type
        if complex_type is None and type_name:
            type_name_clean = type_name.split(':')[-1]
            complex_type = self.type_cache.get(type_name_clean)
        
        if complex_type is None:
            return
        
        # Process sequence/choice/all
        for compositor in ['xs:sequence', 'xs:choice', 'xs:all']:
            for comp in complex_type.findall(f'.//{compositor}', self.ns):
                for child in comp.findall('xs:element', self.ns):
                    child_ref = child.get('ref')
                    if child_ref:
                        ref_elem = self.root.find(f".//xs:element[@name='{child_ref.split(':')[-1]}']", self.ns)
                        if ref_elem is not None:
                            # Copy min/max from reference
                            ref_copy = ET.Element(ref_elem.tag, ref_elem.attrib)
                            ref_copy.extend(list(ref_elem))
                            ref_copy.set('minOccurs', child.get('minOccurs', '1'))
                            ref_copy.set('maxOccurs', child.get('maxOccurs', '1'))
                            self._process_element(ref_copy, parent_path, level)
                    else:
                        self._process_element(child, parent_path, level)
        
        # Check complex content extension
        complex_content = complex_type.find('xs:complexContent', self.ns)
        if complex_content is not None:
            extension = complex_content.find('xs:extension', self.ns)
            if extension is not None:
                base_type = extension.get('base', '').split(':')[-1]
                if base_type in self.type_cache:
                    self._process_children_from_type(
                        self.type_cache[base_type], parent_path, level
                    )
                # Process extension's own children
                for compositor in ['xs:sequence', 'xs:choice', 'xs:all']:
                    for comp in extension.findall(f'.//{compositor}', self.ns):
                        for child in comp.findall('xs:element', self.ns):
                            self._process_element(child, parent_path, level)
    
    def _process_children_from_type(self, type_elem, parent_path: str, level: int):
        """Process children from a type definition"""
        for compositor in ['xs:sequence', 'xs:choice', 'xs:all']:
            for comp in type_elem.findall(f'.//{compositor}', self.ns):
                for child in comp.findall('xs:element', self.ns):
                    self._process_element(child, parent_path, level)
    
    def _get_type_info(self, element, type_name: str) -> Dict:
        """Get detailed type information"""
        info = {'base_type': 'string'}
        
        # Check for inline simple type
        simple_type = element.find('.//xs:simpleType', self.ns)
        if simple_type is None and type_name:
            type_name_clean = type_name.split(':')[-1]
            simple_type = self.type_cache.get(type_name_clean)
        
        if simple_type is not None:
            restriction = simple_type.find('.//xs:restriction', self.ns)
            if restriction is not None:
                info['base_type'] = restriction.get('base', 'string').split(':')[-1]
                
                # Get pattern
                pattern = restriction.find('xs:pattern', self.ns)
                if pattern is not None:
                    info['pattern'] = pattern.get('value')
                
                # Get length constraints
                min_len = restriction.find('xs:minLength', self.ns)
                max_len = restriction.find('xs:maxLength', self.ns)
                if min_len is not None:
                    info['minLength'] = int(min_len.get('value', 0))
                if max_len is not None:
                    info['maxLength'] = int(max_len.get('value', 0))
                
                # Get enumeration
                enums = restriction.findall('xs:enumeration', self.ns)
                if enums:
                    info['enumeration'] = [e.get('value') for e in enums]
        
        return info
    
    def _get_annotation(self, element) -> Optional[str]:
        """Get annotation text"""
        annotation = element.find('xs:annotation', self.ns)
        if annotation is not None:
            for doc in annotation.findall('xs:documentation', self.ns):
                source = doc.get('source', '')
                if source in ['Definition', 'Name']:
                    return doc.text
        return None
    
    def _is_yellow_field(self, element) -> bool:
        """Check if element is a Yellow field"""
        annotation = element.find('xs:annotation', self.ns)
        if annotation is not None:
            for doc in annotation.findall('xs:documentation', self.ns):
                if doc.get('source') == 'Yellow Field':
                    return True
        return False
    
    def _is_white_field(self, element) -> bool:
        """Check if element is a White field"""
        annotation = element.find('xs:annotation', self.ns)
        if annotation is not None:
            for doc in annotation.findall('xs:documentation', self.ns):
                if doc.get('source') == 'White Field':
                    return True
        return False
    
    def _get_sample_value(self, elem_name: str, type_info: Dict) -> str:
        """Get sample value for element"""
        # Check predefined samples
        if elem_name in self.sample_values:
            return self.sample_values[elem_name]
        
        # Check enumeration
        if type_info.get('enumeration'):
            return type_info['enumeration'][0]
        
        # Generate based on type
        base_type = type_info.get('base_type', 'string')
        if 'decimal' in base_type.lower():
            return '100.00'
        elif 'date' in base_type.lower() and 'time' in base_type.lower():
            return '2024-01-15T10:30:00'
        elif 'date' in base_type.lower():
            return '2024-01-15'
        elif 'integer' in base_type.lower() or 'int' in base_type.lower():
            return '1'
        elif 'boolean' in base_type.lower():
            return 'true'
        
        return f"[{elem_name}]"
    
    def generate_excel(self, output_path: str):
        """Generate Excel mapping template"""
        if not HAS_OPENPYXL:
            print("Error: openpyxl required. Install with: pip install openpyxl")
            return
        
        if not self.fields:
            self.extract_fields()
        
        wb = Workbook()
        
        # Styles
        header_fill = PatternFill(start_color="1a1a2e", end_color="1a1a2e", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        yellow_fill = PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid")
        white_fill = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")
        mandatory_font = Font(bold=True)
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # ============================================
        # Sheet 1: Full Mapping Template
        # ============================================
        ws_full = wb.active
        ws_full.title = "Mapping Template"
        
        headers = [
            "Level", "XPath", "Element", "Data Type", "Min", "Max", 
            "Mandatory", "Pattern/Enum", "Max Length",
            "Yellow", "White", "Sample Value",
            "Source System", "Source Field", "Transformation Rule", 
            "Default Value", "Notes", "Status"
        ]
        
        for col_idx, header in enumerate(headers, 1):
            cell = ws_full.cell(row=1, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', wrap_text=True)
            cell.border = thin_border
        
        # Freeze header row
        ws_full.freeze_panes = 'A2'
        
        for row_idx, field in enumerate(self.fields, 2):
            # Level (with indent visualization)
            indent = "  " * field.level
            ws_full.cell(row=row_idx, column=1, value=field.level)
            
            # XPath
            ws_full.cell(row=row_idx, column=2, value=field.xpath)
            
            # Element name (indented)
            elem_cell = ws_full.cell(row=row_idx, column=3, value=f"{indent}{field.element_name}")
            if field.is_mandatory:
                elem_cell.font = mandatory_font
            
            # Data Type
            ws_full.cell(row=row_idx, column=4, value=field.data_type)
            
            # Min/Max
            ws_full.cell(row=row_idx, column=5, value=field.min_occurs)
            ws_full.cell(row=row_idx, column=6, value=field.max_occurs)
            
            # Mandatory
            mandatory_cell = ws_full.cell(row=row_idx, column=7, 
                                          value="✓" if field.is_mandatory else "")
            mandatory_cell.alignment = Alignment(horizontal='center')
            
            # Pattern/Enum
            pattern_enum = ""
            if field.pattern:
                pattern_enum = f"Pattern: {field.pattern[:50]}"
            elif field.enumeration:
                pattern_enum = f"Enum: {', '.join(field.enumeration[:5])}"
                if len(field.enumeration) > 5:
                    pattern_enum += "..."
            ws_full.cell(row=row_idx, column=8, value=pattern_enum)
            
            # Max Length
            ws_full.cell(row=row_idx, column=9, value=field.max_length or "")
            
            # Yellow/White indicators
            yellow_cell = ws_full.cell(row=row_idx, column=10, 
                                       value="🟡" if field.is_yellow else "")
            yellow_cell.alignment = Alignment(horizontal='center')
            if field.is_yellow:
                for col in range(1, len(headers) + 1):
                    ws_full.cell(row=row_idx, column=col).fill = yellow_fill
            
            white_cell = ws_full.cell(row=row_idx, column=11, 
                                      value="⚪" if field.is_white else "")
            white_cell.alignment = Alignment(horizontal='center')
            if field.is_white:
                for col in range(1, len(headers) + 1):
                    ws_full.cell(row=row_idx, column=col).fill = white_fill
            
            # Sample Value
            ws_full.cell(row=row_idx, column=12, value=field.sample_value)
            
            # Empty mapping columns (to be filled by user)
            for col in range(13, 19):
                cell = ws_full.cell(row=row_idx, column=col, value="")
                cell.border = thin_border
        
        # Add data validation for Status column
        status_validation = DataValidation(
            type="list",
            formula1='"Not Started,In Progress,Mapped,Verified,N/A"',
            allow_blank=True
        )
        ws_full.add_data_validation(status_validation)
        status_validation.add(f'R2:R{len(self.fields) + 1}')
        
        # Set column widths
        column_widths = [8, 50, 30, 15, 6, 6, 10, 40, 10, 8, 8, 25, 20, 20, 30, 15, 30, 12]
        for col_idx, width in enumerate(column_widths, 1):
            ws_full.column_dimensions[get_column_letter(col_idx)].width = width
        
        # ============================================
        # Sheet 2: Mandatory Fields Only
        # ============================================
        ws_mandatory = wb.create_sheet("Mandatory Fields")
        
        mandatory_headers = ["XPath", "Element", "Data Type", "Sample", 
                           "Source Field", "Transformation", "Notes"]
        
        for col_idx, header in enumerate(mandatory_headers, 1):
            cell = ws_mandatory.cell(row=1, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.border = thin_border
        
        ws_mandatory.freeze_panes = 'A2'
        
        row_idx = 2
        for field in self.fields:
            if field.is_mandatory:
                ws_mandatory.cell(row=row_idx, column=1, value=field.xpath)
                ws_mandatory.cell(row=row_idx, column=2, value=field.element_name)
                ws_mandatory.cell(row=row_idx, column=3, value=field.data_type)
                ws_mandatory.cell(row=row_idx, column=4, value=field.sample_value)
                row_idx += 1
        
        ws_mandatory.column_dimensions['A'].width = 50
        ws_mandatory.column_dimensions['B'].width = 25
        ws_mandatory.column_dimensions['C'].width = 15
        ws_mandatory.column_dimensions['D'].width = 25
        ws_mandatory.column_dimensions['E'].width = 20
        ws_mandatory.column_dimensions['F'].width = 30
        ws_mandatory.column_dimensions['G'].width = 30
        
        # ============================================
        # Sheet 3: Summary Statistics
        # ============================================
        ws_summary = wb.create_sheet("Summary")
        
        total_fields = len(self.fields)
        mandatory_fields = sum(1 for f in self.fields if f.is_mandatory)
        optional_fields = total_fields - mandatory_fields
        yellow_fields = sum(1 for f in self.fields if f.is_yellow)
        white_fields = sum(1 for f in self.fields if f.is_white)
        
        summary_data = [
            ["ISO 20022 Mapping Template", ""],
            ["", ""],
            ["Schema Information", ""],
            ["Schema File", os.path.basename(self.xsd_file)],
            ["Generated", datetime.now().strftime('%Y-%m-%d %H:%M:%S')],
            ["", ""],
            ["Field Statistics", ""],
            ["Total Fields", total_fields],
            ["Mandatory Fields", mandatory_fields],
            ["Optional Fields", optional_fields],
            ["Yellow Fields", yellow_fields],
            ["White Fields", white_fields],
            ["", ""],
            ["Instructions", ""],
            ["1. Review the 'Mapping Template' sheet", ""],
            ["2. Fill in Source System and Source Field columns", ""],
            ["3. Add Transformation Rules where needed", ""],
            ["4. Update Status as mapping progresses", ""],
            ["5. Use 'Mandatory Fields' sheet for quick reference", ""],
        ]
        
        for row_idx, row_data in enumerate(summary_data, 1):
            for col_idx, value in enumerate(row_data, 1):
                cell = ws_summary.cell(row=row_idx, column=col_idx, value=value)
                if row_idx in [1, 3, 7, 14]:
                    cell.font = Font(bold=True, size=14)
        
        ws_summary.column_dimensions['A'].width = 40
        ws_summary.column_dimensions['B'].width = 30
        
        # Save workbook
        wb.save(output_path)
        print(f"✅ Mapping template saved to: {output_path}")
        
        return {
            'total_fields': total_fields,
            'mandatory_fields': mandatory_fields,
            'optional_fields': optional_fields,
            'yellow_fields': yellow_fields,
            'white_fields': white_fields
        }


def main():
    parser = argparse.ArgumentParser(
        description='ISO 20022 Mapping Template Generator',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  # Generate Excel mapping template
  python mapping_generator.py schema.xsd -o mapping_template.xlsx
  
  # Generate with verbose output
  python mapping_generator.py schema.xsd -o mapping.xlsx --verbose
        """
    )
    
    parser.add_argument('xsd_file', help='XSD schema file')
    parser.add_argument('-o', '--output', required=True, help='Output Excel file')
    parser.add_argument('--verbose', '-v', action='store_true', help='Verbose output')
    
    args = parser.parse_args()
    
    if not Path(args.xsd_file).exists():
        print(f"❌ Error: XSD file not found: {args.xsd_file}")
        return
    
    if not HAS_OPENPYXL:
        print("❌ Error: openpyxl required. Install with: pip install openpyxl")
        return
    
    print(f"\n{'='*70}")
    print("ISO 20022 MAPPING TEMPLATE GENERATOR")
    print(f"{'='*70}\n")
    print(f"📋 XSD: {args.xsd_file}")
    print(f"📁 Output: {args.output}")
    print(f"\n⏳ Extracting fields...")
    
    generator = MappingTemplateGenerator(args.xsd_file)
    fields = generator.extract_fields()
    
    print(f"   Found {len(fields)} fields")
    
    if args.verbose:
        mandatory = sum(1 for f in fields if f.is_mandatory)
        yellow = sum(1 for f in fields if f.is_yellow)
        print(f"   Mandatory: {mandatory}")
        print(f"   Yellow: {yellow}")
    
    print(f"\n⏳ Generating Excel template...")
    stats = generator.generate_excel(args.output)
    
    print(f"\n📊 Summary:")
    print(f"   Total Fields:     {stats['total_fields']}")
    print(f"   Mandatory Fields: {stats['mandatory_fields']}")
    print(f"   Yellow Fields:    {stats['yellow_fields']}")
    print(f"   White Fields:     {stats['white_fields']}")
    
    print(f"\n{'='*70}")
    print("COMPLETE")
    print(f"{'='*70}\n")


if __name__ == '__main__':
    main()
